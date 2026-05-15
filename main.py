# -*- coding: utf-8 -*-
import gc, os, sys, threading, json, csv, smtplib, imaplib, email, sqlite3
import traceback
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict
from email.header import decode_header
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from typing import Optional

from kivy.app import App
from kivy.lang import Builder
from kivy.clock import Clock
from kivy.uix.screenmanager import ScreenManager, SlideTransition, Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.spinner import Spinner
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.uix.widget import Widget
from kivy.animation import Animation
from kivy.metrics import dp
from kivy.utils import get_color_from_hex
from kivy.core.window import Window
from kivy.graphics import Color, RoundedRectangle
from kivy.graphics import Color as GColor, Line as GLine
from kivy.utils import platform

# KV: ajustes_importar.kv
Builder.load_string('#:import utils kivy.utils\n#:import dp kivy.metrics.dp\n\n<AjustesScreen>:\n    BoxLayout:\n        orientation: \'vertical\'\n        padding: dp(15)\n        spacing: dp(10)\n        canvas.before:\n            Color:\n                rgba: utils.get_color_from_hex(\'#0A0A12\')\n            Rectangle:\n                pos: self.pos\n                size: self.size\n\n        BoxLayout:\n            size_hint_y: None\n            height: dp(50)\n            padding: [dp(5), 0]\n            Label:\n                text: "AJUSTES"\n                font_size: \'18sp\'\n                bold: True\n                halign: \'left\'\n                text_size: self.size\n                valign: \'middle\'\n                color: utils.get_color_from_hex(\'#5A6A7A\')\n\n        ScrollView:\n            BoxLayout:\n                orientation: \'vertical\'\n                size_hint_y: None\n                height: self.minimum_height\n                spacing: dp(12)\n                padding: dp(5)\n\n                BoxLayout:\n                    orientation: \'vertical\'\n                    size_hint_y: None\n                    height: dp(180)\n                    spacing: dp(8)\n                    padding: dp(12)\n                    canvas.before:\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#0D1B2A\')\n                        RoundedRectangle:\n                            pos: self.pos\n                            size: self.size\n                            radius: [dp(4)]\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#1A2A4A\')\n                        Line:\n                            rounded_rectangle: (self.x, self.y, self.width, self.height, dp(4))\n                            width: dp(1)\n                    Label:\n                        text: "ALERTAS DE STOCK MINIMO"\n                        bold: True\n                        color: utils.get_color_from_hex(\'#FF0044\')\n                        size_hint_y: None\n                        height: dp(30)\n                    Spinner:\n                        id: alerta_material_spinner\n                        text: "SELECCIONAR MATERIAL"\n                        values: []\n                        size_hint_y: None\n                        height: dp(45)\n                    BoxLayout:\n                        spacing: dp(8)\n                        size_hint_y: None\n                        height: dp(45)\n                        TextInput:\n                            id: stock_minimo_input\n                            hint_text: "Stock minimo (ej: 5)"\n                            input_filter: \'float\'\n                            multiline: False\n                        Button:\n                            text: "GUARDAR"\n                            size_hint_x: None\n                            width: dp(100)\n                            background_color: utils.get_color_from_hex(\'#0066FF\')\n                            on_release: root.guardar_alerta()\n                    Label:\n                        id: info_alerta\n                        text: ""\n                        color: utils.get_color_from_hex(\'#00D4FF\')\n                        size_hint_y: None\n                        height: dp(22)\n\n                BoxLayout:\n                    orientation: \'vertical\'\n                    size_hint_y: None\n                    height: dp(230)\n                    spacing: dp(8)\n                    padding: dp(12)\n                    canvas.before:\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#0D1B2A\')\n                        RoundedRectangle:\n                            pos: self.pos\n                            size: self.size\n                            radius: [dp(4)]\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#1A2A4A\')\n                        Line:\n                            rounded_rectangle: (self.x, self.y, self.width, self.height, dp(4))\n                            width: dp(1)\n                    Label:\n                        text: "ACTUALIZACION DESDE CORREO"\n                        bold: True\n                        color: utils.get_color_from_hex(\'#00D4FF\')\n                        size_hint_y: None\n                        height: dp(30)\n                        halign: \'left\'\n                        text_size: self.size\n                    Label:\n                        text: "Asunto a buscar (vacio = cualquiera):"\n                        font_size: \'11sp\'\n                        color: utils.get_color_from_hex(\'#5A6A7A\')\n                        size_hint_y: None\n                        height: dp(20)\n                        halign: \'left\'\n                        text_size: self.size\n                    TextInput:\n                        id: filtro_asunto\n                        hint_text: "ej: INVENTARIO  (vacio = sin filtro)"\n                        multiline: False\n                        size_hint_y: None\n                        height: dp(40)\n                        font_size: \'13sp\'\n                    Label:\n                        text: "Remitente a buscar (vacio = cualquiera):"\n                        font_size: \'11sp\'\n                        color: utils.get_color_from_hex(\'#5A6A7A\')\n                        size_hint_y: None\n                        height: dp(20)\n                        halign: \'left\'\n                        text_size: self.size\n                    TextInput:\n                        id: filtro_remitente\n                        hint_text: "ej: jefe@empresa.com  (vacio = sin filtro)"\n                        multiline: False\n                        size_hint_y: None\n                        height: dp(40)\n                        font_size: \'13sp\'\n                    Button:\n                        text: "GUARDAR FILTROS"\n                        size_hint_y: None\n                        height: dp(40)\n                        background_color: utils.get_color_from_hex(\'#0066FF\')\n                        on_release: root.guardar_filtros_correo()\n\n                BoxLayout:\n                    orientation: \'vertical\'\n                    size_hint_y: None\n                    height: dp(145)\n                    spacing: dp(8)\n                    padding: dp(12)\n                    canvas.before:\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#0D1B2A\')\n                        RoundedRectangle:\n                            pos: self.pos\n                            size: self.size\n                            radius: [dp(4)]\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#1A2A4A\')\n                        Line:\n                            rounded_rectangle: (self.x, self.y, self.width, self.height, dp(4))\n                            width: dp(1)\n                    Label:\n                        text: "EXPORTAR DATOS"\n                        bold: True\n                        color: utils.get_color_from_hex(\'#00D4FF\')\n                        size_hint_y: None\n                        height: dp(30)\n                    Label:\n                        id: lbl_carpeta_export\n                        text: ""\n                        font_size: \'11sp\'\n                        color: utils.get_color_from_hex(\'#5A6A7A\')\n                        size_hint_y: None\n                        height: dp(28)\n                        halign: \'left\'\n                        text_size: self.size\n                        shorten: True\n                        shorten_from: \'left\'\n                    BoxLayout:\n                        spacing: dp(8)\n                        size_hint_y: None\n                        height: dp(50)\n                        Button:\n                            text: "CSV"\n                            background_color: utils.get_color_from_hex(\'#0066FF\')\n                            on_release: root.exportar_csv()\n                        Button:\n                            text: "EXCEL"\n                            background_color: utils.get_color_from_hex(\'#0066FF\')\n                            on_release: root.exportar_excel_local()\n\n                BoxLayout:\n                    orientation: \'vertical\'\n                    size_hint_y: None\n                    height: dp(120)\n                    spacing: dp(8)\n                    padding: dp(12)\n                    canvas.before:\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#0D1B2A\')\n                        RoundedRectangle:\n                            pos: self.pos\n                            size: self.size\n                            radius: [dp(4)]\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#1A2A4A\')\n                        Line:\n                            rounded_rectangle: (self.x, self.y, self.width, self.height, dp(4))\n                            width: dp(1)\n                    Label:\n                        text: "BACKUP AUTOMATICO"\n                        bold: True\n                        color: utils.get_color_from_hex(\'#00D4FF\')\n                        size_hint_y: None\n                        height: dp(28)\n                    BoxLayout:\n                        spacing: dp(8)\n                        size_hint_y: None\n                        height: dp(45)\n                        TextInput:\n                            id: backup_interval\n                            hint_text: "Cada cuantas horas? (0=desactivado)"\n                            input_filter: \'int\'\n                            multiline: False\n                            font_size: \'13sp\'\n                        Button:\n                            text: "GUARDAR"\n                            size_hint_x: None\n                            width: dp(100)\n                            background_color: utils.get_color_from_hex(\'#0066FF\')\n                            on_release: root.guardar_backup_config()\n                    Label:\n                        text: "Se genera un Excel automaticamente en la carpeta de exportacion"\n                        font_size: \'10sp\'\n                        color: utils.get_color_from_hex(\'#5A6A7A\')\n                        size_hint_y: None\n                        height: dp(34)\n                        text_size: self.width, None\n                        halign: \'left\'\n                        valign: \'top\'\n\n                BoxLayout:\n                    orientation: \'vertical\'\n                    size_hint_y: None\n                    height: dp(100)\n                    spacing: dp(8)\n                    padding: dp(12)\n                    canvas.before:\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#0D1B2A\')\n                        RoundedRectangle:\n                            pos: self.pos\n                            size: self.size\n                            radius: [dp(4)]\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#1A2A4A\')\n                        Line:\n                            rounded_rectangle: (self.x, self.y, self.width, self.height, dp(4))\n                            width: dp(1)\n                    Label:\n                        text: "ZONA DE PELIGRO"\n                        bold: True\n                        color: utils.get_color_from_hex(\'#FF0044\')\n                        size_hint_y: None\n                        height: dp(28)\n                    Button:\n                        text: "PURGAR REGISTROS ANTIGUOS (+90 dias)"\n                        size_hint_y: None\n                        height: dp(45)\n                        background_color: utils.get_color_from_hex(\'#FF0044\')\n                        on_release: root.confirmar_purga()\n        \n        Label:\n            id: info_ajustes\n            text: ""\n            color: utils.get_color_from_hex(\'#00D4FF\')\n            size_hint_y: None\n            height: dp(28)\n\n        Button:\n            text: "VOLVER AL INICIO"\n            size_hint_y: None\n            height: dp(50)\n            font_size: \'14sp\'\n            bold: True\n            background_color: utils.get_color_from_hex(\'#334155\')\n            on_release: root.manager.current = \'inventario\'\n\n<ImportarExcelScreen>:\n    BoxLayout:\n        orientation: \'vertical\'\n        padding: dp(12)\n        spacing: dp(8)\n        canvas.before:\n            Color:\n                rgba: utils.get_color_from_hex(\'#0A0A12\')\n            Rectangle:\n                pos: self.pos\n                size: self.size\n\n        Label:\n            text: "IMPORTAR EXCEL AL INVENTARIO"\n            bold: True\n            font_size: \'16sp\'\n            color: utils.get_color_from_hex(\'#00D4FF\')\n            size_hint_y: None\n            height: dp(36)\n\n        BoxLayout:\n            orientation: \'vertical\'\n            size_hint_y: None\n            height: dp(175)\n            spacing: dp(6)\n            padding: dp(10)\n            canvas.before:\n                Color:\n                    rgba: utils.get_color_from_hex(\'#0D1B2A\')\n                RoundedRectangle:\n                    pos: self.pos\n                    size: self.size\n                    radius: [dp(4)]\n                Color:\n                    rgba: utils.get_color_from_hex(\'#1A2A4A\')\n                Line:\n                    rounded_rectangle: (self.x, self.y, self.width, self.height, dp(4))\n                    width: dp(1)\n\n            Label:\n                text: "SELECCIONA O PEGA LA RUTA DEL ARCHIVO"\n                bold: True\n                font_size: \'12sp\'\n                color: utils.get_color_from_hex(\'#FF0044\')\n                size_hint_y: None\n                height: dp(22)\n                halign: \'left\'\n                text_size: self.size\n\n            Label:\n                id: lbl_hint_carpeta\n                text: ""\n                font_size: \'10sp\'\n                color: utils.get_color_from_hex(\'#1A2A4A\')\n                size_hint_y: None\n                height: dp(18)\n                halign: \'left\'\n                text_size: self.size\n                shorten: True\n                shorten_from: \'left\'\n\n            TextInput:\n                id: ruta_excel\n                hint_text: "La ruta aparece aqui al explorar o pegar"\n                multiline: False\n                size_hint_y: None\n                height: dp(42)\n                font_size: \'12sp\'\n\n            BoxLayout:\n                size_hint_y: None\n                height: dp(44)\n                spacing: dp(6)\n                Button:\n                    text: "EXPLORAR"\n                    font_size: \'13sp\'\n                    bold: True\n                    background_color: utils.get_color_from_hex(\'#0066FF\')\n                    on_release: root.abrir_explorador()\n                Button:\n                    text: "PEGAR"\n                    font_size: \'13sp\'\n                    background_color: utils.get_color_from_hex(\'#1A2A4A\')\n                    on_release: root.pegar_ruta()\n                Button:\n                    text: "PREVISUALIZAR"\n                    font_size: \'13sp\'\n                    bold: True\n                    background_color: utils.get_color_from_hex(\'#0066FF\')\n                    on_release: root.previsualizar()\n\n        BoxLayout:\n            orientation: \'vertical\'\n            size_hint_y: None\n            height: dp(188)\n            spacing: dp(5)\n            padding: dp(10)\n            canvas.before:\n                Color:\n                    rgba: utils.get_color_from_hex(\'#0D1B2A\')\n                RoundedRectangle:\n                    pos: self.pos\n                    size: self.size\n                    radius: [dp(4)]\n                Color:\n                    rgba: utils.get_color_from_hex(\'#1A2A4A\')\n                Line:\n                    rounded_rectangle: (self.x, self.y, self.width, self.height, dp(4))\n                    width: dp(1)\n\n            Label:\n                text: "COLUMNAS DEL EXCEL (se auto-detectan al previsualizar)"\n                bold: True\n                font_size: \'12sp\'\n                color: utils.get_color_from_hex(\'#FF0044\')\n                size_hint_y: None\n                height: dp(22)\n                halign: \'left\'\n                text_size: self.size\n\n            BoxLayout:\n                size_hint_y: None\n                height: dp(36)\n                spacing: dp(6)\n                Label:\n                    text: "Responsable"\n                    font_size: \'11sp\'\n                    size_hint_x: 0.28\n                    color: utils.get_color_from_hex(\'#5A6A7A\')\n                    halign: \'right\'\n                    text_size: self.size\n                    valign: \'middle\'\n                TextInput:\n                    id: col_nombre\n                    text: "Nombre"\n                    multiline: False\n                    font_size: \'12sp\'\n                Label:\n                    text: "Material"\n                    font_size: \'11sp\'\n                    size_hint_x: 0.22\n                    color: utils.get_color_from_hex(\'#5A6A7A\')\n                    halign: \'right\'\n                    text_size: self.size\n                    valign: \'middle\'\n                TextInput:\n                    id: col_material\n                    text: "Material"\n                    multiline: False\n                    font_size: \'12sp\'\n\n            BoxLayout:\n                size_hint_y: None\n                height: dp(36)\n                spacing: dp(6)\n                Label:\n                    text: "SKU"\n                    font_size: \'11sp\'\n                    size_hint_x: 0.28\n                    color: utils.get_color_from_hex(\'#5A6A7A\')\n                    halign: \'right\'\n                    text_size: self.size\n                    valign: \'middle\'\n                TextInput:\n                    id: col_sku\n                    text: "SKU"\n                    multiline: False\n                    font_size: \'12sp\'\n                Label:\n                    text: "Cantidad"\n                    font_size: \'11sp\'\n                    size_hint_x: 0.22\n                    color: utils.get_color_from_hex(\'#5A6A7A\')\n                    halign: \'right\'\n                    text_size: self.size\n                    valign: \'middle\'\n                TextInput:\n                    id: col_cantidad\n                    text: "Cantidad"\n                    multiline: False\n                    font_size: \'12sp\'\n\n            BoxLayout:\n                size_hint_y: None\n                height: dp(36)\n                spacing: dp(6)\n                Label:\n                    text: "Tipo"\n                    font_size: \'11sp\'\n                    size_hint_x: 0.28\n                    color: utils.get_color_from_hex(\'#5A6A7A\')\n                    halign: \'right\'\n                    text_size: self.size\n                    valign: \'middle\'\n                TextInput:\n                    id: col_tipo\n                    text: "Tipo"\n                    multiline: False\n                    font_size: \'12sp\'\n                Label:\n                    text: "Fecha"\n                    font_size: \'11sp\'\n                    size_hint_x: 0.22\n                    color: utils.get_color_from_hex(\'#5A6A7A\')\n                    halign: \'right\'\n                    text_size: self.size\n                    valign: \'middle\'\n                TextInput:\n                    id: col_fecha\n                    text: "Fecha"\n                    multiline: False\n                    font_size: \'12sp\'\n\n            BoxLayout:\n                size_hint_y: None\n                height: dp(36)\n                spacing: dp(6)\n                Label:\n                    text: "Hoja Excel"\n                    font_size: \'11sp\'\n                    size_hint_x: 0.28\n                    color: utils.get_color_from_hex(\'#5A6A7A\')\n                    halign: \'right\'\n                    text_size: self.size\n                    valign: \'middle\'\n                Spinner:\n                    id: hoja_spinner\n                    text: "-- detectar --"\n                    values: ["-- detectar --"]\n                    font_size: \'12sp\'\n\n        Label:\n            id: estado_importar\n            text: "Pulsa EXPLORAR para buscar tu archivo Excel"\n            color: utils.get_color_from_hex(\'#5A6A7A\')\n            markup: True\n            size_hint_y: None\n            height: dp(36)\n            halign: \'left\'\n            text_size: self.size\n            font_size: \'12sp\'\n\n        ScrollView:\n            size_hint_y: 1\n            do_scroll_x: False\n            GridLayout:\n                id: preview_container\n                cols: 1\n                size_hint_y: None\n                height: self.minimum_height\n                spacing: dp(5)\n\n        Button:\n            id: btn_importar\n            text: "IMPORTAR AL INVENTARIO"\n            bold: True\n            font_size: \'15sp\'\n            size_hint_y: None\n            height: dp(52)\n            background_color: utils.get_color_from_hex(\'#0066FF\')\n            disabled: True\n            on_release: root.importar()\n\n        Button:\n            text: "VOLVER"\n            size_hint_y: None\n            height: dp(46)\n            background_color: utils.get_color_from_hex(\'#334155\')\n            on_release: root.manager.current = \'inventario\'\n')

# KV: alertas.kv
Builder.load_string('#:import utils kivy.utils\n#:import dp kivy.metrics.dp\n\n<AlertasScreen>:\n    BoxLayout:\n        orientation: \'vertical\'\n        padding: dp(15)\n        spacing: dp(10)\n        canvas.before:\n            Color:\n                rgba: utils.get_color_from_hex(\'#0A0A12\')\n            Rectangle:\n                pos: self.pos\n                size: self.size\n\n        BoxLayout:\n            size_hint_y: None\n            height: dp(50)\n            spacing: dp(8)\n            Label:\n                id: total_alertas\n                text: "Alertas activas: 0"\n                font_size: \'16sp\'\n                bold: True\n                color: utils.get_color_from_hex(\'#FF0044\')\n                halign: \'left\'\n                valign: \'middle\'\n                text_size: self.size\n            Button:\n                text: "VOLVER"\n                size_hint_x: None\n                width: dp(90)\n                font_size: \'12sp\'\n                background_color: utils.get_color_from_hex(\'#334155\')\n                on_release: root.manager.current = \'inventario\'\n\n        ScrollView:\n            do_scroll_x: False\n            GridLayout:\n                id: container_alertas\n                cols: 1\n                size_hint_y: None\n                height: self.minimum_height\n                spacing: dp(8)\n                padding: [0, dp(5)]\n')

# KV: auditoria.kv
Builder.load_string('#:import utils kivy.utils\n#:import dp kivy.metrics.dp\n\n<AuditoriaScreen>:\n    BoxLayout:\n        orientation: \'vertical\'\n        padding: dp(15)\n        spacing: dp(8)\n        canvas.before:\n            Color:\n                rgba: utils.get_color_from_hex(\'#0A0A12\')\n            Rectangle:\n                pos: self.pos\n                size: self.size\n\n        BoxLayout:\n            size_hint_y: None\n            height: dp(50)\n            spacing: dp(8)\n            Label:\n                text: "AUDITORIA"\n                font_size: \'16sp\'\n                bold: True\n                color: utils.get_color_from_hex(\'#00D4FF\')\n                halign: \'left\'\n                valign: \'middle\'\n                text_size: self.size\n            Button:\n                text: "VOLVER"\n                size_hint_x: None\n                width: dp(90)\n                font_size: \'12sp\'\n                background_color: utils.get_color_from_hex(\'#334155\')\n                on_release: root.manager.current = \'inventario\'\n\n        BoxLayout:\n            size_hint_y: None\n            height: dp(42)\n            spacing: dp(6)\n            TextInput:\n                id: filtro_accion\n                hint_text: "Filtrar por accion..."\n                multiline: False\n                font_size: \'12sp\'\n            TextInput:\n                id: filtro_usuario\n                hint_text: "Filtrar por usuario..."\n                multiline: False\n                font_size: \'12sp\'\n            Button:\n                text: "FILTRAR"\n                size_hint_x: None\n                width: dp(90)\n                font_size: \'12sp\'\n                background_color: utils.get_color_from_hex(\'#0066FF\')\n                on_release: root.filtrar()\n\n        PaginationBar:\n            id: pagination\n            callback: root.cargar_log\n            page_size: 30\n\n        ScrollView:\n            GridLayout:\n                id: log_container\n                cols: 1\n                size_hint_y: None\n                height: self.minimum_height\n                spacing: dp(6)\n                padding: [0, dp(4)]\n')

# KV: categorias.kv
Builder.load_string('#:import utils kivy.utils\n#:import dp kivy.metrics.dp\n\n<CategoriasScreen>:\n    BoxLayout:\n        orientation: \'vertical\'\n        padding: dp(15)\n        spacing: dp(10)\n        canvas.before:\n            Color:\n                rgba: utils.get_color_from_hex(\'#0A0A12\')\n            Rectangle:\n                pos: self.pos\n                size: self.size\n\n        BoxLayout:\n            size_hint_y: None\n            height: dp(50)\n            spacing: dp(8)\n            Label:\n                text: "CATEGORIAS"\n                font_size: \'16sp\'\n                bold: True\n                color: utils.get_color_from_hex(\'#00D4FF\')\n                halign: \'left\'\n                valign: \'middle\'\n                text_size: self.size\n            Button:\n                text: "+ NUEVA"\n                size_hint_x: None\n                width: dp(100)\n                font_size: \'13sp\'\n                bold: True\n                background_color: utils.get_color_from_hex(\'#0066FF\')\n                on_release: root.mostrar_nueva()\n            Button:\n                text: "VOLVER"\n                size_hint_x: None\n                width: dp(90)\n                font_size: \'12sp\'\n                background_color: utils.get_color_from_hex(\'#334155\')\n                on_release: root.manager.current = \'inventario\'\n\n        ScrollView:\n            GridLayout:\n                id: lista_categorias\n                cols: 1\n                size_hint_y: None\n                height: self.minimum_height\n                spacing: dp(6)\n')

# KV: gestion.kv
Builder.load_string('#:import utils kivy.utils\n#:import dp kivy.metrics.dp\n\n<GestionScreen>:\n    BoxLayout:\n        orientation: \'vertical\'\n        padding: dp(15)\n        spacing: dp(10)\n        canvas.before:\n            Color:\n                rgba: utils.get_color_from_hex(\'#0A0A12\')\n            Rectangle:\n                pos: self.pos\n                size: self.size\n\n        BoxLayout:\n            size_hint_y: None\n            height: dp(50)\n            spacing: dp(8)\n            Label:\n                id: titulo_seccion\n                text: "MATERIALES"\n                font_size: \'16sp\'\n                bold: True\n                color: utils.get_color_from_hex(\'#00D4FF\')\n                halign: \'left\'\n                valign: \'middle\'\n                text_size: self.size\n            Button:\n                text: "MATERIALES"\n                size_hint_x: None\n                width: dp(100)\n                font_size: \'11sp\'\n                background_color: utils.get_color_from_hex(\'#0066FF\')\n                on_release: root.cambiar_modo("materiales")\n            Button:\n                text: "RESPONSABLES"\n                size_hint_x: None\n                width: dp(110)\n                font_size: \'11sp\'\n                background_color: utils.get_color_from_hex(\'#0066FF\')\n                on_release: root.cambiar_modo("responsables")\n            Button:\n                text: "VOLVER"\n                size_hint_x: None\n                width: dp(80)\n                font_size: \'12sp\'\n                background_color: utils.get_color_from_hex(\'#334155\')\n                on_release: root.manager.current = \'inventario\'\n\n        BoxLayout:\n            size_hint_y: None\n            height: dp(42)\n            spacing: dp(6)\n            TextInput:\n                id: filtro_gestion\n                hint_text: "Filtrar..."\n                multiline: False\n                font_size: \'12sp\'\n            Button:\n                text: "BUSCAR"\n                size_hint_x: None\n                width: dp(80)\n                font_size: \'12sp\'\n                background_color: utils.get_color_from_hex(\'#0066FF\')\n                on_release: root.filtrar()\n\n        ScrollView:\n            GridLayout:\n                id: lista_gestion\n                cols: 1\n                size_hint_y: None\n                height: self.minimum_height\n                spacing: dp(6)\n')

# KV: limpieza.kv
Builder.load_string('#:import utils kivy.utils\n#:import dp kivy.metrics.dp\n\n<LimpiezaScreen>:\n    BoxLayout:\n        orientation: \'vertical\'\n        padding: dp(15)\n        spacing: dp(10)\n        canvas.before:\n            Color:\n                rgba: utils.get_color_from_hex(\'#0A0A12\')\n            Rectangle:\n                pos: self.pos\n                size: self.size\n\n        BoxLayout:\n            size_hint_y: None\n            height: dp(50)\n            padding: [dp(5), 0]\n            Label:\n                text: "CONTROL DE LIMPIEZA"\n                font_size: \'18sp\'\n                bold: True\n                halign: \'left\'\n                text_size: self.size\n                valign: \'middle\'\n                color: utils.get_color_from_hex(\'#00D4FF\')\n\n        BoxLayout:\n            orientation: \'vertical\'\n            size_hint_y: None\n            height: dp(310)\n            spacing: dp(8)\n            padding: dp(15)\n            canvas.before:\n                Color:\n                    rgba: utils.get_color_from_hex(\'#0D1B2A\')\n                RoundedRectangle:\n                    pos: self.pos\n                    size: self.size\n                    radius: [dp(4)]\n                Color:\n                    rgba: utils.get_color_from_hex(\'#00D4FF\')\n                Line:\n                    rounded_rectangle: [self.pos[0], self.pos[1], self.size[0], self.size[1], dp(4)]\n\n            Label:\n                text: "REGISTRAR ESTADO"\n                bold: True\n                font_size: \'12sp\'\n                color: utils.get_color_from_hex(\'#00D4FF\')\n                size_hint_y: None\n                height: dp(20)\n                text_size: self.size\n                halign: \'left\'\n\n            Spinner:\n                id: limpieza_spinner\n                text: "SELECCIONAR MATERIAL"\n                values: []\n                size_hint_y: None\n                height: dp(45)\n            TextInput:\n                id: m_n\n                hint_text: "O escribe material..."\n                multiline: False\n                size_hint_y: None\n                height: dp(45)\n            TextInput:\n                id: m_c\n                hint_text: "Cantidad"\n                input_filter: \'float\'\n                multiline: False\n                size_hint_y: None\n                height: dp(45)\n            TextInput:\n                id: m_notas\n                hint_text: "Notas..."\n                multiline: False\n                size_hint_y: None\n                height: dp(45)\n            BoxLayout:\n                spacing: dp(10)\n                size_hint_y: None\n                height: dp(50)\n                Button:\n                    text: "LIMPIO"\n                    bold: True\n                    background_color: utils.get_color_from_hex(\'#0066FF\')\n                    on_release: root.registrar_limpieza("LIMPIO")\n                Button:\n                    text: "SUCIO"\n                    bold: True\n                    background_color: utils.get_color_from_hex(\'#FF0044\')\n                    on_release: root.registrar_limpieza("SUCIO")\n\n        Label:\n            id: info_limpieza\n            text: ""\n            color: utils.get_color_from_hex(\'#FF0044\')\n            size_hint_y: None\n            height: dp(22)\n\n        TextInput:\n            id: f_l\n            hint_text: "Filtrar historial..."\n            multiline: False\n            size_hint_y: None\n            height: dp(45)\n            on_text: root.consultar_historial()\n\n        ScrollView:\n            GridLayout:\n                id: hist_limpieza\n                cols: 1\n                size_hint_y: None\n                height: self.minimum_height\n                spacing: dp(10)\n\n        Button:\n            text: "VOLVER AL INICIO"\n            size_hint_y: None\n            height: dp(50)\n            font_size: \'14sp\'\n            bold: True\n            background_color: utils.get_color_from_hex(\'#334155\')\n            on_release: root.manager.current = \'inventario\'\n')

# KV: pendientes_filtros.kv
Builder.load_string('#:import utils kivy.utils\n#:import dp kivy.metrics.dp\n\n<PendientesScreen>:\n    BoxLayout:\n        orientation: \'vertical\'\n        padding: dp(15)\n        spacing: dp(10)\n        canvas.before:\n            Color:\n                rgba: utils.get_color_from_hex(\'#0A0A12\')\n            Rectangle:\n                pos: self.pos\n                size: self.size\n\n        BoxLayout:\n            size_hint_y: None\n            height: dp(50)\n            spacing: dp(8)\n            padding: [dp(5), 0]\n            Label:\n                text: "MATERIALES FUERA"\n                font_size: \'18sp\'\n                bold: True\n                halign: \'left\'\n                text_size: self.size\n                valign: \'middle\'\n                color: utils.get_color_from_hex(\'#FF0044\')\n            Button:\n                text: "SYNC"\n                size_hint_x: None\n                width: dp(45)\n                background_color: utils.get_color_from_hex(\'#334155\')\n                on_release: root.cargar_pendientes()\n\n        TextInput:\n            id: filtro_pendientes\n            hint_text: "Filtrar por responsable o material..."\n            multiline: False\n            size_hint_y: None\n            height: dp(45)\n            on_text: root.cargar_pendientes()\n\n        ScrollView:\n            GridLayout:\n                id: lista_pendientes\n                cols: 1\n                size_hint_y: None\n                height: self.minimum_height\n                spacing: dp(12)\n\n        Button:\n            text: "VOLVER AL INICIO"\n            size_hint_y: None\n            height: dp(50)\n            font_size: \'14sp\'\n            bold: True\n            background_color: utils.get_color_from_hex(\'#334155\')\n            on_release: root.manager.current = \'inventario\'\n\n<FiltroScreen>:\n    BoxLayout:\n        orientation: \'vertical\'\n        padding: dp(15)\n        spacing: dp(10)\n        canvas.before:\n            Color:\n                rgba: utils.get_color_from_hex(\'#0A0A12\')\n            Rectangle:\n                pos: self.pos\n                size: self.size\n\n        BoxLayout:\n            size_hint_y: None\n            height: dp(50)\n            padding: [dp(5), 0]\n            Label:\n                text: "AUDITORIA DE MOVIMIENTOS"\n                font_size: \'18sp\'\n                bold: True\n                halign: \'left\'\n                text_size: self.size\n                valign: \'middle\'\n                color: utils.get_color_from_hex(\'#00D4FF\')\n\n        BoxLayout:\n            orientation: \'vertical\'\n            size_hint_y: None\n            height: dp(175)\n            spacing: dp(6)\n            padding: dp(15)\n            canvas.before:\n                Color:\n                    rgba: utils.get_color_from_hex(\'#0D1B2A\')\n                RoundedRectangle:\n                    pos: self.pos\n                    size: self.size\n                    radius: [dp(4)]\n\n            Label:\n                text: "FILTROS"\n                bold: True\n                font_size: \'12sp\'\n                color: utils.get_color_from_hex(\'#00D4FF\')\n                size_hint_y: None\n                height: dp(20)\n                text_size: self.size\n                halign: \'left\'\n\n            TextInput:\n                id: f_n\n                hint_text: "Filtrar por Responsable..."\n                multiline: False\n                size_hint_y: None\n                height: dp(42)\n                on_text: root.filtrar()\n            TextInput:\n                id: f_p\n                hint_text: "Filtrar por Material..."\n                multiline: False\n                size_hint_y: None\n                height: dp(42)\n                on_text: root.filtrar()\n            Spinner:\n                id: filtro_tipo\n                text: "TODOS LOS TIPOS"\n                values: ["TODOS LOS TIPOS", "ENTRADA", "SALIDA"]\n                size_hint_y: None\n                height: dp(42)\n                on_text: root.filtrar()\n\n        Label:\n            id: filtro_count\n            text: ""\n            color: utils.get_color_from_hex(\'#5A6A7A\')\n            size_hint_y: None\n            height: dp(25)\n\n        ScrollView:\n            GridLayout:\n                id: c_r\n                cols: 1\n                size_hint_y: None\n                height: self.minimum_height\n                spacing: dp(8)\n\n        Button:\n            text: "VOLVER AL INICIO"\n            size_hint_y: None\n            height: dp(50)\n            font_size: \'14sp\'\n            bold: True\n            background_color: utils.get_color_from_hex(\'#334155\')\n            on_release: root.manager.current = \'inventario\'\n')

# KV: stock_estadisticas.kv
Builder.load_string('#:import utils kivy.utils\n#:import dp kivy.metrics.dp\n\n<StockScreen>:\n    BoxLayout:\n        orientation: \'vertical\'\n        padding: dp(15)\n        spacing: dp(10)\n        canvas.before:\n            Color:\n                rgba: utils.get_color_from_hex(\'#0A0A12\')\n            Rectangle:\n                pos: self.pos\n                size: self.size\n\n        BoxLayout:\n            size_hint_y: None\n            height: dp(50)\n            padding: [dp(5), 0]\n            Label:\n                text: "ESTADO DE STOCK"\n                font_size: \'18sp\'\n                bold: True\n                halign: \'left\'\n                text_size: self.size\n                valign: \'middle\'\n                color: utils.get_color_from_hex(\'#00D4FF\')\n\n        BoxLayout:\n            size_hint_y: None\n            height: dp(45)\n            spacing: dp(5)\n            Spinner:\n                id: stock_spinner\n                text: "BUSCAR MATERIAL"\n                values: []\n                on_text: root.consultar()\n            Button:\n                text: ""\n                size_hint_x: None\n                width: dp(45)\n                background_color: utils.get_color_from_hex(\'#334155\')\n                on_release: root.consultar()\n\n        TextInput:\n            id: f_s\n            hint_text: "Escribe para filtrar..."\n            multiline: False\n            size_hint_y: None\n            height: dp(45)\n            on_text: root.consultar()\n\n        BoxLayout:\n            size_hint_y: None\n            height: dp(35)\n            spacing: dp(5)\n            Button:\n                text: "TODO"\n                font_size: \'12sp\'\n                background_color: utils.get_color_from_hex(\'#334155\')\n                on_release: root.set_filtro_estado("")\n            Button:\n                text: "SIN STOCK"\n                font_size: \'12sp\'\n                background_color: utils.get_color_from_hex(\'#FF0044\')\n                on_release: root.set_filtro_estado("sin_stock")\n            Button:\n                text: "BAJO"\n                font_size: \'12sp\'\n                background_color: utils.get_color_from_hex(\'#FF0044\')\n                on_release: root.set_filtro_estado("bajo")\n            Button:\n                text: "OK"\n                font_size: \'12sp\'\n                background_color: utils.get_color_from_hex(\'#0066FF\')\n                on_release: root.set_filtro_estado("ok")\n\n        ScrollView:\n            GridLayout:\n                id: s_r\n                cols: 1\n                size_hint_y: None\n                height: self.minimum_height\n                spacing: dp(10)\n\n        Button:\n            text: "VOLVER AL INICIO"\n            size_hint_y: None\n            height: dp(50)\n            font_size: \'14sp\'\n            bold: True\n            background_color: utils.get_color_from_hex(\'#334155\')\n            on_release: root.manager.current = \'inventario\'\n\n<EstadisticasScreen>:\n    BoxLayout:\n        orientation: \'vertical\'\n        padding: dp(15)\n        spacing: dp(10)\n        canvas.before:\n            Color:\n                rgba: utils.get_color_from_hex(\'#0A0A12\')\n            Rectangle:\n                pos: self.pos\n                size: self.size\n\n        BoxLayout:\n            size_hint_y: None\n            height: dp(50)\n            padding: [dp(5), 0]\n            Label:\n                text: "ESTADISTICAS"\n                font_size: \'18sp\'\n                bold: True\n                halign: \'left\'\n                text_size: self.size\n                valign: \'middle\'\n                color: utils.get_color_from_hex(\'#FF0044\')\n\n        ScrollView:\n            GridLayout:\n                id: stats_container\n                cols: 1\n                size_hint_y: None\n                height: self.minimum_height\n                spacing: dp(10)\n\n        Button:\n            text: "VOLVER AL INICIO"\n            size_hint_y: None\n            height: dp(50)\n            font_size: \'14sp\'\n            bold: True\n            background_color: utils.get_color_from_hex(\'#334155\')\n            on_release: root.manager.current = \'inventario\'\n')

# KV: ubicaciones.kv
Builder.load_string('#:import utils kivy.utils\n#:import dp kivy.metrics.dp\n\n<UbicacionesScreen>:\n    BoxLayout:\n        orientation: \'vertical\'\n        padding: dp(15)\n        spacing: dp(10)\n        canvas.before:\n            Color:\n                rgba: utils.get_color_from_hex(\'#0A0A12\')\n            Rectangle:\n                pos: self.pos\n                size: self.size\n\n        BoxLayout:\n            size_hint_y: None\n            height: dp(50)\n            spacing: dp(8)\n            Label:\n                text: "UBICACIONES"\n                font_size: \'16sp\'\n                bold: True\n                color: utils.get_color_from_hex(\'#00D4FF\')\n                halign: \'left\'\n                valign: \'middle\'\n                text_size: self.size\n            Button:\n                text: "+ NUEVA"\n                size_hint_x: None\n                width: dp(100)\n                font_size: \'13sp\'\n                bold: True\n                background_color: utils.get_color_from_hex(\'#0066FF\')\n                on_release: root.mostrar_nueva()\n            Button:\n                text: "VOLVER"\n                size_hint_x: None\n                width: dp(90)\n                font_size: \'12sp\'\n                background_color: utils.get_color_from_hex(\'#334155\')\n                on_release: root.manager.current = \'inventario\'\n\n        ScrollView:\n            GridLayout:\n                id: lista_ubicaciones\n                cols: 1\n                size_hint_y: None\n                height: self.minimum_height\n                spacing: dp(6)\n')

# KV: valor.kv
Builder.load_string('#:import utils kivy.utils\n#:import dp kivy.metrics.dp\n\n<ValorScreen>:\n    BoxLayout:\n        orientation: \'vertical\'\n        padding: [dp(15), 0, dp(15), dp(15)]\n        spacing: dp(10)\n        canvas.before:\n            Color:\n                rgba: utils.get_color_from_hex(\'#0A0A12\')\n            Rectangle:\n                pos: self.pos\n                size: self.size\n\n        Label:\n            text: "VALOR DEL INVENTARIO"\n            font_size: \'16sp\'\n            bold: True\n            color: utils.get_color_from_hex(\'#00D4FF\')\n            size_hint_y: None\n            height: dp(40)\n            halign: \'left\'\n            valign: \'middle\'\n            text_size: self.size\n            padding: [dp(5), 0]\n\n        BoxLayout:\n            orientation: \'vertical\'\n            size_hint_y: 1\n            spacing: dp(8)\n\n            Widget:\n                size_hint_y: 1\n\n            BoxLayout:\n                orientation: \'vertical\'\n                size_hint_y: None\n                height: self.minimum_height\n                spacing: dp(8)\n\n                BoxLayout:\n                    orientation: \'vertical\'\n                    size_hint_y: None\n                    spacing: dp(6)\n                    padding: dp(12)\n                    canvas.before:\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#0D1B2A\')\n                        RoundedRectangle:\n                            pos: self.pos\n                            size: self.size\n                            radius: [dp(4)]\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#1A2A4A\')\n                        Line:\n                            rounded_rectangle: (self.x, self.y, self.width, self.height, dp(4))\n                            width: dp(1)\n                    Label:\n                        text: "ASIGNAR PRECIO DE VENTA"\n                        bold: True\n                        color: utils.get_color_from_hex(\'#00D4FF\')\n                        size_hint_y: None\n                        height: dp(20)\n                        font_size: \'11sp\'\n                        halign: \'left\'\n                        text_size: self.size\n                    Spinner:\n                        id: spinner_material\n                        text: "-- SELECCIONAR --"\n                        values: []\n                        size_hint_y: None\n                        height: dp(38)\n                        font_size: \'12sp\'\n                        on_text: root.actualizar_valor_unitario()\n                    BoxLayout:\n                        size_hint_y: None\n                        height: dp(38)\n                        spacing: dp(6)\n                        TextInput:\n                            id: input_precio\n                            hint_text: "Precio venta unitario ($ COP)"\n                            input_filter: \'int\'\n                            multiline: False\n                            font_size: \'13sp\'\n                        Button:\n                            text: "ASIGNAR"\n                            size_hint_x: None\n                            width: dp(90)\n                            font_size: \'11sp\'\n                            background_color: utils.get_color_from_hex(\'#0066FF\')\n                            on_release: root.asignar_precio()\n                    Label:\n                        id: info_valor\n                        text: ""\n                        color: utils.get_color_from_hex(\'#00D4FF\')\n                        size_hint_y: None\n                        height: dp(18)\n                        font_size: \'10sp\'\n                        halign: \'left\'\n                        text_size: self.size\n\n                BoxLayout:\n                    size_hint_y: None\n                    height: dp(38)\n                    spacing: dp(6)\n                    padding: [dp(10), 0]\n                    canvas.before:\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#0D1B2A\')\n                        RoundedRectangle:\n                            pos: self.pos\n                            size: self.size\n                            radius: [dp(4)]\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#1A2A4A\')\n                        Line:\n                            rounded_rectangle: (self.x, self.y, self.width, self.height, dp(4))\n                            width: dp(1)\n                    Label:\n                        text: "VALOR UNITARIO:"\n                        font_size: \'12sp\'\n                        bold: True\n                        color: utils.get_color_from_hex(\'#E8EAED\')\n                        halign: \'left\'\n                        valign: \'middle\'\n                        text_size: self.size\n                        size_hint_x: 0.5\n                    Label:\n                        id: unitario_label\n                        text: "$ 0"\n                        font_size: \'16sp\'\n                        bold: True\n                        color: utils.get_color_from_hex(\'#5A6A7A\')\n                        halign: \'right\'\n                        valign: \'middle\'\n                        text_size: self.size\n                        size_hint_x: 0.5\n\n                BoxLayout:\n                    size_hint_y: None\n                    height: dp(38)\n                    spacing: dp(6)\n                    padding: [dp(10), 0]\n                    canvas.before:\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#0D1B2A\')\n                        RoundedRectangle:\n                            pos: self.pos\n                            size: self.size\n                            radius: [dp(4)]\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#1A2A4A\')\n                        Line:\n                            rounded_rectangle: (self.x, self.y, self.width, self.height, dp(4))\n                            width: dp(1)\n                    Label:\n                        text: "VALOR TOTAL:"\n                        font_size: \'12sp\'\n                        bold: True\n                        color: utils.get_color_from_hex(\'#E8EAED\')\n                        halign: \'left\'\n                        valign: \'middle\'\n                        text_size: self.size\n                        size_hint_x: 0.5\n                    Label:\n                        id: total_label\n                        text: "$ 0"\n                        font_size: \'16sp\'\n                        bold: True\n                        color: utils.get_color_from_hex(\'#5A6A7A\')\n                        halign: \'right\'\n                        valign: \'middle\'\n                        text_size: self.size\n                        size_hint_x: 0.5\n\n                BoxLayout:\n                    size_hint_y: None\n                    height: dp(38)\n                    spacing: dp(6)\n                    padding: [dp(10), 0]\n                    canvas.before:\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#0D1B2A\')\n                        RoundedRectangle:\n                            pos: self.pos\n                            size: self.size\n                            radius: [dp(4)]\n                        Color:\n                            rgba: utils.get_color_from_hex(\'#1A2A4A\')\n                        Line:\n                            rounded_rectangle: (self.x, self.y, self.width, self.height, dp(4))\n                            width: dp(1)\n                    Label:\n                        text: "VALOR TOTAL INVENTARIO:"\n                        font_size: \'12sp\'\n                        bold: True\n                        color: utils.get_color_from_hex(\'#E8EAED\')\n                        halign: \'left\'\n                        valign: \'middle\'\n                        text_size: self.size\n                        size_hint_x: 0.5\n                    Label:\n                        id: total_inventario_label\n                        text: "$ 0"\n                        font_size: \'16sp\'\n                        bold: True\n                        color: utils.get_color_from_hex(\'#5A6A7A\')\n                        halign: \'right\'\n                        valign: \'middle\'\n                        text_size: self.size\n                        size_hint_x: 0.5\n\n            Widget:\n                size_hint_y: 1\n\n        ScrollView:\n            id: scroll_items\n            size_hint_y: None\n            height: 0\n            GridLayout:\n                id: lista_valores\n                cols: 1\n                size_hint_y: None\n                height: self.minimum_height\n                spacing: dp(4)\n\n        Button:\n            text: "VOLVER"\n            size_hint_y: None\n            height: dp(42)\n            font_size: \'12sp\'\n            background_color: utils.get_color_from_hex(\'#0D1B2A\')\n            color: utils.get_color_from_hex(\'#00D4FF\')\n            on_release: root.manager.current = \'inventario\'\n')

# KV: inventario.kv
Builder.load_string('#:import utils kivy.utils\n#:import dp kivy.metrics.dp\n\n<Button>:\n    background_normal: \'\'\n    background_down: \'\'\n    canvas.before:\n        Color:\n            rgba: self.background_color if self.state == \'normal\' else [c * 0.6 for c in self.background_color]\n        RoundedRectangle:\n            pos: self.pos\n            size: self.size\n            radius: [dp(4)]\n\n<TextInput>:\n    background_normal: \'\'\n    background_active: \'\'\n    background_color: utils.get_color_from_hex(\'#0D1B2A\')\n    foreground_color: utils.get_color_from_hex(\'#E8EAED\')\n    cursor_color: utils.get_color_from_hex(\'#00D4FF\')\n    hint_text_color: utils.get_color_from_hex(\'#5A6A7A\')\n    padding: [dp(10), (self.height - self.line_height) / 2]\n    canvas.after:\n        Color:\n            rgba: utils.get_color_from_hex(\'#00D4FF\') if self.focus else utils.get_color_from_hex(\'#1A2A4A\')\n        Line:\n            rounded_rectangle: (self.x, self.y, self.width, self.height, dp(4))\n            width: dp(1.5) if self.focus else dp(1)\n\n<SpinnerOption>:\n    background_normal: \'\'\n    background_color: utils.get_color_from_hex(\'#0D1B2A\')\n    color: utils.get_color_from_hex(\'#E8EAED\')\n    font_size: \'13sp\'\n\n<Spinner>:\n    background_normal: \'\'\n    background_down: \'\'\n    background_color: utils.get_color_from_hex(\'#0D1B2A\')\n    color: utils.get_color_from_hex(\'#E8EAED\')\n    font_size: \'13sp\'\n    option_cls: \'SpinnerOption\'\n    canvas.before:\n        Color:\n            rgba: self.background_color\n        RoundedRectangle:\n            pos: self.pos\n            size: self.size\n            radius: [dp(4)]\n        Color:\n            rgba: utils.get_color_from_hex(\'#1A2A4A\')\n        Line:\n            rounded_rectangle: (self.x, self.y, self.width, self.height, dp(4))\n            width: dp(1)\n\n<InventarioScreen>:\n    RelativeLayout:\n\n        BoxLayout:\n            id: main_area\n            orientation: \'vertical\'\n            padding: dp(15)\n            spacing: dp(8)\n            canvas.before:\n                Color:\n                    rgba: utils.get_color_from_hex(\'#0A0A12\')\n                Rectangle:\n                    pos: self.pos\n                    size: self.size\n\n            BoxLayout:\n                size_hint_y: None\n                height: dp(50)\n                spacing: dp(8)\n                padding: [dp(5), 0]\n                Button:\n                    id: btn_menu\n                    text: "MENU"\n                    size_hint_x: None\n                    width: dp(70)\n                    font_size: \'12sp\'\n                    bold: True\n                    background_color: utils.get_color_from_hex(\'#0D1B2A\')\n                    color: utils.get_color_from_hex(\'#00D4FF\')\n                    on_release: root.toggle_menu()\n                Widget:\n\n            ScrollView:\n                do_scroll_x: False\n                BoxLayout:\n                    orientation: \'vertical\'\n                    size_hint_y: None\n                    height: self.minimum_height\n                    spacing: dp(12)\n                    padding: [0, dp(6)]\n\n                    # Dashboard\n                    BoxLayout:\n                        orientation: \'vertical\'\n                        size_hint_y: None\n                        height: dp(140)\n                        padding: dp(10)\n                        spacing: dp(6)\n                        canvas.before:\n                            Color:\n                                rgba: utils.get_color_from_hex(\'#0D1B2A\')\n                            RoundedRectangle:\n                                pos: self.pos\n                                size: self.size\n                                radius: [dp(4)]\n                            Color:\n                                rgba: utils.get_color_from_hex(\'#00D4FF\')\n                            Line:\n                                rounded_rectangle: (self.x, self.y, self.width, self.height, dp(4))\n                                width: dp(1.5)\n                        canvas.after:\n                            Color:\n                                rgba: 0, 0.831, 1, 0.06\n                            Line:\n                                rounded_rectangle: (self.x - dp(2), self.y - dp(2), self.width + dp(4), self.height + dp(4), dp(4))\n                                width: dp(1)\n                        Label:\n                            id: lbl_saludo\n                            text: "Buenos dias"\n                            font_size: \'11sp\'\n                            color: utils.get_color_from_hex(\'#00D4FF\')\n                            size_hint_y: None\n                            height: dp(18)\n                            text_size: self.size\n                            halign: \'left\'\n                        Label:\n                            id: lbl_fecha_hoy\n                            text: ""\n                            font_size: \'9sp\'\n                            color: utils.get_color_from_hex(\'#5A6A7A\')\n                            size_hint_y: None\n                            height: dp(14)\n                            text_size: self.size\n                            halign: \'left\'\n                        BoxLayout:\n                            spacing: dp(8)\n                            size_hint_y: None\n                            height: dp(60)\n                            BoxLayout:\n                                orientation: \'vertical\'\n                                canvas.before:\n                                    Color:\n                                        rgba: utils.get_color_from_hex(\'#0A0A12\')\n                                    RoundedRectangle:\n                                        pos: self.pos\n                                        size: self.size\n                                        radius: [dp(4)]\n                                padding: [dp(4), dp(3)]\n                                Label:\n                                    id: stat_materiales\n                                    text: "0"\n                                    font_size: \'18sp\'\n                                    bold: True\n                                    color: utils.get_color_from_hex(\'#00D4FF\')\n                                    size_hint_y: None\n                                    height: dp(28)\n                                    text_size: self.size\n                                    halign: \'center\'\n                                    valign: \'middle\'\n                                Label:\n                                    text: "MATERIALES"\n                                    font_size: \'8sp\'\n                                    color: utils.get_color_from_hex(\'#5A6A7A\')\n                                    size_hint_y: None\n                                    height: dp(12)\n                                    text_size: self.size\n                                    halign: \'center\'\n                                    valign: \'middle\'\n                            BoxLayout:\n                                orientation: \'vertical\'\n                                canvas.before:\n                                    Color:\n                                        rgba: utils.get_color_from_hex(\'#0A0A12\')\n                                    RoundedRectangle:\n                                        pos: self.pos\n                                        size: self.size\n                                        radius: [dp(4)]\n                                padding: [dp(4), dp(3)]\n                                Label:\n                                    id: stat_movs_hoy\n                                    text: "0"\n                                    font_size: \'18sp\'\n                                    bold: True\n                                    color: utils.get_color_from_hex(\'#00D4FF\')\n                                    size_hint_y: None\n                                    height: dp(28)\n                                    text_size: self.size\n                                    halign: \'center\'\n                                    valign: \'middle\'\n                                Label:\n                                    text: "MOV. HOY"\n                                    font_size: \'8sp\'\n                                    color: utils.get_color_from_hex(\'#5A6A7A\')\n                                    size_hint_y: None\n                                    height: dp(12)\n                                    text_size: self.size\n                                    halign: \'center\'\n                                    valign: \'middle\'\n                            BoxLayout:\n                                orientation: \'vertical\'\n                                canvas.before:\n                                    Color:\n                                        rgba: utils.get_color_from_hex(\'#0A0A12\')\n                                    RoundedRectangle:\n                                        pos: self.pos\n                                        size: self.size\n                                        radius: [dp(4)]\n                                padding: [dp(4), dp(3)]\n                                Label:\n                                    id: stat_alertas\n                                    text: "0"\n                                    font_size: \'18sp\'\n                                    bold: True\n                                    color: utils.get_color_from_hex(\'#FF0044\')\n                                    size_hint_y: None\n                                    height: dp(28)\n                                    text_size: self.size\n                                    halign: \'center\'\n                                    valign: \'middle\'\n                                Label:\n                                    text: "ALERTAS"\n                                    font_size: \'8sp\'\n                                    color: utils.get_color_from_hex(\'#5A6A7A\')\n                                    size_hint_y: None\n                                    height: dp(12)\n                                    text_size: self.size\n                                    halign: \'center\'\n                                    valign: \'middle\'\n                            BoxLayout:\n                                orientation: \'vertical\'\n                                canvas.before:\n                                    Color:\n                                        rgba: utils.get_color_from_hex(\'#0A0A12\')\n                                    RoundedRectangle:\n                                        pos: self.pos\n                                        size: self.size\n                                        radius: [dp(4)]\n                                padding: [dp(4), dp(3)]\n                                Label:\n                                    id: stat_pendientes\n                                    text: "0"\n                                    font_size: \'18sp\'\n                                    bold: True\n                                    color: utils.get_color_from_hex(\'#FF0044\')\n                                    size_hint_y: None\n                                    height: dp(28)\n                                    text_size: self.size\n                                    halign: \'center\'\n                                    valign: \'middle\'\n                                Label:\n                                    text: "FUERA"\n                                    font_size: \'8sp\'\n                                    color: utils.get_color_from_hex(\'#5A6A7A\')\n                                    size_hint_y: None\n                                    height: dp(12)\n                                    text_size: self.size\n                                    halign: \'center\'\n                                    valign: \'middle\'\n                    Label:\n                        id: lbl_resumen_db\n                        text: ""\n                        font_size: \'9sp\'\n                        color: utils.get_color_from_hex(\'#475569\')\n                        size_hint_y: None\n                        height: dp(14)\n                        text_size: self.size\n                        halign: \'right\'\n\n                    # Nuevo Movimiento\n                    BoxLayout:\n                        id: seccion_movimiento\n                        orientation: \'vertical\'\n                        spacing: dp(6)\n                        padding: [dp(12), dp(5), dp(12), dp(10)]\n                        size_hint_y: None\n                        height: self.minimum_height\n                        canvas.before:\n                            Color:\n                                rgba: utils.get_color_from_hex(\'#0D1B2A\')\n                            RoundedRectangle:\n                                pos: self.pos\n                                size: self.size\n                                radius: [dp(4)]\n                            Color:\n                                rgba: utils.get_color_from_hex(\'#1A2A4A\')\n                            Line:\n                                rounded_rectangle: (self.x, self.y, self.width, self.height, dp(4))\n                                width: dp(1)\n                        Label:\n                            text: "NUEVO MOVIMIENTO"\n                            bold: True\n                            font_size: \'11sp\'\n                            color: utils.get_color_from_hex(\'#00D4FF\')\n                            size_hint_y: None\n                            height: dp(18)\n                            text_size: self.size\n                            halign: \'left\'\n                        Spinner:\n                            id: responsable_spinner\n                            text: "SELECCIONAR RESPONSABLE"\n                            values: []\n                            size_hint_y: None\n                            height: dp(38)\n                        TextInput:\n                            id: nombre_nuevo\n                            hint_text: "O escribe nuevo responsable..."\n                            multiline: False\n                            size_hint_y: None\n                            height: dp(38)\n                        Spinner:\n                            id: producto_spinner\n                            text: "SELECCIONAR MATERIAL"\n                            values: []\n                            size_hint_y: None\n                            height: dp(38)\n                        TextInput:\n                            id: producto_nuevo\n                            hint_text: "O escribe material nuevo..."\n                            multiline: False\n                            size_hint_y: None\n                            height: dp(38)\n                        BoxLayout:\n                            spacing: dp(6)\n                            size_hint_y: None\n                            height: dp(38)\n                            TextInput:\n                                id: sku\n                                hint_text: "SKU / ID"\n                                multiline: False\n                            TextInput:\n                                id: cantidad\n                                hint_text: "Cantidad"\n                                input_filter: \'float\'\n                                multiline: False\n                        BoxLayout:\n                            spacing: dp(6)\n                            size_hint_y: None\n                            height: dp(38)\n                            TextInput:\n                                id: d_p\n                                hint_text: "Dias (salida)"\n                                input_filter: \'int\'\n                                multiline: False\n                            TextInput:\n                                id: ubicacion\n                                hint_text: "Ubicacion"\n                                multiline: False\n                        TextInput:\n                            id: f_m\n                            hint_text: "Fecha Manual (dd/mm/aaaa hh:mm)"\n                            multiline: False\n                            size_hint_y: None\n                            height: dp(38)\n                        TextInput:\n                            id: notas\n                            hint_text: "Notas / Observaciones..."\n                            multiline: False\n                            size_hint_y: None\n                            height: dp(38)\n                        BoxLayout:\n                            size_hint_y: None\n                            height: dp(32)\n                            spacing: dp(8)\n                            CheckBox:\n                                id: c_l\n                                active: False\n                                size_hint_x: None\n                                width: dp(36)\n                            Label:\n                                text: "Enviar a Limpieza (solo INVENTARIO)"\n                                halign: \'left\'\n                                valign: \'middle\'\n                                text_size: self.size\n                                font_size: \'11sp\'\n                                color: utils.get_color_from_hex(\'#94A3B8\')\n                        BoxLayout:\n                            size_hint_y: None\n                            height: dp(36)\n                            spacing: dp(8)\n                            Button:\n                                text: "LIMPIAR CAMPOS"\n                                background_color: utils.get_color_from_hex(\'#334155\')\n                                font_size: \'11sp\'\n                                on_release: root.limpiar_campos()\n                            Button:\n                                id: btn_deshacer\n                                text: "DESHACER"\n                                background_color: utils.get_color_from_hex(\'#7C3AED\')\n                                font_size: \'11sp\'\n                                disabled: True\n                                on_release: root.deshacer_ultimo()\n\n                    Label:\n                        id: info_carpeta\n                        text: ""\n                        color: utils.get_color_from_hex(\'#5A6A7A\')\n                        markup: True\n                        font_size: \'9sp\'\n                        size_hint_y: None\n                        height: dp(20)\n                        text_size: self.size\n                        halign: \'left\'\n                        valign: \'middle\'\n                        shorten: True\n                        shorten_from: \'left\'\n                    Label:\n                        id: info\n                        text: "Listo"\n                        color: utils.get_color_from_hex(\'#5A6A7A\')\n                        markup: True\n                        size_hint_y: None\n                        height: dp(30)\n                        text_size: self.size\n                        halign: \'left\'\n                        valign: \'middle\'\n\n                    # Acciones Rapidas\n                    BoxLayout:\n                        id: seccion_acciones\n                        orientation: \'vertical\'\n                        size_hint_y: None\n                        height: dp(120)\n                        padding: dp(12)\n                        spacing: dp(8)\n                        canvas.before:\n                            Color:\n                                rgba: utils.get_color_from_hex(\'#0D1B2A\')\n                            RoundedRectangle:\n                                pos: self.pos\n                                size: self.size\n                                radius: [dp(4)]\n                            Color:\n                                rgba: utils.get_color_from_hex(\'#1A2A4A\')\n                            Line:\n                                rounded_rectangle: (self.x, self.y, self.width, self.height, dp(4))\n                                width: dp(1)\n                        Label:\n                            text: "ACCIONES RAPIDAS"\n                            bold: True\n                            font_size: \'11sp\'\n                            color: utils.get_color_from_hex(\'#00D4FF\')\n                            size_hint_y: None\n                            height: dp(18)\n                            text_size: self.size\n                            halign: \'left\'\n                        BoxLayout:\n                            spacing: dp(10)\n                            Button:\n                                text: "ENTRADA"\n                                font_size: \'14sp\'\n                                bold: True\n                                on_release: root.procesar("ENTRADA")\n                                background_color: utils.get_color_from_hex(\'#0066FF\')\n                            Button:\n                                text: "SALIDA"\n                                font_size: \'14sp\'\n                                bold: True\n                                on_release: root.procesar("SALIDA")\n                                background_color: utils.get_color_from_hex(\'#FF0044\')\n                            Button:\n                                text: "SALIDA CAT"\n                                font_size: \'12sp\'\n                                bold: True\n                                on_release: root.salida_por_categoria()\n                                background_color: utils.get_color_from_hex(\'#CC0033\')\n\n        BoxLayout:\n            id: drawer_overlay\n            size_hint: 1, 1\n            opacity: 0\n            canvas.before:\n                Color:\n                    rgba: 0, 0, 0, 0.6\n                Rectangle:\n                    pos: self.pos\n                    size: self.size\n\n        BoxLayout:\n            id: drawer\n            orientation: \'vertical\'\n            size_hint_x: 0.7\n            size_hint_y: 1\n            opacity: 0\n            padding: dp(15)\n            spacing: dp(6)\n            canvas.before:\n                Color:\n                    rgba: utils.get_color_from_hex(\'#0D1B2A\')\n                RoundedRectangle:\n                    pos: self.pos\n                    size: self.size\n                    radius: [0, dp(4), dp(4), 0]\n            Label:\n                text: "HERRAMIENTAS"\n                bold: True\n                font_size: \'13sp\'\n                color: utils.get_color_from_hex(\'#00D4FF\')\n                size_hint_y: None\n                height: dp(38)\n                halign: \'left\'\n                valign: \'middle\'\n                text_size: self.size\n                padding: [dp(6), 0]\n            GridLayout:\n                id: menu_grid\n                cols: 1\n                spacing: dp(4)\n                size_hint_y: None\n                height: self.minimum_height\n\n\n')

# === config.py ===
import os
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

VERSION_ACTUAL = os.environ.get("APP_VERSION", "3.0.0")
URL_NUEVA_APK  = os.environ.get("UPDATE_URL", "https://tu-servidor.com/app_actualizada.apk")

EMAIL_CONFIG = {
    "remitente":        os.environ.get("APP_EMAIL", ""),
    "password":         os.environ.get("APP_EMAIL_PASS", ""),
    "destinatarios":    [os.environ.get("APP_EMAIL", "")],
    "imap_servidor":    "imap.gmail.com",
    "imap_puerto":      993,
    "asunto_filtro":    os.environ.get("FILTRO_ASUNTO", ""),
    "remitente_filtro": os.environ.get("FILTRO_REMITENTE", ""),
}

def _get_db_path() -> str:
    try:
        from kivy.utils import platform
        if platform == 'android':
            from android.storage import app_storage_path
            import sqlite3
            db = os.path.join(app_storage_path(), "gestion_inventario.db")
            return db
    except Exception:
        pass
    return "gestion_inventario.db"

DB_PATH = _get_db_path()

def _get_export_dir() -> str:
    try:
        from kivy.utils import platform
        if platform == 'android':
            export = "/storage/emulated/0/Download/InventarioPRO"
        else:
            export = os.path.join(str(Path.home()), "Downloads", "InventarioPRO")
        os.makedirs(export, exist_ok=True)
        return export
    except Exception:
        return os.path.dirname(os.path.abspath(__file__))

EXPORT_DIR = _get_export_dir()


# === logging_config.py ===
import logging
import os
import traceback
import sys
from datetime import datetime

LOG_FORMAT = "%(asctime)s [%(levelname)s] %(name)s: %(message)s"
LOG_DIR = os.path.dirname(os.path.abspath(__file__))

def setup_logging() -> logging.Logger:
    logger = logging.getLogger("inventario_pro")
    logger.setLevel(logging.INFO)
    log_file = os.path.join(LOG_DIR, "app.log")
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setFormatter(logging.Formatter(LOG_FORMAT, datefmt="%d/%m/%Y %H:%M:%S"))
    logger.addHandler(fh)
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
    logger.addHandler(ch)
    return logger

logger = setup_logging()

def guardar_crash(tipo, valor, tb) -> None:
    try:
        ruta = os.path.join(LOG_DIR, "crash_log.txt")
        with open(ruta, "w", encoding="utf-8") as f:
            f.write(f"CRASH {datetime.now()}\n")
            traceback.print_exception(tipo, valor, tb, file=f)
        logger.critical("Crash guardado en %s", ruta)
    except Exception:
        pass

sys.excepthook = guardar_crash


# === database/connection.py ===
import sqlite3
import threading
from typing import Optional


class DatabaseConnection:
    _instances: dict = {}
    _lock = threading.Lock()

    def __init__(self, db_path: str):
        self.db_path = db_path
        self._local = threading.local()

    @classmethod
    def get_instance(cls, db_path: str) -> "DatabaseConnection":
        if db_path not in cls._instances:
            with cls._lock:
                if db_path not in cls._instances:
                    cls._instances[db_path] = cls(db_path)
        return cls._instances[db_path]

    def get_connection(self) -> sqlite3.Connection:
        if not hasattr(self._local, "conn") or self._local.conn is None:
            self._local.conn = sqlite3.connect(self.db_path)
            self._local.conn.row_factory = sqlite3.Row
            self._local.conn.execute("PRAGMA journal_mode=WAL")
            self._local.conn.execute("PRAGMA foreign_keys=ON")
        return self._local.conn

    def close(self) -> None:
        if hasattr(self._local, "conn") and self._local.conn:
            self._local.conn.close()
            self._local.conn = None

    @classmethod
    def close_all(cls) -> None:
        for inst in cls._instances.values():
            inst.close()
        cls._instances.clear()

    @classmethod
    def close_path(cls, db_path: str) -> None:
        if db_path in cls._instances:
            cls._instances[db_path].close()
            del cls._instances[db_path]


def db_query(sql: str, params: tuple = (), fetchall: bool = True, db_path: Optional[str] = None):
    
    path = db_path or DB_PATH
    conn = DatabaseConnection.get_instance(path).get_connection()
    c = conn.cursor()
    c.execute(sql, params)
    if fetchall:
        return c.fetchall()
    return c.fetchone()


def db_execute(sql: str, params: tuple = (), db_path: Optional[str] = None):
    
    path = db_path or DB_PATH
    conn = DatabaseConnection.get_instance(path).get_connection()
    try:
        c = conn.cursor()
        c.execute(sql, params)
        conn.commit()
        return c.lastrowid
    except sqlite3.Error as e:
        conn.rollback()
        raise e


# === database/models.py ===
import sqlite3
from typing import Optional



def init_db(db_path: str) -> None:
    conn = DatabaseConnection.get_instance(db_path).get_connection()
    c = conn.cursor()

    c.execute('''CREATE TABLE IF NOT EXISTS movimientos (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre         TEXT,
        material       TEXT,
        sku            TEXT,
        cantidad       REAL,
        tipo           TEXT,
        fecha          TEXT,
        stock_registro REAL,
        dias           INTEGER,
        retorno        TEXT DEFAULT "N/A",
        notas          TEXT DEFAULT "",
        ubicacion      TEXT DEFAULT ""
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS limpieza (
        id       INTEGER PRIMARY KEY AUTOINCREMENT,
        material TEXT,
        cantidad REAL,
        estado   TEXT,
        fecha    TEXT,
        notas    TEXT DEFAULT ""
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS alertas_stock (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        material     TEXT UNIQUE,
        stock_minimo REAL DEFAULT 5
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS ubicaciones (
        id     INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT UNIQUE
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS auditoria (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        accion    TEXT,
        detalle   TEXT,
        fecha     TEXT,
        usuario   TEXT DEFAULT "sistema"
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS actualizaciones_correo (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha        TEXT,
        remitente    TEXT,
        asunto       TEXT,
        archivo      TEXT,
        insertados   INTEGER,
        omitidos     INTEGER,
        errores      INTEGER
    )''')

    # Fase 0: Nuevas tablas
    c.execute('''CREATE TABLE IF NOT EXISTS usuarios (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        username        TEXT UNIQUE NOT NULL,
        password_hash   TEXT NOT NULL,
        rol             TEXT NOT NULL DEFAULT 'editor',
        nombre_completo TEXT NOT NULL DEFAULT '',
        email           TEXT DEFAULT '',
        ultimo_acceso   TEXT,
        creado_en       TEXT DEFAULT (datetime('now','localtime'))
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS categorias (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre      TEXT UNIQUE NOT NULL,
        descripcion TEXT DEFAULT '',
        color_hex   TEXT DEFAULT '#38BDF8'
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS categoria_items (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        categoria_id INTEGER NOT NULL REFERENCES categorias(id),
        nombre       TEXT NOT NULL,
        UNIQUE(categoria_id, nombre)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS sync_log (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        dispositivo_id  TEXT NOT NULL,
        accion          TEXT NOT NULL,
        tabla           TEXT NOT NULL,
        registro_id     INTEGER,
        datos_json      TEXT,
        timestamp       TEXT NOT NULL,
        estado          TEXT DEFAULT 'pendiente'
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS notificaciones (
        id      INTEGER PRIMARY KEY AUTOINCREMENT,
        tipo    TEXT NOT NULL,
        titulo  TEXT NOT NULL,
        mensaje TEXT,
        leida   INTEGER DEFAULT 0,
        fecha   TEXT DEFAULT (datetime('now','localtime'))
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS config_app (
        clave TEXT PRIMARY KEY,
        valor TEXT NOT NULL
    )''')

    _migrate_columns(c)
    _create_indexes(c)
    conn.commit()


def _migrate_columns(c: sqlite3.Cursor) -> None:
    for tabla, columna, definicion in [
        ("movimientos", "notas",    "TEXT DEFAULT ''"),
        ("movimientos", "ubicacion","TEXT DEFAULT ''"),
        ("movimientos", "usuario_id",      "INTEGER DEFAULT NULL"),
        ("movimientos", "categoria_id",    "INTEGER DEFAULT NULL"),
        ("movimientos", "costo_unitario",  "REAL DEFAULT 0"),
        ("movimientos", "precio_venta",    "REAL DEFAULT 0"),
        ("movimientos", "dispositivo_id",  "TEXT DEFAULT ''"),
        ("movimientos", "ultima_sincronizacion", "TEXT DEFAULT ''"),
        ("limpieza",    "notas",    "TEXT DEFAULT ''"),
    ]:
        try:
            c.execute(f"ALTER TABLE {tabla} ADD COLUMN {columna} {definicion}")
        except sqlite3.OperationalError:
            pass


def _create_indexes(c: sqlite3.Cursor) -> None:
    indexes = [
        "CREATE INDEX IF NOT EXISTS idx_movimientos_material ON movimientos(material)",
        "CREATE INDEX IF NOT EXISTS idx_movimientos_nombre ON movimientos(nombre)",
        "CREATE INDEX IF NOT EXISTS idx_movimientos_fecha ON movimientos(fecha)",
        "CREATE INDEX IF NOT EXISTS idx_movimientos_tipo ON movimientos(tipo)",
        "CREATE INDEX IF NOT EXISTS idx_movimientos_mat_nom ON movimientos(material, nombre)",
        "CREATE INDEX IF NOT EXISTS idx_limpieza_material ON limpieza(material)",
        "CREATE INDEX IF NOT EXISTS idx_auditoria_fecha ON auditoria(fecha)",
        "CREATE INDEX IF NOT EXISTS idx_usuarios_username ON usuarios(username)",
        "CREATE INDEX IF NOT EXISTS idx_notificaciones_leida ON notificaciones(leida)",
        "CREATE INDEX IF NOT EXISTS idx_sync_log_estado ON sync_log(estado)",
    ]
    for idx in indexes:
        try:
            c.execute(idx)
        except sqlite3.OperationalError:
            pass


# === services/auditoria.py ===
from datetime import datetime


def log_auditoria(accion: str, detalle: str) -> None:
    try:
        db_execute(
            "INSERT INTO auditoria (accion, detalle, fecha) VALUES (?,?,?)",
            (accion, detalle, datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        )
    except Exception:
        pass


# === services/inventario.py ===
import os
import csv
from datetime import datetime, timedelta


def ruta_exportacion(nombre_archivo: str) -> str:
    try:
        os.makedirs(EXPORT_DIR, exist_ok=True)
    except Exception:
        pass
    return os.path.join(EXPORT_DIR, nombre_archivo)


def obtener_stock_material(material: str) -> tuple:
    material = material.strip().upper()
    row = db_query(
        "SELECT SUM(cantidad) FROM movimientos WHERE UPPER(material)=? AND UPPER(nombre)='INVENTARIO' AND tipo='ENTRADA'",
        (material,), fetchall=False
    )
    stock_inv = float(row[0] or 0.0)
    movs = db_query(
        "SELECT tipo, cantidad FROM movimientos WHERE UPPER(material)=? AND UPPER(nombre)!='INVENTARIO'",
        (material,)
    )
    delta = sum(float(c) if t == "ENTRADA" else -float(c) for t, c in movs)
    return stock_inv, round(stock_inv + delta, 4)


def obtener_alertas_activas() -> list:
    alertas = db_query("SELECT material, stock_minimo FROM alertas_stock")
    resultado = []
    for mat, minimo in alertas:
        _, disponible = obtener_stock_material(mat)
        if disponible <= minimo:
            resultado.append((mat, disponible, minimo))
    return resultado


def obtener_reporte_estadisticas() -> dict:
    hoy = datetime.now().strftime("%d/%m/%Y")
    hace_7_dias  = (datetime.now() - timedelta(days=7)).strftime("%d/%m/%Y")
    hace_30_dias = (datetime.now() - timedelta(days=30)).strftime("%d/%m/%Y")

    total_movs          = int(db_query("SELECT COUNT(*) FROM movimientos", fetchall=False)[0] or 0)
    total_entradas      = int(db_query("SELECT COUNT(*) FROM movimientos WHERE tipo='ENTRADA'", fetchall=False)[0] or 0)
    total_salidas       = int(db_query("SELECT COUNT(*) FROM movimientos WHERE tipo='SALIDA'", fetchall=False)[0] or 0)
    materiales_unicos   = int(db_query("SELECT COUNT(DISTINCT material) FROM movimientos", fetchall=False)[0] or 0)
    responsables_unicos = int(db_query(
        "SELECT COUNT(DISTINCT nombre) FROM movimientos WHERE UPPER(nombre)!='INVENTARIO'",
        fetchall=False
    )[0] or 0)
    top_mat = db_query(
        "SELECT material, COUNT(*) as cnt FROM movimientos GROUP BY UPPER(material) ORDER BY cnt DESC LIMIT 1",
        fetchall=False
    )
    material_top = str(top_mat[0]) if top_mat else "N/A"

    movs_recientes = int(db_query(
        "SELECT COUNT(*) FROM movimientos WHERE fecha >= ?",
        (hace_7_dias,), fetchall=False
    )[0] or 0)
    movs_30dias = int(db_query(
        "SELECT COUNT(*) FROM movimientos WHERE fecha >= ?",
        (hace_30_dias,), fetchall=False
    )[0] or 0)
    ent_7d = int(db_query(
        "SELECT COUNT(*) FROM movimientos WHERE tipo='ENTRADA' AND fecha >= ?",
        (hace_7_dias,), fetchall=False
    )[0] or 0)
    sal_7d = int(db_query(
        "SELECT COUNT(*) FROM movimientos WHERE tipo='SALIDA' AND fecha >= ?",
        (hace_7_dias,), fetchall=False
    )[0] or 0)
    vol_entradas = float(db_query(
        "SELECT COALESCE(SUM(cantidad), 0) FROM movimientos WHERE tipo='ENTRADA'",
        fetchall=False
    )[0] or 0)
    vol_salidas = float(db_query(
        "SELECT COALESCE(SUM(cantidad), 0) FROM movimientos WHERE tipo='SALIDA'",
        fetchall=False
    )[0] or 0)
    prom_cant = float(db_query(
        "SELECT COALESCE(AVG(cantidad), 0) FROM movimientos",
        fetchall=False
    )[0] or 0)

    top5_materiales = db_query(
        "SELECT UPPER(material), COUNT(*) as cnt FROM movimientos "
        "GROUP BY UPPER(material) ORDER BY cnt DESC LIMIT 5"
    )
    top5_responsables = db_query(
        "SELECT UPPER(nombre), COUNT(*) as cnt FROM movimientos "
        "WHERE UPPER(nombre)!='INVENTARIO' GROUP BY UPPER(nombre) ORDER BY cnt DESC LIMIT 5"
    )

    todos_mats = [r[0] for r in db_query("SELECT DISTINCT UPPER(material) FROM movimientos")]
    mats_recientes = {r[0] for r in db_query(
        "SELECT DISTINCT UPPER(material) FROM movimientos WHERE fecha >= ?",
        (hace_30_dias,)
    )}
    mats_inactivos = [m for m in todos_mats if m not in mats_recientes]

    total_limpieza = int(db_query("SELECT COUNT(*) FROM limpieza", fetchall=False)[0] or 0)
    limpios = int(db_query("SELECT COUNT(*) FROM limpieza WHERE UPPER(estado)='LIMPIO'", fetchall=False)[0] or 0)
    sucios  = int(db_query("SELECT COUNT(*) FROM limpieza WHERE UPPER(estado)='SUCIO'", fetchall=False)[0] or 0)

    primer_mov = db_query("SELECT fecha FROM movimientos ORDER BY id ASC LIMIT 1", fetchall=False)
    ultimo_mov = db_query("SELECT fecha FROM movimientos ORDER BY id DESC LIMIT 1", fetchall=False)

    return {
        "total_movimientos": total_movs,
        "entradas":          total_entradas,
        "salidas":           total_salidas,
        "materiales":        materiales_unicos,
        "responsables":      responsables_unicos,
        "material_top":      material_top,
        "movs_7dias":        movs_recientes,
        "movs_30dias":       movs_30dias,
        "ent_7d":            ent_7d,
        "sal_7d":            sal_7d,
        "vol_entradas":      round(vol_entradas, 2),
        "vol_salidas":       round(vol_salidas, 2),
        "prom_cant":         round(prom_cant, 2),
        "top5_materiales":   top5_materiales,
        "top5_responsables": top5_responsables,
        "mats_inactivos":    mats_inactivos[:10],
        "total_limpieza":    total_limpieza,
        "limpios":           limpios,
        "sucios":            sucios,
        "primer_mov":        str(primer_mov[0]) if primer_mov else "N/A",
        "ultimo_mov":        str(ultimo_mov[0]) if ultimo_mov else "N/A",
    }


# === services/alertas.py ===
from datetime import datetime, timedelta


def validar_movimiento(responsable: str, material: str, cantidad_str: str) -> tuple:
    resp = (responsable or "").strip().upper()
    mat  = (material or "").strip().upper()
    cant_str = (cantidad_str or "").strip().replace(",", ".")

    if not resp:
        return False, "El responsable es obligatorio.", 0.0
    if len(resp) < 2:
        return False, "El nombre del responsable es demasiado corto.", 0.0
    if not mat:
        return False, "El equipo/material es obligatorio.", 0.0
    if len(mat) < 2:
        return False, "El nombre del equipo es demasiado corto.", 0.0
    if not cant_str:
        return False, "La cantidad es obligatoria.", 0.0
    try:
        cant = float(cant_str)
    except ValueError:
        return False, f"Cantidad invalida: '{cant_str}'. Usa numeros (ej. 2 o 1.5).", 0.0
    if cant <= 0:
        return False, "La cantidad debe ser mayor a 0.", 0.0
    if cant > 9999:
        return False, "Cantidad fuera de rango maximo (9999).", 0.0
    return True, "OK", round(cant, 4)


def obtener_equipos_en_prestamo() -> list:
    rows = db_query(
        "SELECT nombre, material, sku, cantidad, tipo, fecha, dias, notas "
        "FROM movimientos WHERE UPPER(nombre) != 'INVENTARIO' ORDER BY id ASC"
    )
    prestamos = {}
    for n, m, s, c, t, f, d, notas in rows:
        key = f"{str(n).upper()}|{str(m).upper()}"
        if t == "SALIDA":
            if key not in prestamos:
                prestamos[key] = {
                    "responsable":    str(n).upper(),
                    "material":       str(m).upper(),
                    "sku":            str(s or ""),
                    "cantidad":       float(c),
                    "fecha_salida":   str(f),
                    "dias_acordados": int(d or 0),
                    "notas":          str(notas or ""),
                }
            else:
                prestamos[key]["cantidad"] += float(c)
        elif t == "ENTRADA" and key in prestamos:
            prestamos[key]["cantidad"] -= float(c)
            if prestamos[key]["cantidad"] <= 0:
                del prestamos[key]

    resultado = []
    ahora = datetime.now()
    for item in prestamos.values():
        if item["cantidad"] <= 0:
            continue
        try:
            fecha_dt = datetime.strptime(item["fecha_salida"][:16], "%d/%m/%Y %H:%M")
            dias_fuera = (ahora - fecha_dt).days
        except Exception:
            dias_fuera = 0

        dias_acordados = item["dias_acordados"]
        if dias_acordados > 0 and dias_fuera > dias_acordados:
            estado = "VENCIDO"
        elif dias_acordados > 0 and dias_fuera >= dias_acordados - 1:
            estado = "POR_VENCER"
        else:
            estado = "EN_TIEMPO"

        resultado.append({**item, "dias_fuera": dias_fuera, "estado": estado})

    orden = {"VENCIDO": 0, "POR_VENCER": 1, "EN_TIEMPO": 2}
    resultado.sort(key=lambda x: (orden[x["estado"]], -x["dias_fuera"]))
    return resultado


def obtener_vencimientos(dias_alerta: int = 1) -> dict:
    prestamos = obtener_equipos_en_prestamo()
    vencidos    = [p for p in prestamos if p["estado"] == "VENCIDO"]
    por_vencer  = [p for p in prestamos if p["estado"] == "POR_VENCER"]
    return {"vencidos": vencidos, "por_vencer": por_vencer}


def consultar_disponibilidad(material: str) -> dict:
    material = material.strip().upper()
    _, stock_disponible = obtener_stock_material(material)
    prestamos_equipo = [
        p for p in obtener_equipos_en_prestamo()
        if p["material"] == material
    ]
    total_fuera = sum(p["cantidad"] for p in prestamos_equipo)
    return {
        "material":          material,
        "stock_total":       stock_disponible + total_fuera,
        "stock_disponible":  round(stock_disponible, 4),
        "total_en_prestamo": round(total_fuera, 4),
        "disponible":        stock_disponible > 0,
        "prestamos":         prestamos_equipo,
    }


def limpiar_duplicados() -> int:
    conn = DatabaseConnection.get_instance(DB_PATH).get_connection()
    try:
        c = conn.cursor()
        c.execute("""
            DELETE FROM movimientos
            WHERE id NOT IN (
                SELECT MIN(id)
                FROM movimientos
                GROUP BY nombre, material, sku, cantidad, tipo, fecha
            )
        """)
        eliminados = c.rowcount
        conn.commit()
        if eliminados:
            log_auditoria("LIMPIAR_DUPLICADOS", f"{eliminados} duplicados eliminados")
        return eliminados
    except Exception:
        return 0


def generar_resumen_diario() -> str:
    hoy = datetime.now().strftime("%d/%m/%Y")
    movs_hoy = int(db_query(
        "SELECT COUNT(*) FROM movimientos WHERE fecha LIKE ? || '%'",
        (hoy,), fetchall=False
    )[0] or 0)
    ent_hoy = int(db_query(
        "SELECT COUNT(*) FROM movimientos WHERE tipo='ENTRADA' AND fecha LIKE ? || '%'",
        (hoy,), fetchall=False
    )[0] or 0)
    sal_hoy = int(db_query(
        "SELECT COUNT(*) FROM movimientos WHERE tipo='SALIDA' AND fecha LIKE ? || '%'",
        (hoy,), fetchall=False
    )[0] or 0)

    prestamos = obtener_equipos_en_prestamo()
    vencidos   = [p for p in prestamos if p["estado"] == "VENCIDO"]
    por_vencer = [p for p in prestamos if p["estado"] == "POR_VENCER"]
    alertas    = obtener_alertas_activas()

    lineas = [
        f"=== RESUMEN DIARIO - {hoy} ===",
        "",
        "ACTIVIDAD HOY:",
        f"  * Movimientos totales : {movs_hoy}",
        f"  * Salidas (alquileres): {sal_hoy}",
        f"  * Retornos            : {ent_hoy}",
        "",
        "ESTADO DE PRESTAMOS:",
        f"  * Equipos en prestamo : {len(prestamos)}",
        f"  * Vencidos            : {len(vencidos)}",
        f"  * Proximos a vencer   : {len(por_vencer)}",
        "",
    ]
    if vencidos:
        lineas.append("VENCIDOS (requieren atencion):")
        for p in vencidos:
            lineas.append(
                f"  ! {p['material']} - {p['responsable']} "
                f"({p['dias_fuera']} dias fuera, acordado: {p['dias_acordados']})"
            )
        lineas.append("")
    if alertas:
        lineas.append("ALERTAS DE STOCK BAJO:")
        for mat, disp, minimo in alertas:
            lineas.append(f"  ! {mat}: {disp} disponibles (minimo: {minimo})")
        lineas.append("")
    lineas.append(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    lineas.append(f"- INVENTARIO PRO v{VERSION_ACTUAL}")
    return "\n".join(lineas)




# === services/sync.py ===
import json
import uuid
from datetime import datetime


def obtener_dispositivo_id() -> str:
    row = db_query("SELECT valor FROM config_app WHERE clave = 'dispositivo_id'", fetchall=False)
    if row:
        return row[0]
    new_id = uuid.uuid4().hex[:12].upper()
    db_execute("INSERT INTO config_app (clave, valor) VALUES ('dispositivo_id', ?) ON CONFLICT(clave) DO UPDATE SET valor = ?",
               (new_id, new_id))
    return new_id


def obtener_ultima_sincronizacion() -> str:
    row = db_query("SELECT valor FROM config_app WHERE clave = 'ultima_sync'", fetchall=False)
    return row[0] if row else ""


def guardar_ultima_sincronizacion(fecha: str) -> None:
    db_execute("INSERT INTO config_app (clave, valor) VALUES ('ultima_sync', ?) ON CONFLICT(clave) DO UPDATE SET valor = ?",
               (fecha, fecha))


def exportar_cambios() -> dict:
    ultima = obtener_ultima_sincronizacion()
    device_id = obtener_dispositivo_id()
    cambios = {
        "dispositivo": device_id,
        "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "movimientos": [],
    }
    if ultima:
        rows = db_query(
            "SELECT id, nombre, material, sku, cantidad, tipo, fecha, stock_registro, dias, retorno, notas, ubicacion FROM movimientos WHERE id > (SELECT COALESCE(MAX(registro_id), 0) FROM sync_log WHERE tabla='movimientos' AND dispositivo_id != ?)",
            (device_id,)
        )
    else:
        rows = db_query(
            "SELECT id, nombre, material, sku, cantidad, tipo, fecha, stock_registro, dias, retorno, notas, ubicacion FROM movimientos ORDER BY id DESC LIMIT 500"
        )
    for r in rows:
        cambios["movimientos"].append({
            "id": r[0], "nombre": r[1], "material": r[2], "sku": r[3],
            "cantidad": r[4], "tipo": r[5], "fecha": r[6],
            "stock_registro": r[7], "dias": r[8], "retorno": r[9],
            "notas": r[10], "ubicacion": r[11]
        })
    return cambios


def importar_cambios(datos: dict) -> dict:
    if not datos or "movimientos" not in datos:
        return {"importados": 0, "omitidos": 0}

    device_origen = datos.get("dispositivo", "")
    device_local = obtener_dispositivo_id()
    if device_origen == device_local:
        return {"importados": 0, "omitidos": len(datos["movimientos"]), "motivo": "mismo dispositivo"}

    importados = 0
    omitidos = 0
    for mov in datos["movimientos"]:
        try:
            existe = db_query(
                "SELECT id FROM movimientos WHERE nombre=? AND material=? AND sku=? AND cantidad=? AND tipo=? AND fecha=?",
                (mov["nombre"], mov["material"], mov["sku"], mov["cantidad"], mov["tipo"], mov["fecha"]),
                fetchall=False
            )
            if existe:
                omitidos += 1
                continue

            _, stock_actual = obtener_stock_material(mov["material"])
            nuevo_stock = round(stock_actual + (mov["cantidad"] if mov["tipo"] == "ENTRADA" else -mov["cantidad"]), 4)

            db_execute(
                "INSERT INTO movimientos (nombre, material, sku, cantidad, tipo, fecha, stock_registro, dias, retorno, notas, ubicacion, dispositivo_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                (mov["nombre"], mov["material"], mov["sku"], mov["cantidad"], mov["tipo"], mov["fecha"],
                 nuevo_stock, mov["dias"], mov["retorno"], mov["notas"], mov["ubicacion"], device_origen)
            )
            importados += 1
        except Exception as e:
            logger.error("Error importando sync: %s", e)
            omitidos += 1

    if importados > 0:
        guardar_ultima_sincronizacion(datos.get("timestamp", ""))
        log_sync(device_origen, importados)

    return {"importados": importados, "omitidos": omitidos}


def log_sync(dispositivo: str, insertados: int) -> None:
    try:
        log_auditoria("SYNC", f"Desde {dispositivo}: +{insertados}")
    except Exception:
        pass


# === services/scheduler.py ===
import threading
from datetime import datetime, timedelta
from kivy.clock import Clock


def obtener_config(clave, default=""):
    row = db_query("SELECT valor FROM config_app WHERE clave = ?", (clave,), fetchall=False)
    return row[0] if row else default


def guardar_config(clave, valor):
    db_execute("INSERT INTO config_app (clave, valor) VALUES (?,?) ON CONFLICT(clave) DO UPDATE SET valor = ?",
               (clave, valor, valor))


class ProgramadorTareas:
    def __init__(self, app):
        self.app = app
        self._backup_interval = None

    def iniciar(self):
        self._programar_backup()
        self._programar_purga()
        self._programar_resumen_diario()

    def _programar_backup(self):
        intervalo = obtener_config("backup_interval_horas", "0")
        try:
            horas = int(intervalo)
        except ValueError:
            horas = 0
        if horas <= 0:
            return
        self._backup_interval = Clock.schedule_interval(
            lambda dt: self._ejecutar_backup(), horas * 3600
        )
        logger.info("Backup automatico cada %d horas", horas)

    def _ejecutar_backup(self):
        try:
            nombre = ruta_exportacion(
                f"backup_auto_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            generar_excel(nombre)
            log_auditoria("BACKUP_AUTO", nombre)
        except Exception as e:
            logger.error("Error en backup automatico: %s", e)

    def _programar_purga(self):
        dias_str = obtener_config("purga_automatica_dias", "0")
        try:
            dias = int(dias_str)
        except ValueError:
            dias = 0
        if dias <= 0:
            return
        Clock.schedule_interval(lambda dt: self._ejecutar_purga(dias), 86400)

    def _ejecutar_purga(self, dias):
        try:
            limite = (datetime.now() - timedelta(days=dias)).strftime("%d/%m/%Y")
            db_execute("DELETE FROM movimientos WHERE fecha < ?", (limite,))
            log_auditoria("PURGA_AUTO", f"Registros anteriores a {limite}")
        except Exception as e:
            logger.error("Error en purga automatica: %s", e)

    def _programar_resumen_diario(self):
        hora_str = obtener_config("resumen_diario_hora", "")
        if not hora_str:
            return
        try:
            h, m = hora_str.split(":")
            ahora = datetime.now()
            objetivo = ahora.replace(hour=int(h), minute=int(m), second=0)
            if objetivo <= ahora:
                objetivo += timedelta(days=1)
            segundos = (objetivo - ahora).total_seconds()
            Clock.schedule_once(lambda dt: self._enviar_resumen(), segundos)
            Clock.schedule_interval(lambda dt: self._enviar_resumen(), 86400)
        except Exception as e:
            logger.error("Error programando resumen: %s", e)

    def _enviar_resumen(self):
        try:
            exito, msg = enviar_reporte_por_correo()
            if exito:
                log_auditoria("RESUMEN_DIARIO", "Enviado automaticamente")
        except Exception as e:
            logger.error("Error enviando resumen diario: %s", e)


# === services/excel_service.py ===
import os
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def ruta_exportacion(nombre_archivo: str) -> str:
    try:
        os.makedirs(EXPORT_DIR, exist_ok=True)
    except Exception:
        pass
    return os.path.join(EXPORT_DIR, nombre_archivo)


def generar_excel(ruta: str = None) -> str:
    if ruta is None:
        ruta = ruta_exportacion("Reporte_Inventario.xlsx")

    wb = Workbook()
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1E3A5F")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    entrada_fill = PatternFill("solid", fgColor="D4EDDA")
    salida_fill  = PatternFill("solid", fgColor="F8D7DA")
    alerta_fill  = PatternFill("solid", fgColor="FFF3CD")

    ws1 = wb.active
    ws1.title = "Movimientos"
    ws1.freeze_panes = "A2"
    headers = ["#","Responsable","Material","SKU","Cantidad","Tipo","Fecha",
               "Stock Registrado","Dias","Retorno","Notas","Ubicacion"]
    ws1.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center_align; cell.border = thin_border

    movimientos = db_query(
        "SELECT id,nombre,material,sku,cantidad,tipo,fecha,stock_registro,dias,retorno,notas,ubicacion "
        "FROM movimientos ORDER BY id DESC"
    )
    for row_data in movimientos:
        ws1.append(tuple(row_data))
        row_idx = ws1.max_row
        fill = entrada_fill if row_data[5] == "ENTRADA" else salida_fill
        for col_idx in range(1, len(headers)+1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.fill = fill; cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    widths = [5,20,20,12,10,10,18,15,8,10,25,15]
    for col_idx, width in enumerate(widths, 1):
        ws1.column_dimensions[chr(64+col_idx)].width = width

    ws2 = wb.create_sheet("Stock Actual")
    ws2.freeze_panes = "A2"
    ws2.append(["Material","Stock Total Sistema","Stock Disponible Estante","Estado"])
    for col_idx in range(1, 5):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center_align; cell.border = thin_border
    materiales   = db_query("SELECT DISTINCT UPPER(material) FROM movimientos ORDER BY material")
    alertas_conf = {r[0].upper(): r[1] for r in db_query("SELECT material, stock_minimo FROM alertas_stock")}
    for (mat,) in materiales:
        fijo, disponible = obtener_stock_material(mat)
        minimo = alertas_conf.get(mat, 5)
        if disponible <= 0:        estado = "SIN STOCK"
        elif disponible <= minimo: estado = "STOCK BAJO"
        else:                      estado = "OK"
        ws2.append([mat, fijo, disponible, estado])
        row_idx  = ws2.max_row
        fill_row = alerta_fill if "BAJO" in estado or "SIN" in estado else PatternFill("solid", fgColor="D4EDDA")
        for col_idx in range(1, 5):
            ws2.cell(row=row_idx, column=col_idx).fill = fill_row
            ws2.cell(row=row_idx, column=col_idx).border = thin_border
    for col_idx, width in enumerate([25,20,22,18], 1):
        ws2.column_dimensions[chr(64+col_idx)].width = width

    ws3 = wb.create_sheet("Limpieza")
    ws3.append(["Material","Cantidad","Estado","Fecha","Notas"])
    for col_idx in range(1, 6):
        cell = ws3.cell(row=1, column=col_idx)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center_align; cell.border = thin_border
    for row_data in db_query("SELECT material,cantidad,estado,fecha,notas FROM limpieza ORDER BY id DESC"):
        ws3.append(tuple(row_data))
        row_idx = ws3.max_row
        f = PatternFill("solid", fgColor="D4EDDA") if row_data[2] == "LIMPIO" else PatternFill("solid", fgColor="FFF3CD")
        for col_idx in range(1, 6):
            ws3.cell(row=row_idx, column=col_idx).fill = f
            ws3.cell(row=row_idx, column=col_idx).border = thin_border

    ws4 = wb.create_sheet("Estadisticas")
    stats = obtener_reporte_estadisticas()
    ws4["A1"] = "RESUMEN EJECUTIVO"
    ws4["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws4["A3"] = "Indicador"; ws4["B3"] = "Valor"
    for cell in [ws4["A3"], ws4["B3"]]:
        cell.font = header_font; cell.fill = header_fill
    filas = [
        ("Total de Movimientos",       stats["total_movimientos"]),
        ("Entradas Registradas",        stats["entradas"]),
        ("Salidas Registradas",         stats["salidas"]),
        ("Materiales Distintos",        stats["materiales"]),
        ("Responsables Activos",        stats["responsables"]),
        ("Material Mas Movido",         stats["material_top"]),
        ("Movimientos (ultimos 7 dias)",stats["movs_7dias"]),
        ("Reporte Generado",            datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    for i, (indicador, valor) in enumerate(filas, start=4):
        ws4.cell(row=i, column=1, value=indicador).border = thin_border
        ws4.cell(row=i, column=2, value=valor).border     = thin_border
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 20
    wb.save(ruta)
    return ruta




# === services/correo.py ===
import json
import smtplib
import imaplib
import email
import os
import threading
from email.header import decode_header
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime, timedelta




def enviar_alerta_vencimientos() -> tuple:
    venc = obtener_vencimientos()
    vencidos   = venc["vencidos"]
    por_vencer = venc["por_vencer"]

    if not vencidos and not por_vencer:
        return True, "Sin vencimientos - no se envio correo."

    cfg = EMAIL_CONFIG
    if not cfg.get("password", "").strip():
        return False, "Sin contrasena configurada en APP_EMAIL_PASS."

    now = datetime.now()
    lineas = [
        f"ALERTA DE VENCIMIENTOS - {now.strftime('%d/%m/%Y %H:%M')}",
        "",
    ]
    if vencidos:
        lineas.append(f"EQUIPOS CON RETORNO VENCIDO ({len(vencidos)}):")
        for p in vencidos:
            lineas.append(
                f"  * {p['material']} - Responsable: {p['responsable']}"
                f" | Dias fuera: {p['dias_fuera']} (acordado: {p['dias_acordados']})"
                f" | Salida: {p['fecha_salida']}"
            )
        lineas.append("")
    if por_vencer:
        lineas.append(f"PROXIMOS A VENCER ({len(por_vencer)}):")
        for p in por_vencer:
            lineas.append(
                f"  * {p['material']} - Responsable: {p['responsable']}"
                f" | Dias fuera: {p['dias_fuera']} (acordado: {p['dias_acordados']})"
            )
        lineas.append("")
    lineas.append(f"- INVENTARIO PRO v{VERSION_ACTUAL}")
    cuerpo = "\n".join(lineas)

    try:
        msg = MIMEMultipart()
        msg['From']    = cfg["remitente"]
        msg['To']      = ", ".join(cfg["destinatarios"])
        msg['Subject'] = (
            f"ALERTA ALQUILER: {len(vencidos)} vencido(s), "
            f"{len(por_vencer)} por vencer - {now.strftime('%d/%m/%Y')}"
        )
        msg.attach(MIMEText(cuerpo, 'plain', 'utf-8'))
        s = smtplib.SMTP('smtp.gmail.com', 587)
        s.starttls()
        s.login(cfg["remitente"], cfg["password"])
        for dest in cfg["destinatarios"]:
            s.sendmail(cfg["remitente"], dest, msg.as_string())
        s.quit()
        log_auditoria("ALERTA_VENCIMIENTOS_EMAIL",
                      f"{len(vencidos)} vencidos, {len(por_vencer)} por vencer")
        return True, "Alerta enviada correctamente."
    except Exception as e:
        return False, f"Error al enviar: {e}"


def enviar_reporte_por_correo() -> tuple:
    EXCEL_TEMP = ruta_exportacion(f"Reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    try:
        generar_excel(EXCEL_TEMP)
    except Exception as e:
        logger.error("Error generando Excel para correo: %s", e, exc_info=True)
        return False, f"Error al generar Excel: {e}"
    try:
        msg = MIMEMultipart()
        cfg = EMAIL_CONFIG
        if not cfg.get("remitente") or not cfg.get("password"):
            return False, "Faltan credenciales de correo en .env (APP_EMAIL, APP_EMAIL_PASS)"
        msg['From']    = cfg["remitente"]
        msg['To']      = ", ".join(cfg["destinatarios"])
        msg['Subject'] = f"INVENTARIO PRO - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        stats = obtener_reporte_estadisticas()
        body  = (f"Reporte automatico generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}.\n\n"
                 f"Resumen:\n"
                 f"  * Total movimientos: {stats['total_movimientos']}\n"
                 f"  * Entradas: {stats['entradas']} | Salidas: {stats['salidas']}\n"
                 f"  * Materiales distintos: {stats['materiales']}\n"
                 f"  * Material mas activo: {stats['material_top']}\n\n"
                 f"Ver detalle en el archivo adjunto.\n\n")
        body += f"- INVENTARIO PRO v{VERSION_ACTUAL}"
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        with open(EXCEL_TEMP, "rb") as f:
            parte = MIMEBase('application', 'octet-stream')
            parte.set_payload(f.read())
            encoders.encode_base64(parte)
            parte.add_header('Content-Disposition', f'attachment; filename={os.path.basename(EXCEL_TEMP)}')
            msg.attach(parte)
        s = smtplib.SMTP('smtp.gmail.com', 587)
        s.starttls()
        s.login(cfg["remitente"], cfg["password"])
        for dest in cfg["destinatarios"]:
            s.sendmail(cfg["remitente"], dest, msg.as_string())
        s.quit()
        log_auditoria("EMAIL_ENVIADO", f"Destinatarios: {cfg['destinatarios']}")
        return True, "Reporte enviado correctamente."
    except smtplib.SMTPAuthenticationError:
        return False, ("Error de autenticacion SMTP. Si usas Gmail, necesitas una "
                       "Contrasena de Aplicacion (no la contrasena normal).")
    except smtplib.SMTPException as e:
        logger.error("Error SMTP: %s", e, exc_info=True)
        return False, f"Error SMTP: {e}"
    except Exception as e:
        logger.error("Error al enviar correo: %s", e, exc_info=True)
        return False, f"Error al enviar: {e}"


def enviar_sync_email() -> tuple:
    try:
        cfg = EMAIL_CONFIG
        if not cfg.get("remitente") or not cfg.get("password"):
            return False, "Sin credenciales"
        cambios = exportar_cambios()
        if not cambios.get("movimientos"):
            return True, "Sin cambios para sincronizar"
        payload = json.dumps(cambios, ensure_ascii=False, default=str)
        msg = MIMEMultipart()
        msg['From'] = cfg["remitente"]
        msg['To'] = cfg["remitente"]
        device = obtener_dispositivo_id()
        msg['Subject'] = f"[SYNC] {device} - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        parte = MIMEText(payload, 'plain', 'utf-8')
        parte.add_header('Content-Disposition', 'attachment; filename="sync.json"')
        msg.attach(parte)
        s = smtplib.SMTP('smtp.gmail.com', 587)
        s.starttls()
        s.login(cfg["remitente"], cfg["password"])
        s.sendmail(cfg["remitente"], cfg["remitente"], msg.as_string())
        s.quit()
        log_auditoria("SYNC_ENVIADO", f"{len(cambios['movimientos'])} registros")
        return True, f"Sync enviado: {len(cambios['movimientos'])} registros"
    except Exception as e:
        logger.error("Error enviando sync: %s", e)
        return False, str(e)


def _procesar_sync_email(msg) -> dict:
    for parte in msg.walk():
        filename = parte.get_filename()
        if filename and "sync.json" in filename.lower():
            payload = parte.get_payload(decode=True)
            if payload:
                try:
                    datos = json.loads(payload.decode('utf-8'))
                    return datos
                except Exception:
                    pass
    return {}


class ActualizadorCorreo:
    EXTENSIONES = ('.xlsx', '.xls', '.xlsm')

    def __init__(self, callback_progreso, callback_fin):
        self.cb_prog = callback_progreso
        self.cb_fin  = callback_fin

    def ejecutar(self) -> None:
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self) -> None:
        self._prog("Conectando al servidor de correo...")
        try:
            mail = imaplib.IMAP4_SSL(EMAIL_CONFIG["imap_servidor"], EMAIL_CONFIG["imap_puerto"])
        except Exception as e:
            self._fin(False, {"error": f"No se pudo conectar al servidor IMAP: {e}"})
            return

        self._prog("Iniciando sesion...")
        if not str(EMAIL_CONFIG.get("password") or "").strip():
            try: mail.logout()
            except Exception: pass
            self._fin(False, {"error": "Falta la contrasena de correo."})
            return

        try:
            mail.login(EMAIL_CONFIG["remitente"], EMAIL_CONFIG["password"])
        except imaplib.IMAP4.error as e:
            try: mail.logout()
            except Exception: pass
            self._fin(False, {"error": f"Error de autenticacion: {e}"})
            return

        self._prog("Buscando correos con archivos Excel...")
        try:
            adjunto, meta = self._buscar_adjunto_excel(mail)
        except Exception as e:
            try: mail.logout()
            except Exception: pass
            self._fin(False, {"error": f"Error al buscar correos: {e}"})
            return

        if adjunto is None:
            try: mail.logout()
            except Exception: pass
            self._fin(False, {"error": "No se encontro ningun correo con un adjunto Excel."})
            return

        nombre_archivo = meta["nombre_archivo"]
        self._prog(f"Descargando: {nombre_archivo}...")
        ruta_local = ruta_exportacion(f"correo_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{nombre_archivo}")
        try:
            with open(ruta_local, "wb") as f:
                f.write(adjunto)
        except Exception as e:
            try: mail.logout()
            except Exception: pass
            self._fin(False, {"error": f"No se pudo guardar el archivo: {e}"})
            return

        self._prog("Importando datos al inventario...")
        try:
            resultado = self._importar_excel(ruta_local, meta)
        except Exception as e:
            try: mail.logout()
            except Exception: pass
            self._fin(False, {"error": f"Error al importar el Excel: {e}"})
            return

        try:
            mail.select("INBOX")
            mail.store(meta["uid"], '+FLAGS', '\\Seen')
        except Exception: pass
        try: mail.logout()
        except Exception: pass

        try:
            db_execute(
                "INSERT INTO actualizaciones_correo "
                "(fecha, remitente, asunto, archivo, insertados, omitidos, errores) "
                "VALUES (?,?,?,?,?,?,?)",
                (datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                 meta.get("de", ""), meta.get("asunto", ""),
                 nombre_archivo, resultado["insertados"],
                 resultado["omitidos"], resultado["errores"])
            )
            log_auditoria("ACTUALIZAR_CORREO",
                          f"{nombre_archivo} | +{resultado['insertados']} registros")
        except Exception: pass

        self._fin(True, resultado)

    def _buscar_adjunto_excel(self, mail):
        mail.select("INBOX")

        def _armar_criterio(*grupos):
            partes = []
            for g in grupos:
                partes.extend(g)
            return "(" + " ".join(partes) + ")" if partes else "ALL"

        filtros = []
        if EMAIL_CONFIG.get("asunto_filtro"):
            filtros.append(f'SUBJECT "{EMAIL_CONFIG["asunto_filtro"]}"')
        if EMAIL_CONFIG.get("remitente_filtro"):
            filtros.append(f'FROM "{EMAIL_CONFIG["remitente_filtro"]}"')

        fecha_desde = (datetime.now() - timedelta(days=90)).strftime("%d-%b-%Y")

        # 1) buscar con filtros + ventana de tiempo
        criterio_str = _armar_criterio(filtros, [f'SINCE "{fecha_desde}"'])
        _, data = mail.search(None, criterio_str)
        uids = data[0].split()

        # 2) si no hay, reintentar sin ventana de tiempo
        if not uids and filtros:
            criterio_str = _armar_criterio(filtros)
            _, data = mail.search(None, criterio_str)
            uids = data[0].split()

        # 3) si aun no hay y no hay filtros, buscar todo
        if not uids and not filtros:
            criterio_str = "ALL"
            _, data = mail.search(None, criterio_str)
            uids = data[0].split()

        if not uids:
            return None, {}

        for uid in reversed(uids):
            _, msg_data = mail.fetch(uid, "(RFC822)")
            if not msg_data or not msg_data[0]:
                continue
            raw = msg_data[0][1]
            msg = email.message_from_bytes(raw)
            de      = self._decodificar_header(msg.get("From", ""))
            asunto  = self._decodificar_header(msg.get("Subject", ""))
            for parte in msg.walk():
                filename = parte.get_filename()
                if not filename:
                    continue
                nombre_dec = self._decodificar_header(filename)
                if not nombre_dec.lower().endswith(self.EXTENSIONES):
                    continue
                payload = parte.get_payload(decode=True)
                if not payload:
                    continue
                meta = {
                    "uid": uid, "de": de, "asunto": asunto,
                    "nombre_archivo": nombre_dec,
                }
                return payload, meta
        return None, {}

    @staticmethod
    def _importar_excel(ruta_local, meta):
        from openpyxl import load_workbook

        wb   = load_workbook(ruta_local, read_only=True, data_only=True)
        hoja = wb.sheetnames[0]
        ws   = wb[hoja]
        filas = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(filas) < 2:
            return {"insertados": 0, "omitidos": 0, "errores": 0,
                    "nota": "El archivo no tiene filas de datos"}

        enc_upper = [str(h).strip().upper() if h is not None else "" for h in filas[0]]
        mapeo_campos = {
            'nombre':   ['NOMBRE','RESPONSABLE','NAME','USUARIO','RESP'],
            'material': ['MATERIAL','PRODUCTO','ITEM','DESCRIPCION','DESCRIPTION','MAT'],
            'sku':      ['SKU','ID','CODIGO','CODE','REF'],
            'cantidad': ['CANTIDAD','QTY','QUANTITY','CANT'],
            'tipo':     ['TIPO','TYPE','MOVIMIENTO','MOVEMENT'],
            'fecha':    ['FECHA','DATE','DATETIME','TIMESTAMP'],
        }
        indices = {}
        for campo, alternativas in mapeo_campos.items():
            for alt in alternativas:
                if alt in enc_upper:
                    indices[campo] = enc_upper.index(alt)
                    break
        if 'material' not in indices or 'cantidad' not in indices:
            raise ValueError("No se encontraron las columnas 'Material' y 'Cantidad'.")

        conn = DatabaseConnection.get_instance(DB_PATH).get_connection()
        cursor = conn.cursor()
        insertados = errores = omitidos = 0
        ahora = datetime.now().strftime("%d/%m/%Y %H:%M")

        for fila in filas[1:]:
            try:
                i = indices
                nombre   = str(fila[i.get('nombre')] if i.get('nombre') is not None and i['nombre'] < len(fila) else "INVENTARIO").strip().upper()
                material = str(fila[i['material']] if i['material'] < len(fila) else "").strip().upper()
                sku      = str(fila[i.get('sku')] if i.get('sku') is not None and i['sku'] < len(fila) else "").strip()
                raw_cant = fila[i['cantidad']] if i['cantidad'] < len(fila) else None
                raw_tipo = str(fila[i.get('tipo')] if i.get('tipo') is not None and i['tipo'] < len(fila) else "ENTRADA").strip().upper()
                raw_fec  = fila[i.get('fecha')] if i.get('fecha') is not None and i['fecha'] < len(fila) else None

                if not material:
                    omitidos += 1; continue
                try:
                    cantidad = float(str(raw_cant).replace(',', '.'))
                except (TypeError, ValueError):
                    omitidos += 1; continue
                if cantidad <= 0:
                    omitidos += 1; continue

                if any(k in raw_tipo for k in ['ENT','IN','ING','COMP']):
                    tipo = "ENTRADA"
                elif any(k in raw_tipo for k in ['SAL','OUT','EGR','RET']):
                    tipo = "SALIDA"
                else:
                    tipo = "ENTRADA"

                if raw_fec is None:
                    fecha = ahora
                elif isinstance(raw_fec, datetime):
                    fecha = raw_fec.strftime("%d/%m/%Y %H:%M")
                else:
                    fecha = str(raw_fec).strip() or ahora

                cursor.execute(
                    "SELECT 1 FROM movimientos WHERE nombre=? AND material=? AND sku=? AND cantidad=? AND tipo=? AND fecha=? LIMIT 1",
                    (nombre, material, sku, cantidad, tipo, fecha)
                )
                if cursor.fetchone():
                    omitidos += 1; continue

                _, stock_actual = obtener_stock_material(material)
                nuevo_stock = round(stock_actual + (cantidad if tipo == "ENTRADA" else -cantidad), 4)
                cursor.execute(
                    "INSERT INTO movimientos (nombre,material,sku,cantidad,tipo,fecha,stock_registro,dias,retorno,notas,ubicacion) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    (nombre, material, sku, cantidad, tipo, fecha, nuevo_stock, 0, "N/A",
                     "", "")
                )
                insertados += 1
            except Exception:
                errores += 1

        conn.commit()
        return {"insertados": insertados, "omitidos": omitidos, "errores": errores,
                "archivo": meta.get("nombre_archivo", ""), "de": meta.get("de", ""),
                "asunto": meta.get("asunto", "")}

    @staticmethod
    def _decodificar_header(valor):
        if not valor:
            return ""
        partes = decode_header(valor)
        resultado = []
        for parte, enc in partes:
            if isinstance(parte, bytes):
                resultado.append(parte.decode(enc or "utf-8", errors="replace"))
            else:
                resultado.append(str(parte))
        return "".join(resultado)

    def _prog(self, texto):
        from kivy.clock import Clock
        Clock.schedule_once(lambda dt: self.cb_prog(texto))

    def _fin(self, exito, datos):
        from kivy.clock import Clock
        Clock.schedule_once(lambda dt: self.cb_fin(exito, datos))


class AutoActualizador:
    EXTENSIONES = ('.xlsx', '.xls', '.xlsm')

    def __init__(self, callback_fin):
        self.cb_fin = callback_fin

    def ejecutar(self):
        import threading
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self):
        try:
            resultado = self._revisar_e_importar()
            sync_result = self._revisar_sync()
            total = resultado.get("importados", 0) + sync_result.get("importados", 0)
            resultado["importados"] = total
            resultado["sync"] = sync_result.get("importados", 0)
            from kivy.clock import Clock
            Clock.schedule_once(lambda dt: self.cb_fin(True, resultado))
        except Exception as ex:
            err = str(ex)
            from kivy.clock import Clock
            Clock.schedule_once(lambda dt, e=err: self.cb_fin(False, {"error": e}))

    def _revisar_sync(self):
        sync_count = 0
        try:
            mail = imaplib.IMAP4_SSL(EMAIL_CONFIG["imap_servidor"], EMAIL_CONFIG["imap_puerto"])
            mail.login(EMAIL_CONFIG["remitente"], EMAIL_CONFIG["password"])
            mail.select("INBOX")
            _, data = mail.search(None, 'UNSEEN', 'SUBJECT "[SYNC]"')
            uids = data[0].split() if data[0] else []
            for uid in reversed(uids):
                _, msg_data = mail.fetch(uid, "(RFC822)")
                if not msg_data or not msg_data[0]:
                    continue
                raw = msg_data[0][1]
                msg = email.message_from_bytes(raw)
                datos = _procesar_sync_email(msg)
                if datos:
                    res = importar_cambios(datos)
                    sync_count += res.get("importados", 0)
                mail.select("INBOX")
                mail.store(uid, '+FLAGS', '\\Seen')
            mail.logout()
        except Exception as e:
            logger.error("Error en sync check: %s", e)
        return {"importados": sync_count}

    def _revisar_e_importar(self):
        mail = imaplib.IMAP4_SSL(EMAIL_CONFIG["imap_servidor"], EMAIL_CONFIG["imap_puerto"])
        mail.login(EMAIL_CONFIG["remitente"], EMAIL_CONFIG["password"])
        mail.select("INBOX")

        criterios = ["UNSEEN"]
        if EMAIL_CONFIG.get("asunto_filtro"):
            criterios.append(f'SUBJECT "{EMAIL_CONFIG["asunto_filtro"]}"')
        if EMAIL_CONFIG.get("remitente_filtro"):
            criterios.append(f'FROM "{EMAIL_CONFIG["remitente_filtro"]}"')

        _, data = mail.search(None, *criterios)
        uids = data[0].split() if data[0] else []

        if not uids:
            mail.logout()
            return {"importados": 0, "mensaje": "Sin correos nuevos"}

        total_importados = 0
        for uid in reversed(uids):
            _, msg_data = mail.fetch(uid, "(RFC822)")
            if not msg_data or not msg_data[0]:
                continue
            raw = msg_data[0][1]
            msg = email.message_from_bytes(raw)
            de = ActualizadorCorreo._decodificar_header(None, msg.get("From", ""))
            asunto = ActualizadorCorreo._decodificar_header(None, msg.get("Subject", ""))

            for parte in msg.walk():
                filename = parte.get_filename()
                if not filename:
                    continue
                nombre_dec = ActualizadorCorreo._decodificar_header(filename)
                if not nombre_dec.lower().endswith(self.EXTENSIONES):
                    continue
                payload = parte.get_payload(decode=True)
                if not payload:
                    continue

                meta = {"uid": uid, "de": de, "asunto": asunto, "nombre_archivo": nombre_dec}
                ruta_local = ruta_exportacion(f"auto_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{nombre_dec}")
                with open(ruta_local, "wb") as f:
                    f.write(payload)

                resultado = ActualizadorCorreo._importar_excel(ruta_local, meta)
                total_importados += resultado.get("insertados", 0)

                mail.select("INBOX")
                mail.store(uid, '+FLAGS', '\\Seen')

                db_execute(
                    "INSERT INTO actualizaciones_correo "
                    "(fecha, remitente, asunto, archivo, insertados, omitidos, errores) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                     de, asunto, nombre_dec,
                     resultado.get("insertados", 0),
                     resultado.get("omitidos", 0),
                     resultado.get("errores", 0))
                )
                log_auditoria("AUTO_ACTUALIZAR_CORREO",
                              f"{nombre_dec} | +{resultado.get('insertados', 0)} registros")

        mail.logout()
        return {"importados": total_importados}


# === services/reporting.py ===
import os
from datetime import datetime
from fpdf import FPDF


PDF_STYLES = {
    "bg_dark": (15, 23, 42),
    "bg_card": (30, 41, 59),
    "primary": (56, 189, 248),
    "success": (16, 185, 129),
    "danger": (248, 113, 113),
    "warning": (245, 158, 11),
    "text": (203, 213, 225),
    "text_dim": (100, 116, 139),
    "white": (248, 250, 252),
}


class PDFReport(FPDF):
    def header(self):
        self.set_fill_color(*PDF_STYLES["bg_dark"])
        self.rect(0, 0, 210, 15, 'F')
        self.set_text_color(*PDF_STYLES["primary"])
        self.set_font("Helvetica", "B", 10)
        self.cell(0, 10, "INVENTARIO PRO - Reporte", align='C', new_x="LMARGIN", new_y="NEXT")
        self.ln(3)

    def footer(self):
        self.set_y(-15)
        self.set_fill_color(*PDF_STYLES["bg_dark"])
        self.rect(0, self.get_y(), 210, 15, 'F')
        self.set_text_color(*PDF_STYLES["text_dim"])
        self.set_font("Helvetica", "I", 8)
        self.cell(0, 10, f"Pagina {self.page_no()}/{{nb}}", align='C')

    def section_title(self, title, color=None):
        if color is None:
            color = PDF_STYLES["primary"]
        self.set_fill_color(*color)
        self.set_text_color(*PDF_STYLES["white"])
        self.set_font("Helvetica", "B", 12)
        self.cell(0, 10, f"  {title}", fill=True, new_x="LMARGIN", new_y="NEXT")
        self.ln(3)

    def card(self, lines, col1_width=80):
        self.set_fill_color(*PDF_STYLES["bg_card"])
        x = self.get_x()
        y = self.get_y()
        line_h = 7
        h = len(lines) * line_h + 6
        self.rect(x, y, 190, h, 'F')
        self.set_xy(x + 4, y + 3)
        self.set_text_color(*PDF_STYLES["text"])
        self.set_font("Helvetica", "", 10)
        for label, value in lines:
            self.set_font("Helvetica", "B", 10)
            self.cell(col1_width, line_h, f"{label}:")
            self.set_font("Helvetica", "", 10)
            self.cell(0, line_h, str(value), new_x="LMARGIN", new_y="NEXT")
            self.set_x(x + 4)
        self.set_y(y + h + 4)


def generar_pdf() -> str:
    ruta = ruta_exportacion(f"reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
    pdf = PDFReport()
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    stats = obtener_reporte_estadisticas()
    alertas = obtener_alertas_activas()

    pdf.section_title("RESUMEN EJECUTIVO", PDF_STYLES["warning"])
    pdf.card([
        ("Total Movimientos", stats["total_movimientos"]),
        ("Entradas", stats["entradas"]),
        ("Salidas", stats["salidas"]),
        ("Materiales", stats["materiales"]),
        ("Responsables", stats["responsables"]),
        ("Material mas activo", stats["material_top"]),
    ])

    pdf.section_title("VOLUMENES", PDF_STYLES["success"])
    pdf.card([
        ("Vol. Entradas", stats["vol_entradas"]),
        ("Vol. Salidas", stats["vol_salidas"]),
        ("Balance Neto", round(stats["vol_entradas"] - stats["vol_salidas"], 2)),
        ("Ratio E/S", round(stats["entradas"] / max(stats["salidas"], 1), 2)),
    ])

    if alertas:
        pdf.section_title(f"ALERTAS DE STOCK ({len(alertas)})", PDF_STYLES["danger"])
        pdf.card([(mat, f"Disp: {disp} / Min: {minimo}") for mat, disp, minimo in alertas])

    if stats["top5_materiales"]:
        pdf.section_title("TOP 5 MATERIALES", PDF_STYLES["primary"])
        pdf.card([(m, str(cnt)) for m, cnt in stats["top5_materiales"]])

    if stats["mats_inactivos"]:
        pdf.section_title("MATERIALES INACTIVOS", PDF_STYLES["text_dim"])
        pdf.card([(m, "+30d sin mov") for m in stats["mats_inactivos"][:10]])

    pdf.set_text_color(*PDF_STYLES["text_dim"])
    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(0, 10, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", align='C')

    pdf.output(ruta)
    log_auditoria("PDF_EXPORT", ruta)
    return ruta


# === widgets/tarjeta.py ===
from kivy.graphics import Color, RoundedRectangle
from kivy.uix.boxlayout import BoxLayout
from kivy.metrics import dp
from kivy.utils import get_color_from_hex


def crear_tarjeta(altura=dp(90), color='#0D1B2A', radius=dp(4)) -> BoxLayout:
    card = BoxLayout(orientation='vertical', size_hint_y=None, height=altura, padding=dp(10))
    with card.canvas.before:
        Color(rgba=get_color_from_hex(color))
        card._rect = RoundedRectangle(pos=card.pos, size=card.size, radius=[radius])

    def _upd(inst, val):
        inst._rect.pos  = inst.pos
        inst._rect.size = inst.size

    card.bind(pos=_upd, size=_upd)
    return card


# === widgets/toast.py ===
from kivy.uix.label import Label
from kivy.uix.boxlayout import BoxLayout
from kivy.clock import Clock
from kivy.metrics import dp
from kivy.utils import get_color_from_hex
from kivy.core.window import Window
from kivy.animation import Animation


class Toast(BoxLayout):
    _instance = None

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'vertical'
        self.size_hint = (None, None)
        self.pos_hint = {'center_x': 0.5}
        self.y = -dp(100)
        self.opacity = 0
        self._label = Label(
            text="", font_size='13sp',
            color=get_color_from_hex('#F8FAFC'),
            halign='center', valign='middle',
            padding=(dp(20), dp(12)),
            text_size=(Window.width * 0.85, None),
            size_hint=(None, None),
        )
        self._label.bind(texture_size=self._update_size)
        self.add_widget(self._label)

    def _update_size(self, inst, val):
        self.size = (min(val[0] + dp(40), Window.width * 0.9), val[1] + dp(24))

    @classmethod
    def show(cls, mensaje, color='#0066FF', duracion=3.0):
        if cls._instance is None:
            cls._instance = cls()
        toast = cls._instance
        if toast.parent:
            toast.parent.remove_widget(toast)
        toast._label.text = mensaje
        with toast.canvas.before:
            from kivy.graphics import Color, RoundedRectangle
            Color(rgba=get_color_from_hex(color))
            toast._rect = RoundedRectangle(pos=toast.pos, size=toast.size, radius=[dp(10)])
        toast.bind(pos=lambda i, v: setattr(toast._rect, 'pos', v),
                   size=lambda i, v: setattr(toast._rect, 'size', v))
        Window.add_widget(toast)
        anim = Animation(y=dp(30), opacity=1, duration=0.3, t='out_back')
        anim.bind(on_complete=lambda *a: Clock.schedule_once(
            lambda dt: Animation(y=-dp(100), opacity=0, duration=0.3).start(toast), duracion))
        anim.start(toast)


# === widgets/charts.py ===
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.widget import Widget
from kivy.graphics import Color, RoundedRectangle
from kivy.metrics import dp
from kivy.utils import get_color_from_hex
from kivy.core.window import Window


class BarraHorizontal(BoxLayout):
    def __init__(self, label, valor, maximo, color='#00D4FF', **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'vertical'
        self.size_hint_y = None
        self.height = dp(32)
        self.spacing = dp(2)

        pct = min(valor / maximo, 1.0) if maximo > 0 else 0

        header = BoxLayout(size_hint_y=None, height=dp(16))
        header.add_widget(Label(
            text=f"{label}:  {valor}",
            color=get_color_from_hex('#CBD5E1'),
            font_size='11sp', halign='left',
            text_size=(Window.width * 0.78, None),
            size_hint_y=None, height=dp(16)
        ))
        self.add_widget(header)

        bar_bg = BoxLayout(size_hint_y=None, height=dp(12))
        with bar_bg.canvas.before:
            Color(rgba=get_color_from_hex('#0F172A'))
            bar_bg._rect_bg = RoundedRectangle(pos=bar_bg.pos, size=bar_bg.size, radius=[dp(6)])
        bar_bg.bind(
            pos=lambda i, v: setattr(bar_bg._rect_bg, 'pos', v),
            size=lambda i, v: setattr(bar_bg._rect_bg, 'size', v)
        )

        bar_fill = Widget(size_hint_x=max(pct, 0.02), size_hint_y=1)
        with bar_fill.canvas.before:
            Color(rgba=get_color_from_hex(color))
            bar_fill._rect_fill = RoundedRectangle(pos=bar_fill.pos, size=bar_fill.size, radius=[dp(6)])
        bar_fill.bind(
            pos=lambda i, v: setattr(bar_fill._rect_fill, 'pos', v),
            size=lambda i, v: setattr(bar_fill._rect_fill, 'size', v)
        )
        bar_bg.add_widget(bar_fill)
        self.add_widget(bar_bg)


class GraficoTendencia(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'vertical'
        self.size_hint_y = None
        self.height = dp(110)
        self.spacing = dp(4)

    def mostrar(self, datos, color='#00D4FF'):
        self.clear_widgets()
        if not datos:
            self.add_widget(Label(text="Sin datos", color=get_color_from_hex('#64748B'),
                                  size_hint_y=None, height=dp(30)))
            return
        max_val = max(datos) if max(datos) > 0 else 1
        header = BoxLayout(size_hint_y=None, height=dp(18))
        header.add_widget(Label(text="Tendencia", color=get_color_from_hex('#94A3B8'),
                                font_size='11sp', halign='left',
                                text_size=(Window.width * 0.8, None)))
        self.add_widget(header)

        bars = BoxLayout(size_hint_y=None, height=dp(60), spacing=dp(3), padding=[0, dp(4)])
        for val in datos:
            pct = max(val / max_val, 0.05)
            bar = Widget(size_hint_x=None, width=dp(20), size_hint_y=pct,
                         pos_hint={'y': 0})
            from kivy.graphics import Color as GColor, RoundedRectangle as GRect
            with bar.canvas.before:
                GColor(rgba=get_color_from_hex(color))
                bar._rect = GRect(pos=bar.pos, size=bar.size, radius=[dp(4)])
            bar.bind(
                pos=lambda i, v: setattr(bar._rect, 'pos', v) if hasattr(bar, '_rect') else None,
                size=lambda i, v: setattr(bar._rect, 'size', v) if hasattr(bar, '_rect') else None
            )
            bars.add_widget(bar)
        self.add_widget(bars)

        labels = BoxLayout(size_hint_y=None, height=dp(18), spacing=dp(3))
        for i, val in enumerate(datos):
            labels.add_widget(Label(text=str(val), color=get_color_from_hex('#64748B'),
                                    font_size='9sp', halign='center'))
        self.add_widget(labels)


# === widgets/pagination.py ===
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.metrics import dp
from kivy.utils import get_color_from_hex


class PaginationBar(BoxLayout):
    def __init__(self, callback=None, page_size=20, **kwargs):
        super().__init__(**kwargs)
        self.callback = callback
        self.page_size = page_size
        self.total = 0
        self.current_page = 0
        self.total_pages = 0
        self.orientation = 'horizontal'
        self.size_hint_y = None
        self.height = dp(40)
        self.spacing = dp(6)
        self._build_ui()

    def _build_ui(self):
        self.clear_widgets()
        self._btn_prev = Button(
            text="<", size_hint_x=None, width=dp(48),
            background_color=get_color_from_hex('#334155'),
            font_size='15sp', bold=True,
            on_release=lambda x: self._ir_pagina(self.current_page - 1)
        )
        self.add_widget(self._btn_prev)

        self._lbl_info = Label(
            text="", color=get_color_from_hex('#94A3B8'),
            font_size='12sp', halign='center'
        )
        self.add_widget(self._lbl_info)

        self._btn_next = Button(
            text=">", size_hint_x=None, width=dp(48),
            background_color=get_color_from_hex('#334155'),
            font_size='15sp', bold=True,
            on_release=lambda x: self._ir_pagina(self.current_page + 1)
        )
        self.add_widget(self._btn_next)

    def actualizar(self, total):
        self.total = total
        self.total_pages = max(1, (total + self.page_size - 1) // self.page_size)
        if self.current_page >= self.total_pages:
            self.current_page = max(0, self.total_pages - 1)
        self._actualizar_ui(notificar=False)

    def _actualizar_ui(self, notificar=True):
        inicio = self.current_page * self.page_size + 1
        fin = min((self.current_page + 1) * self.page_size, self.total)
        self._lbl_info.text = f"{inicio}-{fin} / {self.total}" if self.total > 0 else "0"
        self._btn_prev.disabled = self.current_page <= 0
        self._btn_next.disabled = self.current_page >= self.total_pages - 1 or self.total_pages <= 1
        if notificar and self.callback:
            self.callback(self.current_page, self.page_size)

    def _ir_pagina(self, pagina):
        if pagina < 0 or pagina >= self.total_pages:
            return
        self.current_page = pagina
        self._actualizar_ui(notificar=True)


# === screens/inventario.py ===
import os
import sqlite3
import threading
import gc
from datetime import datetime
from kivy.app import App
from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.animation import Animation
from kivy.clock import Clock
from kivy.graphics import Color as GColor, Line as GLine
from kivy.utils import get_color_from_hex
from kivy.core.window import Window
from kivy.metrics import dp

try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False


class InventarioScreen(Screen):
    _ultimo_mov_id = None
    _ultimo_mov_tipo = None
    _menu_abierto = False

    def on_enter(self):
        self.actualizar_spinners()
        self._verificar_alertas_silencioso()
        self._actualizar_dashboard()
        try:
            self.ids.info_carpeta.text = f"Exportaciones: {EXPORT_DIR}"
        except Exception:
            pass
        gc.collect()

    MENU_ITEMS = [
        ("FUERA", 'pendientes', '#FF0044'),
        ("FILTROS", 'filtros', '#00D4FF'),
        ("GESTION", 'gestion', '#FF6600'),
        ("AUDITORIA", 'auditoria', '#8800FF'),
        ("CATEGORIAS", 'categorias', '#FF00FF'),
        ("STOCK", 'stock', '#00FF88'),
        ("LIMPIEZA", 'limpieza', '#00FFFF'),
        ("ESTADISTICAS", 'estadisticas', '#FF0088'),
        ("AJUSTES", 'ajustes', '#FFFF00'),
        ("VALOR", 'valor', '#00D4FF'),
        ("UBICACIONES", 'ubicaciones', '#00FFAA'),
        ("IMPORTAR", 'importar', '#0088FF'),
    ]
    MENU_ACCIONES = [
        ("EXPORTAR", 'mostrar_opciones_exportar', '#FFAA00'),
        ("BACKUP", 'hacer_backup', '#00FF88'),
        ("ACT. CORREO", 'actualizar_desde_correo', '#FF6600'),
        ("ALERTAS", 'alertas', '#FF0044'),
    ]

    def _construir_menu(self):
        grid = self.ids.menu_grid
        grid.clear_widgets()
        from kivy.uix.button import Button
        for item in self.MENU_ITEMS:
            txt = item[0]
            scr = item[1]
            color = item[2] if len(item) > 2 else '#00D4FF'
            btn = Button(
                text=txt, size_hint_y=None, height=dp(38),
                background_normal='', background_color=[0.051, 0.106, 0.165, 1],
                color=get_color_from_hex(color), font_size='12sp',
                halign='left', valign='middle',
                padding=[dp(10), 0], text_size=(Window.width * 0.65, None),
            )
            btn.bind(on_release=lambda x, s=scr: (setattr(self.manager, 'current', s), self.toggle_menu()))
            grid.add_widget(btn)
        for item in self.MENU_ACCIONES:
            txt = item[0]
            metodo = item[1]
            color = item[2] if len(item) > 2 else '#00D4FF'
            btn = Button(
                text=txt, size_hint_y=None, height=dp(38),
                background_normal='', background_color=[0.051, 0.106, 0.165, 1],
                color=get_color_from_hex(color), font_size='12sp',
                halign='left', valign='middle',
                padding=[dp(10), 0], text_size=(Window.width * 0.65, None),
            )
            if metodo == 'mostrar_opciones_exportar':
                btn.bind(on_release=lambda x: (self.mostrar_opciones_exportar(), self.toggle_menu()))
            elif metodo == 'hacer_backup':
                btn.bind(on_release=lambda x: (self.hacer_backup(), self.toggle_menu()))
            elif metodo == 'actualizar_desde_correo':
                btn.bind(on_release=lambda x: (self.actualizar_desde_correo(), self.toggle_menu()))
            elif metodo == 'alertas':
                btn.bind(on_release=lambda x: (setattr(self.manager, 'current', 'alertas'), self.toggle_menu()))
            grid.add_widget(btn)

    def toggle_menu(self):
        drawer = self.ids.drawer
        overlay = self.ids.drawer_overlay
        if self._menu_abierto:
            anim = Animation(opacity=0, duration=0.15) + Animation(x=-drawer.width, duration=0.2)
            anim.start(drawer)
            Animation(opacity=0, duration=0.15).start(overlay)
            self._menu_abierto = False
        else:
            self._construir_menu()
            drawer.x = -drawer.width
            drawer.opacity = 1
            anim = Animation(x=0, duration=0.2)
            anim.start(drawer)
            Animation(opacity=1, duration=0.15).start(overlay)
            self._menu_abierto = True

    def on_touch_down(self, touch):
        if self._menu_abierto and not self.ids.drawer.collide_point(*touch.pos):
            self.toggle_menu()
            return True
        return super().on_touch_down(touch)

    def actualizar_spinners(self) -> None:
        try:
            nombres    = sorted({str(r[0]).upper() for r in db_query("SELECT DISTINCT nombre FROM movimientos") if r[0]})
            materiales = sorted({str(r[0]).upper() for r in db_query("SELECT DISTINCT material FROM movimientos") if r[0]})
            self.ids.responsable_spinner.values = nombres
            self.ids.producto_spinner.values    = materiales
        except Exception as e:
            logger.error("Error actualizando spinners: %s", e)

    def _verificar_alertas_silencioso(self) -> None:
        try:
            alertas = obtener_alertas_activas()
            if alertas:
                n = len(alertas)
                self.ids.info.text = f"[color=#FF0044]{n} alerta{'s' if n > 1 else ''} de stock[/color]"
        except Exception as e:
            logger.error("Error verificando alertas: %s", e)

    def _actualizar_dashboard(self) -> None:
        try:
            hora = datetime.now().hour
            if hora < 7:
                saludo = "Buenas noches"
            elif hora < 12:
                saludo = "Buenos dias"
            elif hora < 18:
                saludo = "Buenas tardes"
            else:
                saludo = "Buenas noches"
            self.ids.lbl_saludo.text = saludo

            meses = ['enero','febrero','marzo','abril','mayo','junio',
                     'julio','agosto','septiembre','octubre','noviembre','diciembre']
            ahora = datetime.now()
            self.ids.lbl_fecha_hoy.text = f"{ahora.day} de {meses[ahora.month-1]} {ahora.year}  *  {ahora.strftime('%H:%M')}"

            total_mat = int(db_query("SELECT COUNT(DISTINCT material) FROM movimientos", fetchall=False)[0] or 0)
            self.ids.stat_materiales.text = str(total_mat)

            hoy = ahora.strftime("%d/%m/%Y")
            movs_hoy = int(db_query("SELECT COUNT(*) FROM movimientos WHERE fecha LIKE ? || '%'", (hoy,), fetchall=False)[0] or 0)
            self.ids.stat_movs_hoy.text = str(movs_hoy)

            alertas = obtener_alertas_activas()
            self.ids.stat_alertas.text = str(len(alertas))

            prestamos_activos = obtener_equipos_en_prestamo()
            self.ids.stat_pendientes.text = str(len(prestamos_activos))

            try:
                size_kb = os.path.getsize(DB_PATH) / 1024
                self.ids.lbl_resumen_db.text = f"BD: {size_kb:.0f} KB"
            except Exception:
                pass
        except Exception as e:
            logger.error("Error actualizando dashboard: %s", e)

    def _cargar_actividad_reciente(self) -> None:
        try:
            movs = db_query(
                "SELECT nombre, material, cantidad, tipo, fecha FROM movimientos ORDER BY id DESC LIMIT 5"
            )
            texto = "\n".join(
                f"{'▲' if t == 'ENTRADA' else '▼'} {m} x{c} - {n} {str(f)[:16] if f else ''}"
                for n, m, c, t, f in movs
            ) if movs else "Sin movimientos registrados"
            self.ids.info.text = f"[color=#5A6A7A]{texto}[/color]"
        except Exception as e:
            logger.error("Error cargando actividad: %s", e)

    def mostrar_alertas(self) -> None:
        rows = db_query(
            "SELECT nombre,material,sku,cantidad,tipo,fecha,notas FROM movimientos "
            "WHERE UPPER(nombre) != 'INVENTARIO' ORDER BY id ASC"
        )
        actualmente_fuera = {}
        for n, m, s, c, t, f, notas in rows:
            key = f"{n}|{m}"
            if t == "SALIDA":
                if key not in actualmente_fuera:
                    actualmente_fuera[key] = [n, m, s, float(c), t, f, notas or ""]
                else:
                    actualmente_fuera[key][3] += float(c)
            elif t == "ENTRADA" and key in actualmente_fuera:
                actualmente_fuera[key][3] -= float(c)
                if actualmente_fuera[key][3] <= 0:
                    del actualmente_fuera[key]
        pendientes = list(actualmente_fuera.values())
        content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))
        pop = Popup(title=f"MATERIALES SIN RETORNAR ({len(pendientes)})",
                    content=content, size_hint=(0.93, 0.82))
        if not pendientes:
            content.add_widget(Label(text="Todo en orden - ningun material pendiente de retorno.",
                                     color=get_color_from_hex('#00D4FF'), halign='center'))
        else:
            sv = ScrollView()
            gl = GridLayout(cols=1, size_hint_y=None, spacing=dp(8))
            gl.bind(minimum_height=gl.setter('height'))
            for item in sorted(pendientes, key=lambda x: x[0]):
                n, m, s, c, _, f, notas = item
                tiene_notas = bool(notas and notas.strip())
                altura = dp(110) if tiene_notas else dp(80)
                card = crear_tarjeta(altura, '#0D1B2A')
                card.add_widget(Label(
                    text=f"[b]{m}[/b]  -  Responsable: [color=#FF0044]{n}[/color]",
                    markup=True, color=get_color_from_hex('#FF0044'),
                    size_hint_y=None, height=dp(24),
                    halign='left', text_size=(Window.width * 0.82, None)
                ))
                card.add_widget(Label(
                    text=f"Cantidad fuera: [b]{round(c,2)}[/b]  |  SKU: {s or 'N/A'}  |  Salida: {f}",
                    color=get_color_from_hex('#FF0044'), font_size='12sp',
                    size_hint_y=None, height=dp(20),
                    halign='left', text_size=(Window.width * 0.82, None)
                ))
                if tiene_notas:
                    card.add_widget(Label(
                        text=f"[color=#5A6A7A]Notas:[/color] [color=#00D4FF]{notas}[/color]",
                        markup=True, font_size='11sp',
                        size_hint_y=None, height=dp(22),
                        halign='left', text_size=(Window.width * 0.82, None)
                    ))
                gl.add_widget(card)
            sv.add_widget(gl)
            content.add_widget(sv)
        content.add_widget(Button(text="CERRAR", size_hint_y=None, height=dp(45),
                                   background_color=get_color_from_hex('#334155'),
                                   on_release=pop.dismiss))
        pop.open()

    def procesar(self, tipo, r=None, p=None, c=None, s="N/A", desde_retorno=False) -> None:
        resp = (r
                or self.ids.nombre_nuevo.text.strip().upper()
                or (self.ids.responsable_spinner.text
                    if self.ids.responsable_spinner.text != "SELECCIONAR RESPONSABLE" else ""))
        prod = (p
                or self.ids.producto_nuevo.text.strip().upper()
                or (self.ids.producto_spinner.text
                    if self.ids.producto_spinner.text != "SELECCIONAR MATERIAL" else ""))
        cant_str = str(c) if c else self.ids.cantidad.text.strip()

        if not all([resp, prod, cant_str]):
            self.ids.info.text = "[color=#FF0044]Responsable, material y cantidad son obligatorios[/color]"
            return
        try:
            val = float(cant_str)
            if val <= 0:
                self.ids.info.text = "[color=#FF0044]La cantidad debe ser mayor a 0[/color]"
                return
            es_inventario = resp.upper() == "INVENTARIO"
            if tipo == "ENTRADA" and not es_inventario and not desde_retorno:
                self.ids.info.text = "[color=#FF0044]Solo INVENTARIO puede registrar entradas[/color]"
                return
            _, actual = obtener_stock_material(prod)
            if tipo == "SALIDA":
                if actual <= 0:
                    self.ids.info.text = f"[color=#FF0044]SIN STOCK DISPONIBLE: {prod}[/color]"
                    return
                if val > actual:
                    self.ids.info.text = f"[color=#FF0044]SOLO HAY {actual} UNIDADES DISPONIBLES[/color]"
                    return
            f_str     = self.ids.f_m.text.strip() or datetime.now().strftime("%d/%m/%Y %H:%M")
            dias      = self.ids.d_p.text.strip() or 0
            notas     = self.ids.notas.text.strip()
            ubicacion = self.ids.ubicacion.text.strip().upper()
            sku_val   = s if s != "N/A" else self.ids.sku.text.strip()
            fijo, actual2 = obtener_stock_material(prod)
            nuevo_st = round(actual2 + (val if tipo == "ENTRADA" else -val), 4)

            last_id = db_execute(
                "INSERT INTO movimientos (nombre,material,sku,cantidad,tipo,fecha,stock_registro,dias,retorno,notas,ubicacion) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (resp, prod, sku_val, val, tipo, f_str, nuevo_st, dias, "N/A", notas, ubicacion)
            )
            self._ultimo_mov_id = last_id
            self._ultimo_mov_tipo = tipo
            try:
                self.ids.btn_deshacer.disabled = False
            except Exception:
                pass
            if resp.upper() == "INVENTARIO" and self.ids.c_l.active:
                db_execute(
                    "INSERT INTO limpieza (material,cantidad,estado,fecha,notas) VALUES (?,?,?,?,?)",
                    (prod, val, "SUCIO", f_str, notas)
                )
            log_auditoria(f"MOV_{tipo}", f"{resp} | {prod} | x{val} | {f_str}")
            icono = "▲" if tipo == "ENTRADA" else "▼"
            self.ids.info.text = f"[color=#00D4FF] {icono} {tipo} REGISTRADA - {prod} x{val}[/color]"
            self.actualizar_spinners()
            self._actualizar_dashboard()
        except ValueError:
            self.ids.info.text = "[color=#FF0044] CANTIDAD INVALIDA[/color]"
        except sqlite3.Error as e:
            self.ids.info.text = f"[color=#FF0044] ERROR BD: {e}[/color]"

    def limpiar_campos(self) -> None:
        for field in ['nombre_nuevo','producto_nuevo','sku','cantidad','d_p','f_m','notas','ubicacion']:
            self.ids[field].text = ""
        self.ids.responsable_spinner.text = "SELECCIONAR RESPONSABLE"
        self.ids.producto_spinner.text    = "SELECCIONAR MATERIAL"
        self.ids.c_l.active = False

    def deshacer_ultimo(self):
        if not self._ultimo_mov_id:
            return
        try:
            db_execute("DELETE FROM movimientos WHERE id = ?", (self._ultimo_mov_id,))
            log_auditoria("DESHACER", f"Movimiento #{self._ultimo_mov_id} eliminado")
            self.ids.info.text = f"[color=#FF0044] Deshecho: ultimo movimiento eliminado[/color]"
            self._ultimo_mov_id = None
            self._ultimo_mov_tipo = None
            self.ids.btn_deshacer.disabled = True
            self.actualizar_spinners()
            self._actualizar_dashboard()
        except Exception as e:
            self.ids.info.text = f"[color=#FF0044] Error al deshacer: {e}[/color]"
        gc.collect()

    def salida_por_categoria(self):
        from kivy.uix.gridlayout import GridLayout
        from kivy.uix.scrollview import ScrollView
        from kivy.uix.spinner import Spinner
        from kivy.uix.textinput import TextInput as TI

        categorias = db_query("SELECT id, nombre FROM categorias ORDER BY nombre")
        if not categorias:
            self.ids.info.text = "[color=#FF0044]Crea categorias primero en CATEGORIAS[/color]"
            return
        if len(categorias) == 1:
            self._mostrar_items_categoria(categorias[0][0], categorias[0][1])
            return
        content = BoxLayout(orientation='vertical', spacing=dp(12), padding=dp(15))
        pop = Popup(title="SELECCIONAR CATEGORIA", content=content, size_hint=(0.85, 0.35))
        spinner = Spinner(text=categorias[0][1], values=[c[1] for c in categorias],
                          size_hint_y=None, height=dp(48))
        content.add_widget(spinner)
        def continuar(btn):
            for cid, cnom in categorias:
                if cnom == spinner.text:
                    pop.dismiss()
                    self._mostrar_items_categoria(cid, cnom)
                    break
        btn = Button(text="CONTINUAR", size_hint_y=None, height=dp(50),
                     background_color=get_color_from_hex('#0066FF'), on_release=continuar)
        content.add_widget(btn)
        pop.open()

    def _mostrar_items_categoria(self, cid, cnom):
        from kivy.uix.gridlayout import GridLayout as GL
        from kivy.uix.scrollview import ScrollView as SV
        from kivy.uix.textinput import TextInput as TI

        items = db_query("SELECT id, nombre FROM categoria_items WHERE categoria_id = ? ORDER BY nombre", (cid,))
        if not items:
            self.ids.info.text = f"[color=#FF0044]La categoria {cnom} no tiene items[/color]"
            return
        content = BoxLayout(orientation='vertical', spacing=dp(8), padding=dp(12))
        pop = Popup(title=f"SALIDA: {cnom}", content=content, size_hint=(0.92, 0.82))

        resp = self.ids.nombre_nuevo.text.strip().upper() or self.ids.responsable_spinner.text
        if not resp or resp == "SELECCIONAR RESPONSABLE":
            content.add_widget(Label(
                text="Primero selecciona un responsable arriba",
                color=get_color_from_hex('#FF0044'), size_hint_y=None, height=dp(30)))
            content.add_widget(Button(text="CERRAR", size_hint_y=None, height=dp(44),
                                       background_color=get_color_from_hex('#334155'),
                                       on_release=pop.dismiss))
            pop.open()
            return

        inputs = {}
        sv = SV()
        gl = GL(cols=1, size_hint_y=None, spacing=dp(4))
        gl.bind(minimum_height=gl.setter('height'))

        for it in items:
            row = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(8))
            row.add_widget(Label(
                text=it[1],
                color=get_color_from_hex('#E8EAED'),
                font_size='13sp', halign='left',
                text_size=(Window.width * 0.5, None),
                size_hint_x=0.6))
            inp = TI(hint_text="0", input_filter='float', multiline=False,
                     size_hint_x=0.25, font_size='14sp')
            inputs[it[0]] = (it[1], inp)
            row.add_widget(inp)
            gl.add_widget(row)

        sv.add_widget(gl)
        content.add_widget(sv)

        def confirmar(btn):
            registrados = 0
            fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
            notas = self.ids.notas.text.strip()
            for iid, (inom, inp) in inputs.items():
                val_str = inp.text.strip()
                if not val_str:
                    continue
                try:
                    val = float(val_str)
                except ValueError:
                    continue
                if val <= 0:
                    continue
                _, actual = obtener_stock_material(inom)
                if actual <= 0:
                    continue
                if val > actual:
                    val = actual
                _, actual2 = obtener_stock_material(inom)
                nuevo_st = round(actual2 - val, 4)
                db_execute(
                    "INSERT INTO movimientos (nombre,material,sku,cantidad,tipo,fecha,stock_registro,dias,retorno,notas,ubicacion,categoria_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                    (resp, inom, "", val, "SALIDA", fecha, nuevo_st, 0, "N/A", notas, "", cid)
                )
                registrados += 1
            pop.dismiss()
            if registrados > 0:
                self.ids.info.text = f"[color=#00D4FF] Salida x{registrados} items de {cnom}[/color]"
                self.actualizar_spinners()
                self._actualizar_dashboard()

        btn = Button(text=f"CONFIRMAR SALIDA", size_hint_y=None, height=dp(48),
                     background_color=get_color_from_hex('#FF0044'), on_release=confirmar)
        content.add_widget(btn)
        pop.open()

    def enviar_correo(self) -> None:
        self.ids.info.text = "Preparando Excel..."
        threading.Thread(target=self._enviar_correo_worker, daemon=True).start()

    def _enviar_correo_worker(self) -> None:
        exito, msg = enviar_reporte_por_correo()
        if not exito:
            logger.error("Error al enviar correo: %s", msg)
        Clock.schedule_once(lambda dt: setattr(
            self.ids.info, 'text',
            f"[color=#00D4FF]ENVIADO EXITOSAMENTE[/color]" if exito else f"[color=#FF0044]ERROR: {msg}[/color]"
        ))

    def mostrar_opciones_exportar(self):
        content = BoxLayout(orientation='vertical', spacing=dp(12), padding=dp(15))
        pop = Popup(title="EXPORTAR", content=content, size_hint=(0.8, 0.35))
        btn_pdf = Button(text="EXPORTAR PDF", font_size='15sp', bold=True,
                         background_color=get_color_from_hex('#0066FF'),
                         size_hint_y=None, height=dp(52))
        btn_pdf.bind(on_release=lambda x: (pop.dismiss(), self.exportar_pdf()))
        btn_excel = Button(text="ENVIAR POR CORREO (EXCEL)", font_size='13sp',
                           background_color=get_color_from_hex('#0066FF'),
                           size_hint_y=None, height=dp(52))
        btn_excel.bind(on_release=lambda x: (pop.dismiss(), self.enviar_correo()))
        content.add_widget(btn_pdf)
        content.add_widget(btn_excel)
        content.add_widget(Button(text="CANCELAR", size_hint_y=None, height=dp(40),
                                   background_color=get_color_from_hex('#334155'),
                                   on_release=pop.dismiss))
        pop.open()

    def hacer_backup(self) -> None:
        self.ids.info.text = " Generando backup..."
        threading.Thread(target=self._backup_worker, daemon=True).start()

    def exportar_pdf(self) -> None:
        self.ids.info.text = " Generando PDF..."
        threading.Thread(target=self._pdf_worker, daemon=True).start()

    def _pdf_worker(self):
        try:
            ruta = generar_pdf()
            Clock.schedule_once(lambda dt: Toast.show(f"PDF guardado", '#0066FF', 4))
            Clock.schedule_once(lambda dt: setattr(self.ids.info, 'text',
                                                   f"[color=#00D4FF]PDF: {os.path.basename(ruta)}[/color]"))
        except Exception as e:
            logger.error("Error PDF: %s", e)
            Clock.schedule_once(lambda dt, err=str(e): setattr(
                self.ids.info, 'text', f"[color=#FF0044]Error PDF: {err}[/color]"))

    def _backup_worker(self) -> None:
        try:
            nombre = ruta_exportacion(f"backup_inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            generar_excel(nombre)
            log_auditoria("BACKUP", nombre)
            Clock.schedule_once(lambda dt: setattr(self.ids.info, 'text',
                                                   f"[color=#00D4FF]Backup guardado en:\n{nombre}[/color]"))
        except Exception as e:
            logger.error("Error backup: %s", e)
            Clock.schedule_once(lambda dt: setattr(self.ids.info, 'text',
                                                   f"[color=#FF0044]Error backup: {e}[/color]"))

    def actualizar_desde_correo(self) -> None:
        self._popup_correo = self._crear_popup_progreso()
        self._popup_correo.open()
        actualizador = ActualizadorCorreo(
            callback_progreso=self._actualizar_progreso_correo,
            callback_fin=self._finalizar_actualizacion_correo,
        )
        actualizador.ejecutar()

    def _crear_popup_progreso(self):
        content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(12))
        pop = Popup(title="Actualizando desde correo...", content=content,
                    size_hint=(0.92, 0.55), auto_dismiss=False)
        self._lbl_progreso = Label(
            text="Iniciando...", color=get_color_from_hex('#00D4FF'),
            halign='center', font_size='14sp',
            text_size=(Window.width * 0.85, None),
        )
        content.add_widget(self._lbl_progreso)
        btn_cerrar = Button(
            text="Procesando... (toca para cerrar si ya termino)",
            size_hint_y=None, height=dp(46),
            background_color=get_color_from_hex('#334155'),
        )
        btn_cerrar.bind(on_release=pop.dismiss)
        content.add_widget(btn_cerrar)
        self._btn_cerrar_popup = btn_cerrar
        return pop

    def _actualizar_progreso_correo(self, texto):
        try:
            self._lbl_progreso.text = texto
        except Exception:
            pass

    def _finalizar_actualizacion_correo(self, exito, datos):
        try:
            self._popup_correo.dismiss()
        except Exception:
            pass
        if exito:
            self._mostrar_resultado_correo(datos)
            self.actualizar_spinners()
        else:
            self._mostrar_error_correo(datos.get("error", "Error desconocido"))

    def _mostrar_resultado_correo(self, datos):
        content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(10))
        pop = Popup(title="Inventario actualizado", content=content, size_hint=(0.92, 0.72))
        card_ok = crear_tarjeta(dp(60), '#0D1B2A')
        card_ok.add_widget(Label(
            text="[b][color=#00D4FF]ACTUALIZACION COMPLETADA[/color][/b]",
            markup=True, font_size='15sp',
        ))
        content.add_widget(card_ok)
        if datos.get("de") or datos.get("asunto"):
            card_origen = crear_tarjeta(dp(80), '#0D1B2A')
            card_origen.add_widget(Label(
                text=f"De: {datos.get('de', 'N/A')[:50]}",
                color=get_color_from_hex('#5A6A7A'), font_size='12sp',
            ))
            card_origen.add_widget(Label(
                text=f"Asunto: {datos.get('asunto', 'N/A')[:50]}",
                color=get_color_from_hex('#5A6A7A'), font_size='12sp',
            ))
            content.add_widget(card_origen)
        card_stats = crear_tarjeta(dp(110), '#0D1B2A')
        card_stats.add_widget(Label(
            text="RESULTADOS DE LA IMPORTACION",
            color=get_color_from_hex('#00D4FF'), font_size='12sp', bold=True,
        ))
        card_stats.add_widget(Label(
            text=f"[b][color=#00D4FF]Insertados: {datos.get('insertados', 0)}[/color][/b]",
            markup=True, font_size='16sp',
        ))
        card_stats.add_widget(Label(
            text=f"Omitidos: {datos.get('omitidos', 0)}   Errores: {datos.get('errores', 0)}",
            color=get_color_from_hex('#5A6A7A'), font_size='13sp',
        ))
        content.add_widget(card_stats)
        card_archivo = crear_tarjeta(dp(65), '#0D1B2A')
        card_archivo.add_widget(Label(
            text=f"Archivo: {datos.get('archivo', 'N/A')[:45]}",
            color=get_color_from_hex('#E8EAED'), font_size='12sp',
        ))
        card_archivo.add_widget(Label(
            text=f"{datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
            color=get_color_from_hex('#5A6A7A'), font_size='11sp',
        ))
        content.add_widget(card_archivo)
        content.add_widget(Button(
            text="ENTENDIDO", size_hint_y=None, height=dp(48),
            background_color=get_color_from_hex('#0066FF'), on_release=pop.dismiss,
        ))
        self.ids.info.text = f"[color=#00D4FF]Actualizado desde correo: +{datos.get('insertados', 0)} registros[/color]"
        pop.open()

    def _mostrar_error_correo(self, mensaje_error):
        content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(10))
        pop = Popup(title="Error al actualizar desde correo", content=content, size_hint=(0.92, 0.65))
        sv = ScrollView()
        lbl = Label(
            text=mensaje_error, color=get_color_from_hex('#FF0044'),
            halign='center', font_size='13sp', size_hint_y=None,
            text_size=(Window.width * 0.82, None),
        )
        lbl.bind(texture_size=lbl.setter('size'))
        sv.add_widget(lbl)
        content.add_widget(sv)
        card_tip = crear_tarjeta(dp(130), '#0D1B2A')
        card_tip.add_widget(Label(
            text="POSIBLES SOLUCIONES",
            color=get_color_from_hex('#FF0044'), bold=True, font_size='13sp',
        ))
        card_tip.add_widget(Label(
            text="* Activa IMAP en Gmail -> Configuracion -> Ver toda la config -> Reenvio e IMAP",
            color=get_color_from_hex('#5A6A7A'), font_size='11sp', halign='center',
            text_size=(Window.width * 0.75, None),
        ))
        card_tip.add_widget(Label(
            text="* Usa una contrasena de aplicacion (no la contrasena normal)",
            color=get_color_from_hex('#5A6A7A'), font_size='11sp', halign='center',
            text_size=(Window.width * 0.75, None),
        ))
        card_tip.add_widget(Label(
            text="* Verifica que el Excel adjunto este en la bandeja de entrada",
            color=get_color_from_hex('#5A6A7A'), font_size='11sp', halign='center',
            text_size=(Window.width * 0.75, None),
        ))
        content.add_widget(card_tip)
        content.add_widget(Button(
            text="CERRAR", size_hint_y=None, height=dp(46),
            background_color=get_color_from_hex('#334155'), on_release=pop.dismiss,
        ))
        self.ids.info.text = "[color=#FF0044]Error al actualizar desde correo[/color]"
        pop.open()


# === screens/valor.py ===
from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.spinner import Spinner
from kivy.uix.scrollview import ScrollView
from kivy.metrics import dp
from kivy.utils import get_color_from_hex
from kivy.core.window import Window


class ValorScreen(Screen):
    _precios = {}

    def on_enter(self):
        self._cargar_precios_guardados()
        self.cargar_materiales()

    def _cargar_precios_guardados(self):
        rows = db_query("SELECT clave, valor FROM config_app WHERE clave LIKE 'precio_%'")
        for clave, valor in rows:
            try:
                mat = clave.replace("precio_", "", 1)
                self._precios[mat] = int(valor)
            except Exception:
                pass

    def _guardar_precio(self, material, precio):
        db_execute(
            "INSERT INTO config_app (clave, valor) VALUES (?, ?) ON CONFLICT(clave) DO UPDATE SET valor = ?",
            (f"precio_{material}", str(precio), str(precio))
        )

    def cargar_materiales(self):
        self.ids.lista_valores.clear_widgets()
        rows = db_query("SELECT DISTINCT UPPER(material) FROM movimientos ORDER BY material")
        materiales = [r[0] for r in rows]
        self.ids.spinner_material.values = materiales
        self.actualizar_valor_unitario()

    def _ajustar_scroll(self):
        children = self.ids.lista_valores.children
        if children:
            total = sum(c.height for c in children) + (len(children) - 1) * dp(4)
            self.ids.scroll_items.height = min(total, dp(300))
        else:
            self.ids.scroll_items.height = 0

    def _calcular_total_inventario(self):
        rows = db_query("""
            SELECT UPPER(material), SUM(CASE WHEN tipo='ENTRADA' THEN cantidad ELSE -cantidad END)
            FROM movimientos GROUP BY UPPER(material)
        """)
        total = 0
        for mat, disp in rows:
            if disp and disp > 0:
                precio = self._precios.get(mat, 0)
                total += disp * precio
        return total

    def actualizar_valor_unitario(self):
        self.ids.lista_valores.clear_widgets()
        self._ajustar_scroll()
        mat = self.ids.spinner_material.text

        total_inventario = self._calcular_total_inventario()
        self.ids.total_inventario_label.text = f"$ {total_inventario:,.0f}"
        self.ids.total_inventario_label.color = get_color_from_hex('#00D4FF') if total_inventario > 0 else get_color_from_hex('#5A6A7A')

        if mat == "-- SELECCIONAR --" or not mat:
            self.ids.unitario_label.text = "$ 0"
            self.ids.unitario_label.color = get_color_from_hex('#5A6A7A')
            self.ids.total_label.text = "$ 0"
            self.ids.total_label.color = get_color_from_hex('#5A6A7A')
            return
        precio = self._precios.get(mat, 0)

        _, disp = obtener_stock_material(mat)
        if disp <= 0:
            self.ids.unitario_label.text = "$ 0"
            self.ids.unitario_label.color = get_color_from_hex('#5A6A7A')
            self.ids.total_label.text = "$ 0"
            self.ids.total_label.color = get_color_from_hex('#5A6A7A')
            return

        if precio > 0:
            self.ids.unitario_label.text = f"$ {precio:,.0f}"
            self.ids.unitario_label.color = get_color_from_hex('#00D4FF')
            subtotal = disp * precio
            self.ids.total_label.text = f"$ {subtotal:,.0f}"
            self.ids.total_label.color = get_color_from_hex('#00D4FF')
        else:
            self.ids.unitario_label.text = "SIN PRECIO"
            self.ids.unitario_label.color = get_color_from_hex('#5A6A7A')
            self.ids.total_label.text = "SIN PRECIO"
            self.ids.total_label.color = get_color_from_hex('#5A6A7A')

        card = crear_tarjeta(dp(56), '#1A2035')
        header = BoxLayout(size_hint_y=None, height=dp(28), spacing=dp(8))
        header.add_widget(Label(
            text=f"[b]{mat}[/b]  x [color=#00D4FF]{disp}[/color]",
            markup=True, color=get_color_from_hex('#00D4FF'),
            font_size='12sp', halign='left', valign='middle',
            text_size=(Window.width * 0.45, None),
            size_hint_x=0.5, shorten=True, shorten_from='right'))
        if precio > 0:
            subtotal = disp * precio
            header.add_widget(Label(
                text=f"$ {precio:,.0f}",
                color=get_color_from_hex('#00D4FF'),
                font_size='12sp', halign='center', valign='middle',
                text_size=self.size, size_hint_x=0.25, shorten=True))
            header.add_widget(Label(
                text=f"$ {subtotal:,.0f}",
                color=get_color_from_hex('#E8EAED'),
                font_size='13sp', halign='center', valign='middle',
                text_size=self.size, size_hint_x=0.25, bold=True, shorten=True))
        else:
            header.add_widget(Label(
                text="-",
                color=get_color_from_hex('#475569'),
                font_size='11sp', halign='center', valign='middle',
                size_hint_x=0.25))
            header.add_widget(Label(
                text="-",
                color=get_color_from_hex('#475569'),
                font_size='11sp', halign='center', valign='middle',
                size_hint_x=0.25))
        card.add_widget(header)
        self.ids.lista_valores.add_widget(card)
        self._ajustar_scroll()

    def asignar_precio(self):
        mat = self.ids.spinner_material.text
        precio_str = self.ids.input_precio.text.strip()
        if mat == "-- SELECCIONAR --" or not mat:
            self.ids.info_valor.text = "Selecciona un material"
            return
        if not precio_str:
            self.ids.info_valor.text = "Ingresa un precio"
            return
        try:
            precio = int(float(precio_str.replace(",", "")))
        except ValueError:
            self.ids.info_valor.text = "Precio invalido"
            return
        if precio <= 0:
            self.ids.info_valor.text = "Precio debe ser mayor a 0"
            return
        self._precios[mat] = precio
        self._guardar_precio(mat, precio)
        self.ids.input_precio.text = ""
        self.ids.info_valor.text = f"{mat}: $ {precio:,} asignado"
        self.cargar_materiales()


# === screens/ajustes.py ===
import csv
import threading
from datetime import datetime, timedelta
from kivy.app import App
from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.popup import Popup
from kivy.clock import Clock
from kivy.metrics import dp
from kivy.utils import get_color_from_hex


class AjustesScreen(Screen):
    def on_enter(self):
        self.actualizar_spinner_alertas()
        try:
            self.ids.lbl_carpeta_export.text = f"Carpeta: {EXPORT_DIR}"
            self.ids.filtro_asunto.text    = EMAIL_CONFIG.get("asunto_filtro", "") or obtener_config("filtro_asunto", "")
            self.ids.filtro_remitente.text = EMAIL_CONFIG.get("remitente_filtro", "") or obtener_config("filtro_remitente", "")
            self.ids.backup_interval.text  = obtener_config("backup_interval_horas", "0")
        except Exception as e:
            logger.error("Error en on_enter ajustes: %s", e)

    def actualizar_spinner_alertas(self) -> None:
        try:
            mats = sorted({str(r[0]).upper() for r in db_query("SELECT DISTINCT material FROM movimientos") if r[0]})
            self.ids.alerta_material_spinner.values = mats
        except Exception as e:
            logger.error("Error actualizando spinner alertas: %s", e)

    def guardar_alerta(self) -> None:
        mat    = (self.ids.alerta_material_spinner.text
                  if self.ids.alerta_material_spinner.text != "SELECCIONAR MATERIAL" else "")
        minimo = self.ids.stock_minimo_input.text.strip()
        if not mat or not minimo:
            self.ids.info_alerta.text = "Selecciona material y escribe minimo"
            return
        try:
            db_execute(
                "INSERT INTO alertas_stock (material,stock_minimo) VALUES (?,?) "
                "ON CONFLICT(material) DO UPDATE SET stock_minimo=excluded.stock_minimo",
                (mat.upper(), float(minimo))
            )
            self.ids.info_alerta.text = f"Alerta guardada: {mat} -> min {minimo}"
            self.ids.stock_minimo_input.text = ""
        except Exception as e:
            logger.error("Error guardando alerta: %s", e)
            self.ids.info_alerta.text = f"Error: {e}"

    def guardar_filtros_correo(self) -> None:
        EMAIL_CONFIG["asunto_filtro"]    = self.ids.filtro_asunto.text.strip()
        EMAIL_CONFIG["remitente_filtro"] = self.ids.filtro_remitente.text.strip()
        guardar_config("filtro_asunto", EMAIL_CONFIG["asunto_filtro"])
        guardar_config("filtro_remitente", EMAIL_CONFIG["remitente_filtro"])
        self.ids.info_ajustes.text = "Filtros de correo guardados"

    def exportar_csv(self) -> None:
        self.ids.info_ajustes.text = "Exportando CSV..."
        threading.Thread(target=self._exportar_csv_worker, daemon=True).start()

    def _exportar_csv_worker(self) -> None:
        nombre = ruta_exportacion(f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
        try:
            rows   = db_query("SELECT nombre,material,sku,cantidad,tipo,fecha,stock_registro,dias,retorno,notas,ubicacion FROM movimientos")
            with open(nombre, 'w', newline='', encoding='utf-8') as f:
                w = csv.writer(f)
                w.writerow(["Responsable","Material","SKU","Cantidad","Tipo","Fecha",
                             "Stock","Dias","Retorno","Notas","Ubicacion"])
                w.writerows(rows)
            log_auditoria("EXPORT_CSV", nombre)
            Clock.schedule_once(lambda dt, n=nombre: self._set_exportado(f"CSV guardado en:\n{n}"))
        except Exception as ex:
            logger.error("Error exportando CSV: %s", ex)
            err = str(ex)
            Clock.schedule_once(lambda dt, e=err: self._set_exportado(f"Error: {e}"))

    def exportar_excel_local(self) -> None:
        self.ids.info_ajustes.text = "Exportando Excel..."
        threading.Thread(target=self._exportar_excel_worker, daemon=True).start()

    def _exportar_excel_worker(self) -> None:
        nombre = ruta_exportacion(f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        try:
            generar_excel(nombre)
            log_auditoria("EXPORT_XLSX", nombre)
            Clock.schedule_once(lambda dt, n=nombre: self._set_exportado(f"Excel guardado en:\n{n}"))
        except Exception as ex:
            logger.error("Error exportando Excel: %s", ex)
            err = str(ex)
            Clock.schedule_once(lambda dt, e=err: self._set_exportado(f"Error: {e}"))

    def _set_exportado(self, texto):
        self.ids.info_ajustes.text = texto
        try:
            App.get_running_app().root.get_screen('inventario')._actualizar_dashboard()
        except Exception:
            pass

    def confirmar_purga(self) -> None:
        content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(10))
        pop = Popup(title="Confirmar Purga", content=content, size_hint=(0.88, 0.42))
        content.add_widget(Label(
            text="Eliminar movimientos de mas de\n[b]90 dias[/b]?\nEsta accion [b]NO[/b] se puede deshacer.",
            markup=True, halign='center'
        ))
        btn = Button(text="SI, PURGAR",
                     background_color=get_color_from_hex('#991B1B'),
                     size_hint_y=None, height=dp(45))
        btn.bind(on_release=lambda x: self._ejecutar_purga(pop))
        content.add_widget(btn)
        content.add_widget(Button(text="Cancelar", size_hint_y=None, height=dp(40),
                                   background_color=get_color_from_hex('#334155'),
                                   on_release=pop.dismiss))
        pop.open()

    def guardar_backup_config(self) -> None:
        horas = self.ids.backup_interval.text.strip()
        guardar_config("backup_interval_horas", horas)
        self.ids.info_ajustes.text = f"Backup automatico: cada {horas}h" if horas != "0" else "Backup automatico desactivado"

    def _ejecutar_purga(self, pop) -> None:
        try:
            limite = (datetime.now() - timedelta(days=90)).strftime("%d/%m/%Y")
            db_execute("DELETE FROM movimientos WHERE fecha < ?", (limite,))
            eliminados = db_query(
                "SELECT COUNT(*) FROM movimientos WHERE fecha < ?",
                (limite,), fetchall=False
            )[0] or 0
            log_auditoria("PURGA", f"{eliminados} registros eliminados")
            pop.dismiss()
            self.ids.info_ajustes.text = f"{eliminados} registros eliminados"
        except Exception as e:
            pop.dismiss()
            logger.error("Error purgando: %s", e)
            self.ids.info_ajustes.text = f"Error: {e}"


# === screens/importar.py ===
import os
import threading
from datetime import datetime
from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
from kivy.uix.textinput import TextInput
from kivy.clock import Clock
from kivy.metrics import dp
from kivy.utils import get_color_from_hex
from kivy.core.window import Window
from kivy.app import App

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class ExploradorArchivos(Popup):
    EXTENSIONES_EXCEL = ('.xlsx', '.xls', '.xlsm', '.xlsb')

    def __init__(self, callback, **kwargs):
        super().__init__(**kwargs)
        self.callback    = callback
        self.title       = "Seleccionar archivo Excel"
        self.size_hint   = (0.97, 0.93)
        self.ruta_actual = self._ruta_inicial()

        root_layout = BoxLayout(orientation='vertical', spacing=dp(6), padding=dp(8))

        accesos = [("Mis Reportes (app)", EXPORT_DIR)]
        try:
            from kivy.utils import platform
            if platform == 'android':
                accesos += [
                    ("Descargas",    "/storage/emulated/0/Download"),
                    ("Documentos",   "/storage/emulated/0/Documents"),
                    ("WhatsApp",     "/storage/emulated/0/Android/media/com.whatsapp/WhatsApp/Media/WhatsApp Documents"),
                    ("Telegram",     "/storage/emulated/0/Android/media/org.telegram.messenger/Telegram/Telegram Documents"),
                    ("Raiz",         "/storage/emulated/0"),
                ]
            else:
                import pathlib
                accesos += [
                    ("Descargas",    os.path.join(str(pathlib.Path.home()), "Downloads")),
                    ("Inicio",       str(pathlib.Path.home())),
                ]
        except Exception:
            import pathlib
            accesos += [
                ("Descargas", os.path.join(str(pathlib.Path.home()), "Downloads")),
                ("Inicio",    str(pathlib.Path.home())),
            ]

        atajos_scroll = ScrollView(size_hint_y=None, height=dp(48), do_scroll_y=False)
        atajos_box    = BoxLayout(size_hint_x=None, spacing=dp(6))
        atajos_box.bind(minimum_width=atajos_box.setter('width'))
        for nombre, ruta in accesos:
            if os.path.exists(ruta):
                btn = Button(
                    text=nombre, size_hint=(None, 1), width=dp(150),
                    font_size='12sp', background_color=get_color_from_hex('#1E40AF')
                )
                btn.bind(on_release=lambda x, r=ruta: self.listar(r))
                atajos_box.add_widget(btn)
        atajos_scroll.add_widget(atajos_box)
        root_layout.add_widget(atajos_scroll)

        self.lbl_ruta = Label(
            text=self.ruta_actual, size_hint_y=None, height=dp(30),
            font_size='11sp', color=get_color_from_hex('#94A3B8'),
            halign='left', shorten=True, shorten_from='left',
            text_size=(Window.width * 0.92, None)
        )
        root_layout.add_widget(self.lbl_ruta)

        self.filtro = TextInput(
            hint_text='Filtrar archivos...', multiline=False,
            size_hint_y=None, height=dp(42)
        )
        self.filtro.bind(text=lambda inst, val: self.listar(self.ruta_actual))
        root_layout.add_widget(self.filtro)

        self.lbl_debug = Label(
            text='', size_hint_y=None, height=dp(24), font_size='11sp',
            color=get_color_from_hex('#F59E0B'), halign='left',
            text_size=(Window.width * 0.92, None)
        )
        root_layout.add_widget(self.lbl_debug)

        self.scroll = ScrollView()
        self.lista  = GridLayout(cols=1, size_hint_y=None, spacing=dp(5))
        self.lista.bind(minimum_height=self.lista.setter('height'))
        self.scroll.add_widget(self.lista)
        root_layout.add_widget(self.scroll)

        btns = BoxLayout(size_hint_y=None, height=dp(46), spacing=dp(8))
        btn_back = Button(text='SUBIR NIVEL', background_color=get_color_from_hex('#334155'), font_size='13sp')
        btn_back.bind(on_release=lambda x: self.subir_nivel())
        btns.add_widget(btn_back)
        btn_cerrar = Button(text='CANCELAR', background_color=get_color_from_hex('#7F1D1D'), font_size='13sp')
        btn_cerrar.bind(on_release=self.dismiss)
        btns.add_widget(btn_cerrar)
        root_layout.add_widget(btns)
        self.content = root_layout
        self.listar(self.ruta_actual)

    def _ruta_inicial(self):
        if os.path.exists(EXPORT_DIR):
            return EXPORT_DIR
        try:
            from kivy.utils import platform
            if platform == 'android':
                for r in ['/storage/emulated/0/Download', '/storage/emulated/0']:
                    if os.path.exists(r):
                        return r
        except Exception:
            pass
        import pathlib
        descargas = os.path.join(str(pathlib.Path.home()), 'Downloads')
        return descargas if os.path.exists(descargas) else str(pathlib.Path.home())

    def listar(self, ruta):
        self.lista.clear_widgets()
        self.lbl_ruta.text = ruta
        self.ruta_actual   = ruta
        filtro_txt = self.filtro.text.lower().strip() if hasattr(self, 'filtro') else ''
        try:
            entradas = sorted(os.listdir(ruta))
        except PermissionError:
            self.lista.add_widget(Label(
                text='Sin permiso para acceder a esta carpeta',
                color=get_color_from_hex('#F87171'), size_hint_y=None, height=dp(60), halign='center'
            ))
            return
        except Exception as e:
            self.lista.add_widget(Label(text=f'Error: {e}', color=get_color_from_hex('#F87171'), size_hint_y=None, height=dp(50)))
            return

        carpetas = []; archivos = []; total_enc = 0
        for nombre in entradas:
            ns = nombre.strip()
            if ns.startswith('.'):
                continue
            rc = os.path.join(ruta, ns)
            if filtro_txt and filtro_txt not in ns.lower():
                continue
            try:
                es_dir = os.path.isdir(rc)
            except Exception:
                continue
            if es_dir:
                carpetas.append((ns, rc))
            elif ns.lower().endswith(self.EXTENSIONES_EXCEL):
                archivos.append((ns, rc))
                total_enc += 1

        total_en_carpeta = len([n for n in entradas if not n.startswith('.')])
        if hasattr(self, 'lbl_debug'):
            self.lbl_debug.text = f"{total_en_carpeta} entradas  |  {len(carpetas)} carpetas  |  {total_enc} Excel encontrado(s)"

        if not carpetas and not archivos:
            self.lista.add_widget(Label(
                text='No hay archivos Excel en esta carpeta',
                color=get_color_from_hex('#94A3B8'), size_hint_y=None, height=dp(50), halign='center'
            ))
            return

        for ns, rc in carpetas:
            btn = Button(text=f'DIR  {ns}', size_hint_y=None, height=dp(52), halign='left',
                         font_size='13sp', background_color=get_color_from_hex('#1E3A5F'))
            btn.bind(on_release=lambda x, r=rc: self.listar(r))
            self.lista.add_widget(btn)

        for ns, rc in archivos:
            try:
                size_txt = f'{os.path.getsize(rc)/1024:.1f} KB'
            except Exception:
                size_txt = ''
            btn = Button(
                text=f'XLS  {ns}\n[size=11sp][color=#94A3B8]{size_txt}[/color][/size]',
                markup=True, size_hint_y=None, height=dp(60), halign='left',
                font_size='13sp', background_color=get_color_from_hex('#065F46')
            )
            btn.bind(on_release=lambda x, r=rc: self._seleccionar(r))
            self.lista.add_widget(btn)

    def subir_nivel(self):
        padre = os.path.dirname(self.ruta_actual)
        if padre and padre != self.ruta_actual:
            self.listar(padre)

    def _seleccionar(self, ruta):
        self.dismiss()
        Clock.schedule_once(lambda dt: self.callback(ruta), 0.15)


class ImportarExcelScreen(Screen):
    _columnas_excel = []

    def on_enter(self):
        self._columnas_excel = []
        self.ids.preview_container.clear_widgets()
        self.ids.btn_importar.disabled = True
        self.ids.estado_importar.text  = "Pulsa EXPLORAR para buscar tu archivo Excel"
        try:
            self.ids.lbl_hint_carpeta.text = f"Reportes de la app en: {EXPORT_DIR}"
        except Exception:
            pass

    def abrir_explorador(self) -> None:
        try:
            explorador = ExploradorArchivos(callback=self._al_seleccionar_archivo)
            explorador.open()
        except Exception as e:
            self.ids.estado_importar.text = f"[color=#F87171]Error al abrir explorador: {e}[/color]"

    def _al_seleccionar_archivo(self, ruta):
        self.ids.ruta_excel.text      = ruta
        self.ids.estado_importar.text = f"[color=#10B981]Archivo seleccionado[/color]"
        Clock.schedule_once(lambda dt: self.previsualizar(), 0.1)

    def pegar_ruta(self) -> None:
        try:
            from kivy.core.clipboard import Clipboard
            texto = Clipboard.paste()
            if texto:
                texto = texto.strip().strip('\n').strip('\r')
                texto = texto.strip('"').strip("'").strip('\u201c').strip('\u201d')
                self.ids.ruta_excel.text = texto
        except Exception:
            self.ids.estado_importar.text = "[color=#F87171]No se pudo acceder al portapapeles[/color]"

    def previsualizar(self) -> None:
        ruta = self._normalizar_ruta(self.ids.ruta_excel.text)
        if not ruta:
            self.ids.estado_importar.text = "[color=#F87171]Primero selecciona un archivo[/color]"
            return
        if not os.path.exists(ruta):
            self.ids.estado_importar.text = (
                f"[color=#F87171]Archivo no encontrado: {ruta}[/color]"
            )
            return
        self.ids.ruta_excel.text       = ruta
        self.ids.estado_importar.text  = "Leyendo archivo..."
        self.ids.btn_importar.disabled = True
        self.ids.preview_container.clear_widgets()
        threading.Thread(target=self._leer_excel_seguro, args=(ruta,), daemon=True).start()

    def _normalizar_ruta(self, ruta_raw) -> str:
        if not ruta_raw:
            return ""
        ruta = ruta_raw.strip()
        for ch in ('"', "'", '\u201c', '\u201d', '\u2018', '\u2019'):
            ruta = ruta.strip(ch)
        ruta = ruta.strip('\n').strip('\r').strip()
        ruta = os.path.expanduser(ruta)
        ruta = os.path.normpath(ruta)
        return ruta

    def _leer_excel_seguro(self, ruta):
        try:
            if load_workbook is None:
                Clock.schedule_once(lambda dt: self._set_estado("[color=#F87171]openpyxl no instalado[/color]"))
                return
            wb    = load_workbook(ruta, read_only=True, data_only=True)
            hojas = wb.sheetnames
            Clock.schedule_once(lambda dt: self._set_hojas(hojas))
            hoja_sel    = self.ids.hoja_spinner.text
            hoja_nombre = hoja_sel if hoja_sel in hojas else hojas[0]
            ws          = wb[hoja_nombre]
            filas       = list(ws.iter_rows(values_only=True))
            wb.close()
            if len(filas) < 2:
                Clock.schedule_once(lambda dt: self._set_estado("[color=#F87171]El archivo no tiene datos[/color]"))
                return
            encabezados = [str(h).strip() if h is not None else "" for h in filas[0]]
            datos       = filas[1:]
            Clock.schedule_once(lambda dt: self._mostrar_preview(encabezados, datos, hoja_nombre))
        except MemoryError:
            Clock.schedule_once(lambda dt: self._set_estado("[color=#F87171]Archivo demasiado grande[/color]"))
        except Exception as e:
            logger.error("Error leyendo Excel: %s", e)
            Clock.schedule_once(lambda dt: self._set_estado(f"[color=#F87171]Error al leer: {e}[/color]"))

    def _set_hojas(self, hojas):
        self.ids.hoja_spinner.values = hojas
        if self.ids.hoja_spinner.text == "-- detectar --" and hojas:
            self.ids.hoja_spinner.text = hojas[0]

    def _set_estado(self, texto):
        self.ids.estado_importar.text = texto

    def _mostrar_preview(self, encabezados, datos, hoja_nombre):
        self._columnas_excel = encabezados
        self.ids.preview_container.clear_widgets()
        mapeo_auto = {
            'col_nombre':   ['nombre','responsable','name','usuario','resp'],
            'col_material': ['material','producto','item','descripcion','description','mat'],
            'col_sku':      ['sku','id','codigo','code','ref'],
            'col_cantidad': ['cantidad','qty','quantity','cant'],
            'col_tipo':     ['tipo','type','movimiento','movement'],
            'col_fecha':    ['fecha','date','datetime','timestamp'],
        }
        enc_lower = [e.lower() for e in encabezados]
        for campo, alternativas in mapeo_auto.items():
            for alt in alternativas:
                if alt in enc_lower:
                    self.ids[campo].text = encabezados[enc_lower.index(alt)]
                    break

        card_enc = crear_tarjeta(dp(75), '#0F3460')
        card_enc.add_widget(Label(
            text=f"[b]Hoja:[/b] {hoja_nombre}  |  [b]Columnas:[/b] {len(encabezados)}",
            markup=True, color=get_color_from_hex('#38BDF8'), font_size='13sp',
            size_hint_y=None, height=dp(24),
            halign='left', valign='middle',
            text_size=(Window.width * 0.85, None)
        ))
        card_enc.add_widget(Label(
            text="  ".join(encabezados[:5]),
            font_size='11sp', color=get_color_from_hex('#94A3B8'),
            size_hint_y=None, height=dp(22),
            halign='left', valign='middle',
            text_size=(Window.width * 0.85, None),
            shorten=True, shorten_from='right'
        ))
        self.ids.preview_container.add_widget(card_enc)

        total = len(datos)
        for fila in datos[:10]:
            vals = [str(v) if v is not None else "-" for v in fila]
            card = crear_tarjeta(dp(50))
            card.add_widget(Label(
                text="  |  ".join(vals[:5]),
                font_size='11sp', color=get_color_from_hex('#CBD5E1'),
                size_hint_y=None, height=dp(28),
                halign='left', valign='middle',
                text_size=(Window.width * 0.85, None),
                shorten=True, shorten_from='right'
            ))
            self.ids.preview_container.add_widget(card)

        resumen = f"[color=#10B981]{total} fila(s) listas - revisa el mapeo y pulsa IMPORTAR[/color]"
        if total > 10:
            resumen += f"  [color=#94A3B8](mostrando 10 de {total})[/color]"
        self._set_estado(resumen)
        self.ids.btn_importar.disabled = False

    def importar(self) -> None:
        ruta = self._normalizar_ruta(self.ids.ruta_excel.text)
        if not ruta or not os.path.exists(ruta):
            self.ids.estado_importar.text = f"[color=#F87171]Archivo no encontrado.[/color]"
            return
        self.ids.btn_importar.disabled = True
        self.ids.estado_importar.text  = "Importando datos..."
        threading.Thread(target=self._importar_worker, args=(ruta,), daemon=True).start()

    def _importar_worker(self, ruta):
        try:
            if load_workbook is None:
                Clock.schedule_once(lambda dt: self._set_estado("[color=#F87171]openpyxl no instalado[/color]"))
                Clock.schedule_once(lambda dt: setattr(self.ids.btn_importar, 'disabled', False))
                return
            wb          = load_workbook(ruta, read_only=True, data_only=True)
            hoja_nombre = self.ids.hoja_spinner.text
            if hoja_nombre not in wb.sheetnames:
                hoja_nombre = wb.sheetnames[0]
            ws    = wb[hoja_nombre]
            filas = list(ws.iter_rows(values_only=True))
            wb.close()
            if len(filas) < 2:
                Clock.schedule_once(lambda dt: self._set_estado("[color=#F87171]No hay datos[/color]"))
                Clock.schedule_once(lambda dt: setattr(self.ids.btn_importar, 'disabled', False))
                return

            enc_upper = [str(h).strip().upper() if h is not None else "" for h in filas[0]]

            def idx(campo_id):
                nombre_col = self.ids[campo_id].text.strip().upper()
                try:
                    return enc_upper.index(nombre_col)
                except ValueError:
                    return None

            i_nombre   = idx('col_nombre')
            i_material = idx('col_material')
            i_sku      = idx('col_sku')
            i_cantidad = idx('col_cantidad')
            i_tipo     = idx('col_tipo')
            i_fecha    = idx('col_fecha')

            if i_material is None or i_cantidad is None:
                Clock.schedule_once(lambda dt: self._set_estado("[color=#F87171]Columnas 'Material' y 'Cantidad' requeridas[/color]"))
                Clock.schedule_once(lambda dt: setattr(self.ids.btn_importar, 'disabled', False))
                return

            insertados = errores = omitidos = 0
            ahora = datetime.now().strftime("%d/%m/%Y %H:%M")

            for fila in filas[1:]:
                try:
                    def celda(i):
                        if i is None or i >= len(fila):
                            return None
                        return fila[i]

                    nombre   = str(celda(i_nombre)   or "INVENTARIO").strip().upper()
                    material = str(celda(i_material) or "").strip().upper()
                    sku      = str(celda(i_sku)      or "").strip()
                    raw_cant = celda(i_cantidad)
                    raw_tipo = str(celda(i_tipo)     or "ENTRADA").strip().upper()
                    raw_fec  = celda(i_fecha)

                    if not material:
                        omitidos += 1; continue
                    try:
                        cantidad = float(str(raw_cant).replace(',', '.'))
                    except (TypeError, ValueError):
                        omitidos += 1; continue
                    if cantidad <= 0:
                        omitidos += 1; continue

                    if any(k in raw_tipo for k in ['ENT','IN','ING','COMP']):
                        tipo = "ENTRADA"
                    elif any(k in raw_tipo for k in ['SAL','OUT','EGR','RET']):
                        tipo = "SALIDA"
                    else:
                        tipo = "ENTRADA"

                    if raw_fec is None:
                        fecha = ahora
                    elif isinstance(raw_fec, datetime):
                        fecha = raw_fec.strftime("%d/%m/%Y %H:%M")
                    else:
                        fecha = str(raw_fec).strip() or ahora

                    _, stock_actual = obtener_stock_material(material)
                    nuevo_stock = round(stock_actual + (cantidad if tipo == "ENTRADA" else -cantidad), 4)
                    db_execute(
                        "INSERT INTO movimientos (nombre,material,sku,cantidad,tipo,fecha,stock_registro,dias,retorno,notas,ubicacion) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                        (nombre, material, sku, cantidad, tipo, fecha, nuevo_stock, 0, "N/A", "", "")
                    )
                    insertados += 1
                except Exception:
                    errores += 1

            log_auditoria("IMPORTAR_EXCEL", f"{insertados} importadas, {errores} errores, {omitidos} omitidas")
            resumen = (
                f"[color=#10B981]Importacion completada[/color]\n"
                f"Insertados: [b]{insertados}[/b]  |  "
                f"Omitidos: {omitidos}  |  Errores: {errores}"
            )
            Clock.schedule_once(lambda dt: self._finalizar(resumen))
        except MemoryError:
            Clock.schedule_once(lambda dt: self._set_estado("[color=#F87171]Memoria insuficiente[/color]"))
            Clock.schedule_once(lambda dt: setattr(self.ids.btn_importar, 'disabled', False))
        except Exception as e:
            logger.error("Error importando: %s", e)
            Clock.schedule_once(lambda dt: self._set_estado(f"[color=#F87171]Error: {e}[/color]"))
            Clock.schedule_once(lambda dt: setattr(self.ids.btn_importar, 'disabled', False))

    def _finalizar(self, resumen):
        self._set_estado(resumen)
        self.ids.btn_importar.disabled = False
        self.ids.preview_container.clear_widgets()
        try:
            inv = App.get_running_app().root.get_screen('inventario')
            inv.actualizar_spinners()
            inv._actualizar_dashboard()
            inv._cargar_actividad_reciente()
            inv._verificar_alertas_silencioso()
        except Exception:
            pass


# === screens/auditoria.py ===
from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.spinner import Spinner
from kivy.uix.scrollview import ScrollView
from kivy.metrics import dp
from kivy.utils import get_color_from_hex
from kivy.core.window import Window


class AuditoriaScreen(Screen):
    PAGE_SIZE = 30

    def on_enter(self):
        self._filtro_accion = ""
        self._filtro_usuario = ""
        self.cargar_log()

    def cargar_log(self, pagina=0, page_size=None):
        self.ids.log_container.clear_widgets()
        if page_size is None:
            page_size = self.PAGE_SIZE

        where = []
        params = []
        if self._filtro_accion:
            where.append("accion LIKE ?")
            params.append(f"%{self._filtro_accion}%")
        if self._filtro_usuario:
            where.append("usuario LIKE ?")
            params.append(f"%{self._filtro_usuario}%")

        sql_where = (" WHERE " + " AND ".join(where)) if where else ""
        try:
            count = db_query(
                f"SELECT COUNT(*) FROM auditoria{sql_where}", params, fetchall=False
            )
            total = count[0] if count else 0

            offset = pagina * page_size
            rows = db_query(
                f"SELECT id, accion, detalle, fecha, usuario FROM auditoria{sql_where} ORDER BY id DESC LIMIT ? OFFSET ?",
                params + [page_size, offset]
            )

            if not rows:
                lbl = Label(text="No hay registros de auditoria",
                            color=get_color_from_hex('#94A3B8'), size_hint_y=None, height=dp(60))
                self.ids.log_container.add_widget(lbl)
            else:
                for r in rows:
                    card = crear_tarjeta(dp(72), '#1A2035')
                    card.add_widget(Label(
                        text=f"[b]{r[1]}[/b]  [color=#64748B]{r[3]}[/color]",
                        markup=True, color=get_color_from_hex('#38BDF8'),
                        size_hint_y=None, height=dp(22), font_size='12sp',
                        halign='left', text_size=(Window.width * 0.82, None)
                    ))
                    card.add_widget(Label(
                        text=r[2] if len(r[2]) < 80 else r[2][:77] + "...",
                        color=get_color_from_hex('#CBD5E1'),
                        size_hint_y=None, height=dp(22), font_size='11sp',
                        halign='left', text_size=(Window.width * 0.82, None)
                    ))
                    card.add_widget(Label(
                        text=f"Usuario: {r[4] or 'sistema'}",
                        color=get_color_from_hex('#64748B'),
                        size_hint_y=None, height=dp(18), font_size='10sp',
                        halign='left', text_size=(Window.width * 0.82, None)
                    ))
                    self.ids.log_container.add_widget(card)

            if hasattr(self, 'ids') and 'pagination' in self.ids:
                self.ids.pagination.actualizar(total)

        except Exception as e:
            logger.error("Error cargando auditoria: %s", e)
            self.ids.log_container.add_widget(Label(
                text=f"Error: {e}", color=get_color_from_hex('#F87171'),
                size_hint_y=None, height=dp(40)
            ))

    def filtrar(self):
        self._filtro_accion = self.ids.filtro_accion.text.strip().upper()
        self._filtro_usuario = self.ids.filtro_usuario.text.strip().upper()
        self.cargar_log()


# === screens/alertas.py ===
from datetime import datetime, timedelta
from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.widget import Widget
from kivy.uix.popup import Popup
from kivy.clock import Clock
from kivy.metrics import dp
from kivy.utils import get_color_from_hex
from kivy.core.window import Window
from kivy.app import App


def _calcular_vencimiento(fecha_salida, dias_acordados):
    if not dias_acordados or dias_acordados <= 0:
        return "Sin vencimiento"
    try:
        fecha_dt = datetime.strptime(fecha_salida[:16], "%d/%m/%Y %H:%M")
        vence = fecha_dt + timedelta(days=dias_acordados)
        return vence.strftime("%d/%m/%Y")
    except Exception:
        return "Fecha invalida"


class AlertasScreen(Screen):
    def on_enter(self):
        self.cargar_alertas()

    def cargar_alertas(self):
        self.ids.container_alertas.clear_widgets()
        prestamos = obtener_equipos_en_prestamo()

        vencidos = [p for p in prestamos if p["estado"] == "VENCIDO"]
        por_vencer = [p for p in prestamos if p["estado"] == "POR_VENCER"]
        en_tiempo = [p for p in prestamos if p["estado"] == "EN_TIEMPO"]

        total_alertas = len(vencidos) + len(por_vencer)
        self.ids.total_alertas.text = f"Alertas: {total_alertas}"

        if total_alertas == 0 and not en_tiempo:
            card = crear_tarjeta(dp(70), '#0D1B2A')
            card.add_widget(Label(
                text="[color=#00D4FF]Todo en orden[/color]",
                markup=True, font_size='15sp', halign='center'))
            card.add_widget(Label(
                text="No hay materiales pendientes de retorno",
                color=get_color_from_hex('#5A6A7A'), font_size='12sp', halign='center'))
            self.ids.container_alertas.add_widget(card)
            return

        if vencidos:
            self.ids.container_alertas.add_widget(Label(
                text=f"VENCIDOS ({len(vencidos)})",
                bold=True, font_size='14sp',
                color=get_color_from_hex('#FF0044'),
                size_hint_y=None, height=dp(32),
                halign='left', text_size=(Window.width * 0.88, None)))

            for p in vencidos:
                self._agregar_tarjeta_prestamo(p, '#0D1B2A', '#FF0044')

        if por_vencer:
            self.ids.container_alertas.add_widget(Label(
                text=f"POR VENCER ({len(por_vencer)})",
                bold=True, font_size='14sp',
                color=get_color_from_hex('#FF0044'),
                size_hint_y=None, height=dp(32),
                halign='left', text_size=(Window.width * 0.88, None)))

            for p in por_vencer:
                self._agregar_tarjeta_prestamo(p, '#0D1B2A', '#FF0044')

        if en_tiempo:
            self.ids.container_alertas.add_widget(Label(
                text=f"EN TIEMPO ({len(en_tiempo)})",
                bold=True, font_size='13sp',
                color=get_color_from_hex('#5A6A7A'),
                size_hint_y=None, height=dp(28),
                halign='left', text_size=(Window.width * 0.88, None)))

            for p in en_tiempo:
                self._agregar_tarjeta_prestamo(p, '#0D1B2A', '#00D4FF')

    def _agregar_tarjeta_prestamo(self, p, color_fondo, color_estado):
        tiene_notas = bool(p["notas"] and p["notas"].strip())
        vence = _calcular_vencimiento(p["fecha_salida"], p["dias_acordados"])
        altura = dp(130) if tiene_notas else dp(108)
        card = crear_tarjeta(altura, color_fondo)

        header = BoxLayout(size_hint_y=None, height=dp(24), spacing=dp(6))
        header.add_widget(Label(
            text=f"[b]{p['material']}[/b]  —  {p['responsable']}",
            markup=True, color=get_color_from_hex(color_estado),
            font_size='12sp', halign='left', valign='middle',
            text_size=(Window.width * 0.55, None),
            size_hint_x=0.6, shorten=True, shorten_from='right'))
        header.add_widget(Label(
            text=f"[color={color_estado}]{p['estado']}[/color]",
            markup=True, font_size='11sp', bold=True,
            size_hint_x=0.2, halign='center'))
        header.add_widget(Label(
            text=f"{p['dias_fuera']}d",
            color=get_color_from_hex('#5A6A7A'), font_size='11sp',
            size_hint_x=0.15, halign='center'))
        card.add_widget(header)

        sub = BoxLayout(size_hint_y=None, height=dp(20), spacing=dp(6))
        sub.add_widget(Label(
            text=f"Cant: {round(p['cantidad'], 2)}  |  SKU: {p['sku'] or 'N/A'}",
            color=get_color_from_hex('#5A6A7A'), font_size='11sp',
            halign='left', text_size=(Window.width * 0.55, None),
            size_hint_x=0.6))
        vence_label = vence if "Sin" in vence else f"Vence: {vence}"
        sub.add_widget(Label(
            text=vence_label,
            color=get_color_from_hex('#FF0044') if "Vence" in vence_label else get_color_from_hex('#5A6A7A'),
            font_size='10sp', size_hint_x=0.4, halign='center'))
        card.add_widget(sub)

        sub2 = BoxLayout(size_hint_y=None, height=dp(18), spacing=dp(6))
        sub2.add_widget(Label(
            text=f"Salida: {p['fecha_salida'][:16]}",
            color=get_color_from_hex('#5A6A7A'), font_size='10sp',
            halign='left', text_size=(Window.width * 0.55, None),
            size_hint_x=0.6))
        dias_txt = f"Acordado: {p['dias_acordados']}d" if p['dias_acordados'] > 0 else "Sin plazo"
        sub2.add_widget(Label(
            text=dias_txt,
            color=get_color_from_hex('#5A6A7A'), font_size='10sp',
            size_hint_x=0.4, halign='center'))
        card.add_widget(sub2)

        if tiene_notas:
            card.add_widget(Label(
                text=f"Notas: {p['notas']}",
                color=get_color_from_hex('#5A6A7A'),
                size_hint_y=None, height=dp(16), font_size='10sp',
                halign='left', text_size=(Window.width * 0.82, None)))

        btn_row = BoxLayout(size_hint_y=None, height=dp(28), spacing=dp(6))
        btn_row.add_widget(Widget(size_hint_x=0.6))
        btn_retornar = Button(
            text="RETORNAR", size_hint_x=0.4,
            font_size='10sp', bold=True,
            background_color=get_color_from_hex('#0066FF'))
        btn_retornar.bind(on_release=lambda x, item=p: self.confirmar_retorno(item))
        btn_row.add_widget(btn_retornar)
        card.add_widget(btn_row)

        self.ids.container_alertas.add_widget(card)

    def confirmar_retorno(self, p):
        content = BoxLayout(orientation='vertical', padding=dp(12), spacing=dp(10))
        pop = Popup(title="Registrar Retorno", content=content, size_hint=(0.88, 0.45))
        content.add_widget(Label(
            text=f"Reingresar [b]{p['material']}[/b]  x{round(p['cantidad'], 2)}\nde [color=#FF0044]{p['responsable']}[/color]?",
            markup=True, halign='center', font_size='13sp',
            size_hint_y=None, height=dp(50)))
        btn = Button(text="SI, REGISTRAR RETORNO",
                     background_color=get_color_from_hex('#0066FF'),
                     bold=True, size_hint_y=None, height=dp(48))
        btn.bind(on_release=lambda x: self._ejecutar_retorno(p, pop))
        content.add_widget(btn)
        content.add_widget(Button(text="CANCELAR", size_hint_y=None, height=dp(42),
                                   background_color=get_color_from_hex('#334155'),
                                   on_release=pop.dismiss))
        pop.open()

    def _ejecutar_retorno(self, p, pop):
        pop.dismiss()
        App.get_running_app().root.get_screen('inventario').procesar(
            "ENTRADA", r=p['responsable'], p=p['material'], c=p['cantidad'], s=p['sku'], desde_retorno=True)
        Clock.schedule_once(lambda dt: self.cargar_alertas(), 0.3)


# === screens/categorias.py ===
from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.spinner import Spinner
from kivy.uix.scrollview import ScrollView
from kivy.metrics import dp
from kivy.utils import get_color_from_hex
from kivy.core.window import Window


class CategoriasScreen(Screen):
    def on_enter(self):
        self.cargar_categorias()

    def cargar_categorias(self):
        self.ids.lista_categorias.clear_widgets()
        rows = db_query("SELECT id, nombre, descripcion, color_hex FROM categorias ORDER BY nombre")
        if not rows:
            lbl = Label(text="Sin categorias - crea una con + NUEVA",
                        color=get_color_from_hex('#94A3B8'),
                        size_hint_y=None, height=dp(60))
            self.ids.lista_categorias.add_widget(lbl)
            return
        for r in rows:
            items = db_query("SELECT nombre FROM categoria_items WHERE categoria_id = ? ORDER BY nombre", (r[0],))
            item_count = len(items)
            item_names = ", ".join([it[0] for it in items[:5]])
            if len(items) > 5:
                item_names += f" ... (+{len(items)-5})"
            card_height = dp(80) + (min(len(items), 5) * dp(22) if items else 0)
            card = crear_tarjeta(card_height, '#1A2035')

            header = BoxLayout(size_hint_y=None, height=dp(26))
            header.add_widget(Label(
                text=f"[b]{r[1]}[/b]  [color=#64748B]({item_count} items)[/color]",
                markup=True, color=get_color_from_hex(r[3] or '#38BDF8'),
                font_size='13sp', halign='left',
                text_size=(Window.width * 0.5, None), size_hint_x=0.55))
            btn_items = Button(text="ITEMS", size_hint_x=0.18,
                               font_size='11sp',
                               background_color=get_color_from_hex('#0E7490'))
            btn_items.bind(on_release=lambda x, cid=r[0], cnom=r[1]: self._gestionar_items(cid, cnom))
            header.add_widget(btn_items)
            btn_del = Button(text="X", size_hint_x=0.12,
                             font_size='13sp', bold=True,
                             background_color=get_color_from_hex('#991B1B'))
            btn_del.bind(on_release=lambda x, cid=r[0]: self._eliminar(cid))
            header.add_widget(btn_del)
            card.add_widget(header)

            if items:
                for it in items[:5]:
                    card.add_widget(Label(
                        text=f"  • {it[0]}",
                        color=get_color_from_hex('#94A3B8'),
                        size_hint_y=None, height=dp(20), font_size='11sp',
                        halign='left', text_size=(Window.width * 0.8, None)))

            self.ids.lista_categorias.add_widget(card)

    def mostrar_nueva(self):
        content = BoxLayout(orientation='vertical', spacing=dp(10), padding=dp(15))
        pop = Popup(title="NUEVA CATEGORIA", content=content, size_hint=(0.88, 0.48))
        inp_nombre = TextInput(hint_text="Nombre (ej: H, ANGULO, PLATINA)", multiline=False,
                               size_hint_y=None, height=dp(45))
        inp_desc = TextInput(hint_text="Descripcion (opcional)", multiline=False,
                             size_hint_y=None, height=dp(45))
        colores = ["#38BDF8", "#F59E0B", "#10B981", "#F87171", "#A78BFA", "#F472B6", "#34D399"]
        spinner_color = Spinner(text=colores[0], values=colores, size_hint_y=None, height=dp(45))
        content.add_widget(Label(text="Nombre", color=get_color_from_hex('#94A3B8'),
                                  font_size='12sp', size_hint_y=None, height=dp(18)))
        content.add_widget(inp_nombre)
        content.add_widget(inp_desc)
        content.add_widget(spinner_color)

        def guardar(btn):
            nombre = inp_nombre.text.strip().upper()
            if not nombre:
                inp_nombre.text = "Obligatorio"
                return
            try:
                db_execute("INSERT INTO categorias (nombre, descripcion, color_hex) VALUES (?,?,?)",
                           (nombre, inp_desc.text.strip(), spinner_color.text))
                pop.dismiss()
                self.cargar_categorias()
            except Exception as e:
                inp_nombre.text = f"Error: {e}"

        btn = Button(text="CREAR", size_hint_y=None, height=dp(48),
                     background_color=get_color_from_hex('#059669'), on_release=guardar)
        content.add_widget(btn)
        content.add_widget(Button(text="CANCELAR", size_hint_y=None, height=dp(40),
                                   background_color=get_color_from_hex('#334155'),
                                   on_release=pop.dismiss))
        pop.open()

    def _gestionar_items(self, cid, cnom):
        content = BoxLayout(orientation='vertical', spacing=dp(8), padding=dp(12))
        pop = Popup(title=f"ITEMS: {cnom}", content=content, size_hint=(0.92, 0.78))

        inp = TextInput(hint_text="Nuevo item (ej: H 80)", multiline=False,
                        size_hint_y=None, height=dp(42))
        btn_add = Button(text="AGREGAR", size_hint_y=None, height=dp(40),
                         background_color=get_color_from_hex('#059669'))
        input_row = BoxLayout(size_hint_y=None, height=dp(42), spacing=dp(6))
        input_row.add_widget(inp)
        input_row.add_widget(btn_add)
        content.add_widget(input_row)

        sv = ScrollView()
        gl = GridLayout(cols=1, size_hint_y=None, spacing=dp(4))
        gl.bind(minimum_height=gl.setter('height'))

        def refrescar():
            gl.clear_widgets()
            items = db_query("SELECT id, nombre FROM categoria_items WHERE categoria_id = ? ORDER BY nombre", (cid,))
            if not items:
                gl.add_widget(Label(text="Sin items todavia", color=get_color_from_hex('#64748B'),
                                     size_hint_y=None, height=dp(40)))
            for it in items:
                row = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(6))
                row.add_widget(Label(
                    text=f"  {it[1]}",
                    color=get_color_from_hex('#CBD5E1'),
                    font_size='12sp', halign='left',
                    text_size=(Window.width * 0.7, None),
                    size_hint_x=0.8))
                btn_rm = Button(text="X", size_hint_x=0.2,
                                font_size='12sp',
                                background_color=get_color_from_hex('#991B1B'))
                btn_rm.bind(on_release=lambda x, iid=it[0]: (_eliminar_item(iid), refrescar()))
                row.add_widget(btn_rm)
                gl.add_widget(row)

        def _eliminar_item(iid):
            db_execute("DELETE FROM categoria_items WHERE id = ?", (iid,))

        def agregar():
            nombre = inp.text.strip().upper()
            if nombre:
                try:
                    db_execute("INSERT INTO categoria_items (categoria_id, nombre) VALUES (?,?)",
                               (cid, nombre))
                    inp.text = ""
                    refrescar()
                except Exception as e:
                    inp.text = f"Ya existe o error"
            inp.focus = True

        btn_add.bind(on_release=lambda x: agregar())
        inp.bind(on_text_validate=lambda x: agregar())
        sv.add_widget(gl)
        content.add_widget(sv)

        btn_cerrar = Button(text="CERRAR", size_hint_y=None, height=dp(44),
                            background_color=get_color_from_hex('#334155'),
                            on_release=pop.dismiss)
        content.add_widget(btn_cerrar)

        refrescar()
        pop.open()

    def _eliminar(self, cid):
        db_execute("DELETE FROM categoria_items WHERE categoria_id = ?", (cid,))
        db_execute("DELETE FROM categorias WHERE id = ?", (cid,))
        self.cargar_categorias()


# === screens/ubicaciones.py ===
from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.metrics import dp
from kivy.utils import get_color_from_hex
from kivy.core.window import Window


class UbicacionesScreen(Screen):
    def on_enter(self):
        self.cargar_ubicaciones()

    def cargar_ubicaciones(self):
        self.ids.lista_ubicaciones.clear_widgets()
        rows = db_query("SELECT id, nombre FROM ubicaciones ORDER BY nombre")
        if not rows:
            lbl = Label(text="Sin ubicaciones", color=get_color_from_hex('#94A3B8'),
                        size_hint_y=None, height=dp(60))
            self.ids.lista_ubicaciones.add_widget(lbl)
            return
        for r in rows:
            card = crear_tarjeta(dp(56), '#1A2035')
            row = BoxLayout(size_hint_y=None, height=dp(30), spacing=dp(8))
            row.add_widget(Label(
                text=f"[b]{r[1]}[/b]",
                markup=True, color=get_color_from_hex('#38BDF8'),
                font_size='13sp', halign='left',
                text_size=(Window.width * 0.6, None), size_hint_x=0.7))
            btn = Button(text="ELIMINAR", size_hint_x=0.3,
                         font_size='11sp',
                         background_color=get_color_from_hex('#991B1B'))
            btn.bind(on_release=lambda x, uid=r[0]: self._eliminar(uid))
            row.add_widget(btn)
            card.add_widget(row)
            self.ids.lista_ubicaciones.add_widget(card)

    def mostrar_nueva(self):
        inp = TextInput(hint_text="Nombre de la ubicacion", multiline=False,
                        size_hint_y=None, height=dp(48))
        content = BoxLayout(orientation='vertical', spacing=dp(10), padding=dp(15))
        content.add_widget(inp)
        pop = Popup(title="NUEVA UBICACION", content=content, size_hint=(0.85, 0.32))

        def guardar(btn):
            if inp.text.strip():
                db_execute("INSERT INTO ubicaciones (nombre) VALUES (?)", (inp.text.strip().upper(),))
            pop.dismiss()
            self.cargar_ubicaciones()

        btn = Button(text="CREAR", size_hint_y=None, height=dp(48),
                     background_color=get_color_from_hex('#059669'), on_release=guardar)
        content.add_widget(btn)
        pop.open()

    def _eliminar(self, uid):
        db_execute("DELETE FROM ubicaciones WHERE id = ?", (uid,))
        self.cargar_ubicaciones()


# === screens/gestion.py ===
from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
from kivy.metrics import dp
from kivy.utils import get_color_from_hex
from kivy.core.window import Window


class GestionScreen(Screen):
    def on_enter(self):
        self._modo = "materiales"
        self.cargar()

    def cargar(self, filtro=""):
        self.ids.lista_gestion.clear_widgets()
        filtro = filtro.upper().strip()

        if self._modo == "materiales":
            self.ids.titulo_seccion.text = "MATERIALES"
            rows = db_query(
                "SELECT DISTINCT UPPER(material) FROM movimientos ORDER BY material"
            )
            for (mat,) in rows:
                if filtro and filtro not in mat:
                    continue
                fijo, disp = obtener_stock_material(mat)
                card = crear_tarjeta(dp(72), '#0D1B2A')
                card.add_widget(Label(
                    text=f"[b]{mat}[/b]",
                    markup=True, color=get_color_from_hex('#00D4FF'),
                    size_hint_y=None, height=dp(22), font_size='13sp',
                    halign='left', text_size=(Window.width * 0.5, None)))

                estado_color = '#00D4FF' if disp > 0 else '#FF0044'
                card.add_widget(Label(
                    text=f"Stock: [color={estado_color}]{disp}[/color]  |  Total sistema: {fijo}",
                    markup=True, color=get_color_from_hex('#5A6A7A'),
                    size_hint_y=None, height=dp(18), font_size='11sp',
                    halign='left', text_size=(Window.width * 0.5, None)))

                total_movs = db_query(
                    "SELECT COUNT(*) FROM movimientos WHERE UPPER(material)=?",
                    (mat,), fetchall=False
                )[0] or 0

                btn_row = BoxLayout(size_hint_y=None, height=dp(26), spacing=dp(6))
                btn_limpiar = Button(
                    text="ELIMINAR MATERIAL", size_hint_x=0.5,
                    font_size='10sp', bold=True,
                    background_color=get_color_from_hex('#FF0044'))
                btn_limpiar.bind(on_release=lambda x, m=mat, c=total_movs: self._confirmar_eliminar("material", m, c))
                btn_row.add_widget(btn_limpiar)
                btn_row.add_widget(Label(
                    text=f"{total_movs} movs",
                    color=get_color_from_hex('#5A6A7A'), font_size='10sp',
                    size_hint_x=0.5, halign='center'))
                card.add_widget(btn_row)
                self.ids.lista_gestion.add_widget(card)

        else:
            self.ids.titulo_seccion.text = "RESPONSABLES"
            rows = db_query(
                "SELECT DISTINCT UPPER(nombre) FROM movimientos WHERE UPPER(nombre) != 'INVENTARIO' ORDER BY nombre"
            )
            for (nom,) in rows:
                if filtro and filtro not in nom:
                    continue
                total_movs = db_query(
                    "SELECT COUNT(*) FROM movimientos WHERE UPPER(nombre)=?",
                    (nom,), fetchall=False
                )[0] or 0
                ultima_fecha = db_query(
                    "SELECT fecha FROM movimientos WHERE UPPER(nombre)=? ORDER BY id DESC LIMIT 1",
                    (nom,), fetchall=False
                )
                ultimo = str(ultima_fecha[0]) if ultima_fecha else "N/A"

                card = crear_tarjeta(dp(72), '#0D1B2A')
                card.add_widget(Label(
                    text=f"[b]{nom}[/b]",
                    markup=True, color=get_color_from_hex('#FF0044'),
                    size_hint_y=None, height=dp(22), font_size='13sp',
                    halign='left', text_size=(Window.width * 0.5, None)))
                card.add_widget(Label(
                    text=f"Ultimo movimiento: {ultimo}",
                    color=get_color_from_hex('#5A6A7A'),
                    size_hint_y=None, height=dp(18), font_size='11sp',
                    halign='left', text_size=(Window.width * 0.5, None)))

                btn_row = BoxLayout(size_hint_y=None, height=dp(26), spacing=dp(6))
                btn_limpiar = Button(
                    text="ELIMINAR RESPONSABLE", size_hint_x=0.5,
                    font_size='10sp', bold=True,
                    background_color=get_color_from_hex('#FF0044'))
                btn_limpiar.bind(on_release=lambda x, n=nom, c=total_movs: self._confirmar_eliminar("responsable", n, c))
                btn_row.add_widget(btn_limpiar)
                btn_row.add_widget(Label(
                    text=f"{total_movs} movs",
                    color=get_color_from_hex('#5A6A7A'), font_size='10sp',
                    size_hint_x=0.5, halign='center'))
                card.add_widget(btn_row)
                self.ids.lista_gestion.add_widget(card)

    def cambiar_modo(self, modo):
        self._modo = modo
        self.ids.filtro_gestion.text = ""
        self.cargar()

    def filtrar(self):
        self.cargar(self.ids.filtro_gestion.text)

    def _confirmar_eliminar(self, tipo, nombre, total_movs):
        tipo_txt = "material" if tipo == "material" else "responsable"
        content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(10))
        pop = Popup(
            title=f"ELIMINAR {tipo_txt.upper()}",
            content=content, size_hint=(0.88, 0.42))
        content.add_widget(Label(
            text=f"Eliminar [b]{nombre}[/b] y sus {total_movs} movimiento(s)?",
            markup=True, halign='center', color=get_color_from_hex('#FF0044'),
            font_size='13sp'))
        content.add_widget(Label(
            text="Esta accion no se puede deshacer.",
            color=get_color_from_hex('#5A6A7A'), font_size='11sp',
            halign='center'))

        def ejecutar(btn):
            try:
                if tipo == "material":
                    db_execute("DELETE FROM movimientos WHERE UPPER(material)=?", (nombre,))
                    db_execute("DELETE FROM alertas_stock WHERE UPPER(material)=?", (nombre,))
                    log_auditoria("ELIMINAR_MATERIAL", f"{nombre}: {total_movs} movs eliminados")
                else:
                    db_execute("DELETE FROM movimientos WHERE UPPER(nombre)=?", (nombre,))
                    log_auditoria("ELIMINAR_RESPONSABLE", f"{nombre}: {total_movs} movs eliminados")
                pop.dismiss()
                Toast.show(f"{nombre} eliminado", '#FF0044', 3)
                self.cargar(self.ids.filtro_gestion.text)
            except Exception as e:
                pop.dismiss()
                Toast.show(f"Error: {e}", '#FF0044', 4)

        btn = Button(text="SI, ELIMINAR TODO",
                     size_hint_y=None, height=dp(48),
                     background_color=get_color_from_hex('#FF0044'),
                     on_release=ejecutar)
        content.add_widget(btn)
        content.add_widget(Button(text="CANCELAR", size_hint_y=None, height=dp(40),
                                   background_color=get_color_from_hex('#334155'),
                                   on_release=pop.dismiss))
        pop.open()


# === screens/pendientes.py ===
from collections import defaultdict
from kivy.app import App
from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
from kivy.clock import Clock
from kivy.metrics import dp
from kivy.utils import get_color_from_hex
from kivy.graphics import Color, RoundedRectangle
from kivy.core.window import Window


class PendientesScreen(Screen):
    def on_enter(self):
        self.cargar_pendientes()

    def cargar_pendientes(self) -> None:
        try:
            self.ids.lista_pendientes.clear_widgets()
            filtro = self.ids.filtro_pendientes.text.upper()
            rows = db_query(
                "SELECT nombre,material,sku,cantidad,tipo,fecha,notas FROM movimientos "
                "WHERE UPPER(nombre) != 'INVENTARIO' ORDER BY id ASC"
            )
            actualmente_fuera = {}
            for n, m, s, c, t, f, notas in rows:
                key = f"{n}|{m}"
                if t == "SALIDA":
                    if key not in actualmente_fuera:
                        actualmente_fuera[key] = [n, m, s, float(c), t, f, notas or ""]
                    else:
                        actualmente_fuera[key][3] += float(c)
                        nota_existente = actualmente_fuera[key][6]
                        nota_nueva = notas or ""
                        if nota_nueva and nota_nueva not in nota_existente:
                            actualmente_fuera[key][6] = (
                                (nota_existente + " | " + nota_nueva).strip(" | ")
                                if nota_existente else nota_nueva
                            )
                elif t == "ENTRADA" and key in actualmente_fuera:
                    actualmente_fuera[key][3] -= float(c)
                    if actualmente_fuera[key][3] <= 0:
                        del actualmente_fuera[key]

            grupos = defaultdict(list)
            for item in actualmente_fuera.values():
                if filtro and filtro not in item[0].upper() and filtro not in item[1].upper():
                    continue
                grupos[item[0]].append(item)

            if not grupos:
                card = crear_tarjeta(dp(70), '#1E293B')
                card.add_widget(Label(text="No hay materiales fuera",
                                      color=get_color_from_hex('#10B981')))
                self.ids.lista_pendientes.add_widget(card)
                return

            for responsable, materiales in sorted(grupos.items()):
                extra_por_item = [dp(75) if (it[6] and it[6].strip()) else dp(58) for it in materiales]
                btns_height = dp(32) if len(materiales) > 1 else 0
                card_h = dp(50) + sum(extra_por_item) + btns_height
                card = crear_tarjeta(card_h, '#1A2035')

                header = BoxLayout(size_hint_y=None, height=dp(30))
                header.add_widget(Label(
                    text=f"[b]{responsable}[/b]  ({len(materiales)} material{'es' if len(materiales)>1 else ''})",
                    markup=True, color=get_color_from_hex('#FF0044'),
                    size_hint_y=None, height=dp(30),
                    halign='center', valign='middle',
                    text_size=(Window.width * 0.9, None),
                    size_hint_x=1, shorten=True, shorten_from='right'))
                if len(materiales) > 1:
                    btn_entrar_todo = Button(
                        text="ENTRAR TODO", size_hint_x=0.4,
                        font_size='11sp', bold=True,
                        background_color=get_color_from_hex('#059669'))
                    btn_entrar_todo.bind(on_release=lambda x, r=responsable, mats=materiales: self.confirmar_entrar_todo(r, mats))
                    header.add_widget(btn_entrar_todo)
                else:
                    header.add_widget(Label(size_hint_x=0.4))
                card.add_widget(header)

                for it in materiales:
                    n, m, s, c, _, f, notas = it
                    tiene_notas = bool(notas and notas.strip())
                    item_box = BoxLayout(
                        orientation='vertical',
                        size_hint_y=None,
                        height=dp(72) if tiene_notas else dp(55),
                        spacing=dp(2)
                    )
                    with item_box.canvas.before:
                        Color(rgba=get_color_from_hex('#334155'))
                        item_box._rect = RoundedRectangle(
                            pos=item_box.pos, size=item_box.size, radius=[dp(8)]
                        )

                    def _make_rect_updater(rect):
                        def upd_pos(i, v):
                            rect.pos = v
                        def upd_size(i, v):
                            rect.size = v
                        return upd_pos, upd_size

                    _upd_pos, _upd_size = _make_rect_updater(item_box._rect)
                    item_box.bind(pos=_upd_pos, size=_upd_size)

                    btn_text = (
                        f"[b]{m}[/b]   x{round(c,2)}\n"
                        f"[size=11sp][color=#94A3B8]SKU: {s or 'N/A'}  |  Salida: {f}[/color][/size]"
                    )
                    btn = Button(
                        text=btn_text, markup=True, size_hint_y=None, height=dp(55),
                        background_color=(0, 0, 0, 0),
                        halign='left', valign='middle', padding=[dp(8), 0])
                    btn.bind(on_release=lambda x, i=it: self.confirmar_retorno_individual(i))
                    item_box.add_widget(btn)

                    if tiene_notas:
                        notas_box = BoxLayout(
                            orientation='horizontal',
                            size_hint_y=None, height=dp(22),
                            padding=[dp(10), 0])
                        with notas_box.canvas.before:
                            Color(rgba=get_color_from_hex('#1E3A5F'))
                            notas_box._rect2 = RoundedRectangle(
                                pos=notas_box.pos, size=notas_box.size,
                                radius=[dp(0), dp(0), dp(8), dp(8)])

                        def _make_rect2_updater(rect2):
                            def upd_pos2(i, v):
                                rect2.pos = v
                            def upd_size2(i, v):
                                rect2.size = v
                            return upd_pos2, upd_size2

                        _upd_pos2, _upd_size2 = _make_rect2_updater(notas_box._rect2)
                        notas_box.bind(pos=_upd_pos2, size=_upd_size2)
                        notas_box.add_widget(Label(
                            text=f"[color=#64748B]NOTAS:[/color] [color=#CBD5E1]{notas}[/color]",
                            markup=True, font_size='11sp',
                            halign='left', valign='middle',
                            text_size=(Window.width * 0.82, None)))
                        item_box.add_widget(notas_box)
                    card.add_widget(item_box)
                self.ids.lista_pendientes.add_widget(card)
        except Exception as e:
            logger.error("Error cargando pendientes: %s", e)

    def confirmar_retorno_individual(self, item):
        n, m, s, c, _, f, notas = item
        content = BoxLayout(orientation='vertical', spacing=dp(8), padding=dp(12))
        pop = Popup(title="RETORNAR MATERIAL", content=content, size_hint=(0.9, 0.55))

        content.add_widget(Label(
            text=f"[b]{m}[/b]  x[color=#F59E0B]{round(c, 2)}[/color]  de [color=#F87171]{n}[/color]",
            markup=True, font_size='14sp', halign='center', valign='middle',
            size_hint_y=None, height=dp(34),
            text_size=(Window.width * 0.8, None), shorten=True, shorten_from='right'))
        content.add_widget(Label(
            text=f"SKU: {s or 'N/A'}  |  Salida: {f[:16]}",
            color=get_color_from_hex('#64748B'), font_size='11sp',
            size_hint_y=None, height=dp(20), halign='center'))

        info = Label(
            text="", color=get_color_from_hex('#F59E0B'),
            size_hint_y=None, height=dp(22), font_size='12sp', halign='center')
        content.add_widget(info)

        inp = TextInput(
            hint_text=f"Cantidad a retornar (max {round(c,2)})",
            input_filter='float', multiline=False,
            size_hint_y=None, height=dp(48), font_size='16sp')
        content.add_widget(inp)

        def retornar_todo(btn):
            pop.dismiss()
            self._ejecutar_retorno(item, round(c, 2))

        def retornar_parcial(btn):
            val_str = inp.text.strip()
            if not val_str:
                info.text = "Ingresa una cantidad"
                return
            try:
                val = float(val_str)
            except ValueError:
                info.text = "Cantidad invalida"
                return
            if val <= 0:
                info.text = "Debe ser mayor a 0"
                return
            if val > c:
                info.text = f"Maximo {round(c, 2)}"
                return
            pop.dismiss()
            self._ejecutar_retorno(item, round(val, 4))

        btns = BoxLayout(size_hint_y=None, height=dp(48), spacing=dp(8))
        btn_todo = Button(
            text=f"ENTRAR TODO ({round(c, 2)})",
            font_size='13sp', bold=True,
            background_color=get_color_from_hex('#059669'))
        btn_todo.bind(on_release=retornar_todo)
        btns.add_widget(btn_todo)
        btn_parcial = Button(
            text="ENTRAR CANTIDAD",
            font_size='11sp',
            background_color=get_color_from_hex('#2563EB'))
        btn_parcial.bind(on_release=retornar_parcial)
        btns.add_widget(btn_parcial)
        content.add_widget(btns)
        content.add_widget(Button(text="CANCELAR", size_hint_y=None, height=dp(40),
                                   background_color=get_color_from_hex('#334155'),
                                   on_release=pop.dismiss))
        pop.open()

    def confirmar_entrar_todo(self, responsable, materiales):
        content = BoxLayout(orientation='vertical', spacing=dp(10), padding=dp(15))
        pop = Popup(title=f"RETORNAR TODO: {responsable}", content=content, size_hint=(0.88, 0.45))
        content.add_widget(Label(
            text=f"Entrar todos los materiales de [b]{responsable}[/b]?",
            markup=True, font_size='13sp', halign='center',
            size_hint_y=None, height=dp(30)))
        for it in materiales:
            content.add_widget(Label(
                text=f"  {it[1]}  x{round(it[3], 2)}",
                color=get_color_from_hex('#94A3B8'),
                size_hint_y=None, height=dp(20), font_size='11sp',
                halign='left', text_size=(Window.width * 0.7, None)))

        def ejecutar(btn):
            pop.dismiss()
            for it in materiales:
                self._ejecutar_retorno(it, round(it[3], 2))
            Clock.schedule_once(lambda dt: self.cargar_pendientes(), 0.3)

        btn = Button(text="SI, ENTRAR TODO",
                     size_hint_y=None, height=dp(48),
                     background_color=get_color_from_hex('#059669'),
                     on_release=ejecutar)
        content.add_widget(btn)
        content.add_widget(Button(text="CANCELAR", size_hint_y=None, height=dp(40),
                                   background_color=get_color_from_hex('#334155'),
                                   on_release=pop.dismiss))
        pop.open()

    def _ejecutar_retorno(self, item, cantidad):
        n, m, s, c, _, _, _ = item
        App.get_running_app().root.get_screen('inventario').procesar(
            "ENTRADA", r=n, p=m, c=cantidad, s=s, desde_retorno=True)


# === screens/filtros.py ===
from kivy.uix.screenmanager import Screen
from kivy.uix.label import Label
from kivy.metrics import dp
from kivy.utils import get_color_from_hex
from kivy.core.window import Window


class FiltroScreen(Screen):
    def on_enter(self):
        self.filtrar()

    def filtrar(self) -> None:
        try:
            self.ids.c_r.clear_widgets()
            bn   = self.ids.f_n.text.upper()
            bp   = self.ids.f_p.text.upper()
            tipo = self.ids.filtro_tipo.text
            if len(bn) < 2 and len(bp) < 2:
                self.ids.filtro_count.text = "Escribe al menos 2 caracteres para buscar"
                return
            sql    = "SELECT material,tipo,cantidad,nombre,fecha,sku FROM movimientos WHERE UPPER(nombre) LIKE ? AND UPPER(material) LIKE ?"
            params = [f'%{bn}%', f'%{bp}%']
            if tipo in ("ENTRADA", "SALIDA"):
                sql += " AND tipo=?"
                params.append(tipo)
            sql += " ORDER BY id DESC LIMIT 200"
            rows = db_query(sql, params)
            self.ids.filtro_count.text = f"{len(rows)} resultado(s)"
            for m, t, c, n, f, s in rows:
                color_tipo = '#10B981' if t == "ENTRADA" else '#F87171'
                card = crear_tarjeta(dp(90))
                card.add_widget(Label(text=f"[b]{m}[/b]  [color={color_tipo}]{t}[/color]",
                                      markup=True, color=get_color_from_hex('#38BDF8'),
                                      size_hint_y=None, height=dp(26), font_size='12sp',
                                      halign='left', valign='middle',
                                      text_size=(Window.width * 0.82, None)))
                card.add_widget(Label(text=f"Cant: {c} | SKU: {s or 'N/A'}",
                                      color=get_color_from_hex('#94A3B8'),
                                      size_hint_y=None, height=dp(22), font_size='11sp',
                                      halign='left', valign='middle',
                                      text_size=(Window.width * 0.82, None)))
                card.add_widget(Label(text=f"{n}  *  {f}",
                                      font_size='11sp', color=get_color_from_hex('#94A3B8'),
                                      size_hint_y=None, height=dp(22),
                                      halign='left', valign='middle',
                                      text_size=(Window.width * 0.82, None)))
                self.ids.c_r.add_widget(card)
        except Exception as e:
            logger.error("Error filtrando: %s", e)


# === screens/stock.py ===
from kivy.uix.screenmanager import Screen
from kivy.uix.label import Label
from kivy.metrics import dp
from kivy.utils import get_color_from_hex


class StockScreen(Screen):
    _filtro_estado = ""

    def on_enter(self):
        self.actualizar_spinner_stock()
        self.consultar()

    def actualizar_spinner_stock(self) -> None:
        try:
            mats = sorted({str(r[0]).upper() for r in db_query("SELECT DISTINCT material FROM movimientos") if r[0]})
            self.ids.stock_spinner.values = mats
        except Exception as e:
            logger.error("Error actualizando spinner stock: %s", e)

    def set_filtro_estado(self, estado: str) -> None:
        self._filtro_estado = estado
        self.consultar()

    def consultar(self, *args) -> None:
        try:
            self.ids.s_r.clear_widgets()
            f_texto = self.ids.f_s.text.upper()
            f_spin  = (self.ids.stock_spinner.text.upper()
                       if self.ids.stock_spinner.text != "BUSCAR MATERIAL" else "")
            alertas_conf = {r[0].upper(): r[1] for r in db_query("SELECT material,stock_minimo FROM alertas_stock")}
            mats = sorted({r[0] for r in db_query("SELECT DISTINCT UPPER(material) FROM movimientos") if r[0]})
            for m in mats:
                if f_spin and m != f_spin:
                    continue
                if not f_spin and f_texto and f_texto not in m:
                    continue
                fijo, disponible = obtener_stock_material(m)
                minimo = alertas_conf.get(m, 5)
                if disponible <= 0:
                    estado = "sin_stock"
                elif disponible <= minimo:
                    estado = "bajo"
                else:
                    estado = "ok"
                if self._filtro_estado and estado != self._filtro_estado:
                    continue
                color_text = {'sin_stock':'#FF0044','bajo':'#FF0044','ok':'#00D4FF'}[estado]
                etiqueta   = {'sin_stock':'SIN STOCK','bajo':'STOCK BAJO','ok':'DISPONIBLE'}[estado]
                color_card = {'sin_stock':'#0D1B2A','bajo':'#0D1B2A','ok':'#0D1B2A'}[estado]
                card = crear_tarjeta(dp(110), color_card)
                card.add_widget(Label(text=f"[b]{m}[/b]", markup=True,
                                       color=get_color_from_hex('#00D4FF'),
                                       size_hint_y=None, height=dp(28),
                                       halign='left', valign='middle',
                                       text_size=(Window.width * 0.8, None)))
                card.add_widget(Label(text=f"Total sistema: {fijo}  |  Minimo alerta: {minimo}",
                                       color=get_color_from_hex('#5A6A7A'),
                                       size_hint_y=None, height=dp(24), font_size='11sp',
                                       halign='left', valign='middle',
                                       text_size=(Window.width * 0.8, None)))
                card.add_widget(Label(
                    text=f"[b][color={color_text}] {etiqueta}: {disponible}[/color][/b]",
                    markup=True,
                    size_hint_y=None, height=dp(28), font_size='14sp',
                    halign='left', valign='middle',
                    text_size=(Window.width * 0.8, None)))
                self.ids.s_r.add_widget(card)
        except Exception as e:
            logger.error("Error consultando stock: %s", e)


# === screens/estadisticas.py ===
from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.metrics import dp
from kivy.utils import get_color_from_hex
from kivy.core.window import Window


class EstadisticasScreen(Screen):
    def on_enter(self):
        self.cargar_estadisticas()

    def _crear_seccion(self, titulo, color_titulo='#00D4FF'):
        return Label(
            text=f"\n{titulo}", font_size='13sp', bold=True,
            color=get_color_from_hex(color_titulo),
            size_hint_y=None, height=dp(35),
            halign='left', valign='bottom',
            text_size=(Window.width * 0.85, None),
        )

    def _crear_barra(self, label, valor, maximo, color='#00D4FF'):
        return BarraHorizontal(label, valor, maximo, color)

    def cargar_estadisticas(self) -> None:
        try:
            self.ids.stats_container.clear_widgets()
            stats   = obtener_reporte_estadisticas()
            alertas = obtener_alertas_activas()

            self.ids.stats_container.add_widget(self._crear_seccion("RESUMEN GENERAL", '#FF0044'))

            kpis = [
                ("Total Movimientos",    stats["total_movimientos"], '#00D4FF'),
                ("Entradas",             stats["entradas"],          '#00D4FF'),
                ("Salidas",              stats["salidas"],           '#FF0044'),
                ("Materiales",           stats["materiales"],        '#00D4FF'),
                ("Responsables",         stats["responsables"],      '#FF0044'),
                ("Movs. 7 dias",          stats["movs_7dias"],        '#00D4FF'),
                ("Movs. 30 dias",         stats["movs_30dias"],       '#00D4FF'),
                ("Prom. cantidad",       stats["prom_cant"],         '#FF0044'),
            ]
            gl = GridLayout(cols=2, size_hint_y=None, spacing=dp(8))
            gl.bind(minimum_height=gl.setter('height'))
            for label, valor, color in kpis:
                card = crear_tarjeta(dp(80))
                card.add_widget(Label(text=label, font_size='11sp',
                                      color=get_color_from_hex('#5A6A7A')))
                card.add_widget(Label(
                    text=f"[b][color={color}]{valor}[/color][/b]",
                    markup=True, font_size='22sp'))
                gl.add_widget(card)
            self.ids.stats_container.add_widget(gl)

            self.ids.stats_container.add_widget(self._crear_seccion("VOLUMENES MOVIDOS"))
            card_vol = crear_tarjeta(dp(130))
            card_vol.add_widget(Label(
                text=f"Cantidad total entrada:  [b][color=#00D4FF]{stats['vol_entradas']}[/color][/b]",
                markup=True, font_size='13sp', color=get_color_from_hex('#E8EAED'),
                size_hint_y=None, height=dp(28), halign='left', text_size=(Window.width*0.8, None)
            ))
            card_vol.add_widget(Label(
                text=f"Cantidad total salida:   [b][color=#FF0044]{stats['vol_salidas']}[/color][/b]",
                markup=True, font_size='13sp', color=get_color_from_hex('#E8EAED'),
                size_hint_y=None, height=dp(28), halign='left', text_size=(Window.width*0.8, None)
            ))
            balance = stats['vol_entradas'] - stats['vol_salidas']
            bal_color = '#00D4FF' if balance >= 0 else '#FF0044'
            card_vol.add_widget(Label(
                text=f"Balance neto:  [b][color={bal_color}]{round(balance, 2)}[/color][/b]",
                markup=True, font_size='14sp', color=get_color_from_hex('#E8EAED'),
                size_hint_y=None, height=dp(30), halign='left', text_size=(Window.width*0.8, None)
            ))
            ratio = round(stats['entradas'] / max(stats['salidas'], 1), 2)
            card_vol.add_widget(Label(
                text=f"Ratio E/S:  [b][color=#00D4FF]{ratio}[/color][/b]",
                markup=True, font_size='12sp', color=get_color_from_hex('#5A6A7A'),
                size_hint_y=None, height=dp(24), halign='left', text_size=(Window.width*0.8, None)
            ))
            self.ids.stats_container.add_widget(card_vol)

            valor_inv = db_query(
                "SELECT COALESCE(SUM(cantidad * costo_unitario), 0), COALESCE(SUM(cantidad * precio_venta), 0) FROM movimientos WHERE tipo='ENTRADA'",
                fetchall=False
            )
            if valor_inv:
                costo_total, venta_total = valor_inv
                self.ids.stats_container.add_widget(self._crear_seccion("VALOR DEL INVENTARIO", '#00D4FF'))
                card_val = crear_tarjeta(dp(70), '#0D1B2A')
                card_val.add_widget(Label(
                    text=f"Costo total: [b][color=#00D4FF]$ {round(costo_total or 0, 2)}[/color][/b]",
                    markup=True, font_size='13sp', color=get_color_from_hex('#E8EAED'),
                    size_hint_y=None, height=dp(26), halign='left',
                    text_size=(Window.width*0.8, None)))
                margen = (venta_total or 0) - (costo_total or 0)
                color_margen = '#00D4FF' if margen >= 0 else '#FF0044'
                card_val.add_widget(Label(
                    text=f"Valor venta estimado: [b][color={color_margen}]$ {round(venta_total or 0, 2)}[/color][/b]",
                    markup=True, font_size='13sp', color=get_color_from_hex('#E8EAED'),
                    size_hint_y=None, height=dp(26), halign='left',
                    text_size=(Window.width*0.8, None)))
                self.ids.stats_container.add_widget(card_val)

            self.ids.stats_container.add_widget(self._crear_seccion("TENDENCIA 7 DIAS"))
            card_trend = crear_tarjeta(dp(90))
            max_7d = max(stats['ent_7d'], stats['sal_7d'], 1)
            card_trend.add_widget(self._crear_barra("Entradas", stats['ent_7d'], max_7d, '#00D4FF'))
            card_trend.add_widget(self._crear_barra("Salidas", stats['sal_7d'], max_7d, '#FF0044'))
            self.ids.stats_container.add_widget(card_trend)

            chart = GraficoTendencia()
            datos = db_query(
                "SELECT COUNT(*), strftime('%w', fecha) FROM movimientos WHERE fecha >= date('now','-7 days') GROUP BY strftime('%w', fecha) ORDER BY 2",
                fetchall=False
            )
            if datos:
                vals = [0]*7
                for cnt, dia in datos:
                    vals[int(dia)] = cnt
                chart.mostrar(vals, '#00D4FF')
                self.ids.stats_container.add_widget(chart)

            if stats['top5_materiales']:
                self.ids.stats_container.add_widget(self._crear_seccion("TOP 5 MATERIALES", '#00D4FF'))
                max_cnt = stats['top5_materiales'][0][1] if stats['top5_materiales'] else 1
                card_top = crear_tarjeta(dp(40 * len(stats['top5_materiales']) + 10))
                for i, (m, cnt) in enumerate(stats['top5_materiales'], 1):
                    colors = ['#FF0044', '#5A6A7A', '#5A6A7A', '#5A6A7A', '#5A6A7A']
                    prefix = '*' if i == 1 else '-'
                    card_top.add_widget(self._crear_barra(f"{prefix} {m}", cnt, max_cnt, colors[i-1]))
                self.ids.stats_container.add_widget(card_top)

            if stats['top5_responsables']:
                self.ids.stats_container.add_widget(self._crear_seccion("TOP 5 RESPONSABLES", '#FF0044'))
                max_r = stats['top5_responsables'][0][1] if stats['top5_responsables'] else 1
                card_resp = crear_tarjeta(dp(40 * len(stats['top5_responsables']) + 10))
                for i, (n, cnt) in enumerate(stats['top5_responsables'], 1):
                    card_resp.add_widget(self._crear_barra(f"{i}. {n}", cnt, max_r, '#FF0044'))
                self.ids.stats_container.add_widget(card_resp)

            if stats["material_top"] != "N/A":
                card = crear_tarjeta(dp(70), '#0D1B2A')
                card.add_widget(Label(text="Material mas activo",
                                      color=get_color_from_hex('#00D4FF'), font_size='12sp'))
                card.add_widget(Label(text=f"[b]{stats['material_top']}[/b]",
                                      markup=True, font_size='18sp',
                                      color=get_color_from_hex('#E8EAED')))
                self.ids.stats_container.add_widget(card)

            if alertas:
                self.ids.stats_container.add_widget(self._crear_seccion(
                    f"ALERTAS DE STOCK ({len(alertas)})", '#FF0044'))
                card = crear_tarjeta(dp(30 + len(alertas)*36), '#0D1B2A')
                for mat, disp, minimo in alertas:
                    card.add_widget(Label(
                        text=f"  {mat}:  {disp}  (min: {minimo})",
                        font_size='13sp', color=get_color_from_hex('#FF0044'),
                        size_hint_y=None, height=dp(32),
                        halign='left', valign='middle',
                        text_size=(Window.width*0.8, None),
                        padding=(dp(4), dp(2))
                    ))
                self.ids.stats_container.add_widget(card)

            if stats['total_limpieza'] > 0:
                self.ids.stats_container.add_widget(self._crear_seccion("LIMPIEZA", '#00D4FF'))
                card_limp = crear_tarjeta(dp(90))
                max_l = max(stats['limpios'], stats['sucios'], 1)
                card_limp.add_widget(self._crear_barra("Limpios", stats['limpios'], max_l, '#00D4FF'))
                card_limp.add_widget(self._crear_barra("Sucios", stats['sucios'], max_l, '#FF0044'))
                self.ids.stats_container.add_widget(card_limp)

            if stats['mats_inactivos']:
                self.ids.stats_container.add_widget(self._crear_seccion(
                    f"MATERIALES INACTIVOS (+30d):  {len(stats['mats_inactivos'])}", '#5A6A7A'))
                card_inact = crear_tarjeta(dp(20 + len(stats['mats_inactivos'])*24))
                for m in stats['mats_inactivos']:
                    card_inact.add_widget(Label(
                        text=f"  {m}", font_size='11sp',
                        color=get_color_from_hex('#5A6A7A'),
                        size_hint_y=None, height=dp(22),
                        halign='left', text_size=(Window.width*0.8, None)
                    ))
                self.ids.stats_container.add_widget(card_inact)

            self.ids.stats_container.add_widget(self._crear_seccion("SISTEMA", '#5A6A7A'))
            card_sys = crear_tarjeta(dp(100))
            card_sys.add_widget(Label(
                text=f"Primer movimiento:  {stats['primer_mov']}",
                font_size='11sp', color=get_color_from_hex('#5A6A7A'),
                size_hint_y=None, height=dp(22), halign='left', text_size=(Window.width*0.8, None)
            ))
            card_sys.add_widget(Label(
                text=f"Ultimo movimiento:  {stats['ultimo_mov']}",
                font_size='11sp', color=get_color_from_hex('#5A6A7A'),
                size_hint_y=None, height=dp(22), halign='left', text_size=(Window.width*0.8, None)
            ))
            import os
            try:
                size_kb = os.path.getsize(DB_PATH) / 1024
                card_sys.add_widget(Label(
                    text=f"Base de datos:  {size_kb:.1f} KB",
                    font_size='11sp', color=get_color_from_hex('#5A6A7A'),
                    size_hint_y=None, height=dp(22), halign='left', text_size=(Window.width*0.8, None)
                ))
            except Exception:
                pass
            card_sys.add_widget(Label(
                text=f"Version:  {VERSION_ACTUAL}",
                font_size='11sp', color=get_color_from_hex('#5A6A7A'),
                size_hint_y=None, height=dp(22), halign='left', text_size=(Window.width*0.8, None)
            ))
            self.ids.stats_container.add_widget(card_sys)
        except Exception as e:
            logger.error("Error cargando estadisticas: %s", e)


# === screens/limpieza.py ===
from datetime import datetime
from collections import defaultdict
from kivy.uix.screenmanager import Screen
from kivy.uix.label import Label
from kivy.uix.boxlayout import BoxLayout
from kivy.metrics import dp
from kivy.utils import get_color_from_hex


class LimpiezaScreen(Screen):
    def on_enter(self):
        self.actualizar_spinner_limpieza()
        self.consultar_historial()

    def actualizar_spinner_limpieza(self) -> None:
        try:
            mats = sorted({str(r[0]).upper() for r in db_query("SELECT DISTINCT material FROM limpieza") if r[0]})
            self.ids.limpieza_spinner.values = mats
        except Exception as e:
            logger.error("Error actualizando spinner limpieza: %s", e)

    def registrar_limpieza(self, estado: str) -> None:
        mat   = (self.ids.m_n.text.strip().upper()
                 or (self.ids.limpieza_spinner.text
                     if self.ids.limpieza_spinner.text != "SELECCIONAR MATERIAL" else ""))
        cant  = self.ids.m_c.text.strip()
        notas = self.ids.m_notas.text.strip()
        if not mat or not cant:
            self.ids.info_limpieza.text = "Material y cantidad requeridos"
            return
        existe = db_query("SELECT 1 FROM movimientos WHERE UPPER(material)=? LIMIT 1", (mat,), fetchall=False)
        if not existe:
            self.ids.info_limpieza.text = f"'{mat}' no existe en movimientos"
            return
        try:
            db_execute(
                "INSERT INTO limpieza (material,cantidad,estado,fecha,notas) VALUES (?,?,?,?,?)",
                (mat, float(cant), estado, datetime.now().strftime("%d/%m/%Y %H:%M"), notas)
            )
            self.ids.m_n.text = ""
            self.ids.m_c.text = ""
            self.ids.m_notas.text = ""
            self.ids.info_limpieza.text = ""
            self.ids.limpieza_spinner.text = "SELECCIONAR MATERIAL"
            self.consultar_historial()
            self.actualizar_spinner_limpieza()
        except Exception as e:
            logger.error("Error registrando limpieza: %s", e)
            self.ids.info_limpieza.text = f"Error: {e}"

    def consultar_historial(self) -> None:
        try:
            self.ids.hist_limpieza.clear_widgets()
            busq  = self.ids.f_l.text.lower()
            rows  = db_query("SELECT material, cantidad, estado FROM limpieza")
            totales = defaultdict(lambda: {"LIMPIO": 0.0, "SUCIO": 0.0})
            for m, c, e in rows:
                totales[str(m).upper()][e] += float(c)
            for mat, v in sorted(totales.items()):
                if busq and busq not in mat.lower():
                    continue
                neto = v["SUCIO"] - v["LIMPIO"]
                if neto <= 0:
                    continue
                card = crear_tarjeta(dp(90))
                card.add_widget(Label(text=f"[b]{mat}[/b]", markup=True,
                                      color=get_color_from_hex('#00D4FF'),
                                      size_hint_y=None, height=dp(26), font_size='12sp',
                                      halign='left', valign='middle',
                                      text_size=(Window.width * 0.82, None)))
                card.add_widget(Label(text=f"Sucio: {v['SUCIO']}  |  Limpio: {v['LIMPIO']}",
                                      color=get_color_from_hex('#5A6A7A'),
                                      size_hint_y=None, height=dp(22), font_size='11sp',
                                      halign='left', valign='middle',
                                      text_size=(Window.width * 0.82, None)))
                card.add_widget(Label(
                    text=f"[color=#FF0044]Pendiente de limpiar: {max(0, neto):.2f}[/color]",
                    markup=True,
                    size_hint_y=None, height=dp(22), font_size='11sp',
                    halign='left', valign='middle',
                    text_size=(Window.width * 0.82, None)))
                self.ids.hist_limpieza.add_widget(card)
        except Exception as e:
            logger.error("Error consultando historial limpieza: %s", e)



# === App Class ===
class InventarioApp(App):
    def build(self):
        self.title = "Gestion Inventario PRO v" + VERSION_ACTUAL
        sm = ScreenManager(transition=SlideTransition())
        sm.add_widget(InventarioScreen(name='inventario'))
        sm.add_widget(LimpiezaScreen(name='limpieza'))
        sm.add_widget(PendientesScreen(name='pendientes'))
        sm.add_widget(FiltroScreen(name='filtros'))
        sm.add_widget(StockScreen(name='stock'))
        sm.add_widget(EstadisticasScreen(name='estadisticas'))
        sm.add_widget(AjustesScreen(name='ajustes'))
        sm.add_widget(ImportarExcelScreen(name='importar'))
        sm.add_widget(AuditoriaScreen(name='auditoria'))
        sm.add_widget(AlertasScreen(name='alertas'))
        sm.add_widget(CategoriasScreen(name='categorias'))
        sm.add_widget(UbicacionesScreen(name='ubicaciones'))
        sm.add_widget(GestionScreen(name='gestion'))
        sm.add_widget(ValorScreen(name='valor'))
        self._iniciar_auto_actualizacion()
        self._iniciar_scheduler()
        return sm

    def _iniciar_auto_actualizacion(self):
        Clock.schedule_once(lambda dt: self._auto_check(None), 2)
        Clock.schedule_interval(self._auto_check, 30)

    def _auto_check(self, dt):
        threading.Thread(target=self._auto_worker, daemon=True).start()
        threading.Thread(target=self._enviar_sync, daemon=True).start()

    def _auto_worker(self):
        try:
            actualizador = AutoActualizador(callback_fin=self._auto_fin)
            actualizador.ejecutar()
        except Exception:
            pass

    def _auto_fin(self, exito, datos):
        if exito and datos.get("importados", 0) > 0:
            n = datos["importados"]
            Clock.schedule_once(lambda dt, cant=n: self._notificar_actualizacion(cant))

    def _notificar_actualizacion(self, cant):
        try:
            inv = self.root.get_screen('inventario')
            inv.ids.info.text = f"[color=#00D4FF]Auto-actualizado: +{cant} registros[/color]"
            inv.actualizar_spinners()
            inv._actualizar_dashboard()
            inv._verificar_alertas_silencioso()
            self._enviar_sync()
        except Exception:
            pass

    def _enviar_sync(self):
        try:
            enviar_sync_email()
        except Exception:
            pass

    def on_pause(self):
        return True

    def _iniciar_scheduler(self):
        self._scheduler = ProgramadorTareas(self)
        self._scheduler.iniciar()

    def on_start(self):
        Window.bind(on_key_down=self._on_key_down)

    def _on_key_down(self, window, key, scancode, codepoint, modifiers):
        if self.root.current != 'inventario':
            return
        ctrl = 'ctrl' in modifiers
        if ctrl and codepoint == 'e':
            self.root.get_screen('inventario').procesar("ENTRADA")
        elif ctrl and codepoint == 's':
            self.root.get_screen('inventario').procesar("SALIDA")
        elif ctrl and codepoint == 'b':
            self.root.get_screen('inventario').hacer_backup()

    def on_stop(self):
        gc.collect()


if __name__ == "__main__":
    try:
        InventarioApp().run()
    except Exception as e:
        logger.critical("Error fatal en main: %s", e, exc_info=True)
        try:
            ruta_err = os.path.join(os.path.dirname(os.path.abspath(__file__)), "crash_main.txt")
            with open(ruta_err, "w", encoding="utf-8") as f:
                import traceback
                f.write("CRASH en __main__ " + str(__import__('datetime').datetime.now()) + "\n")
                traceback.print_exc(file=f)
        except Exception:
            pass
        raise
