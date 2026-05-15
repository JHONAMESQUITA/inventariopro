import json
import smtplib
import imaplib
import email
import os
import threading
from email.header import decode_header
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime, timedelta
from config import EMAIL_CONFIG, EXPORT_DIR
from database.connection import db_query, db_execute, DatabaseConnection
from config import DB_PATH


from services.auditoria import log_auditoria
from services.excel_service import ruta_exportacion, generar_excel
from services.inventario import obtener_reporte_estadisticas
from services.alertas import obtener_vencimientos
from services.sync import exportar_cambios, importar_cambios, obtener_dispositivo_id
from logging_config import logger


def enviar_alerta_vencimientos() -> tuple:
    venc = obtener_vencimientos()
    vencidos   = venc["vencidos"]
    por_vencer = venc["por_vencer"]

    if not vencidos and not por_vencer:
        return True, "Sin vencimientos - no se envio correo."

    cfg = EMAIL_CONFIG
    if not cfg.get("password", "").strip():
        return False, "Sin contrasena configurada en APP_EMAIL_PASS."

    now = datetime.now()
    lineas = [
        f"ALERTA DE VENCIMIENTOS - {now.strftime('%d/%m/%Y %H:%M')}",
        "",
    ]
    if vencidos:
        lineas.append(f"EQUIPOS CON RETORNO VENCIDO ({len(vencidos)}):")
        for p in vencidos:
            lineas.append(
                f"  * {p['material']} - Responsable: {p['responsable']}"
                f" | Dias fuera: {p['dias_fuera']} (acordado: {p['dias_acordados']})"
                f" | Salida: {p['fecha_salida']}"
            )
        lineas.append("")
    if por_vencer:
        lineas.append(f"PROXIMOS A VENCER ({len(por_vencer)}):")
        for p in por_vencer:
            lineas.append(
                f"  * {p['material']} - Responsable: {p['responsable']}"
                f" | Dias fuera: {p['dias_fuera']} (acordado: {p['dias_acordados']})"
            )
        lineas.append("")
    from config import VERSION_ACTUAL
    lineas.append(f"- INVENTARIO PRO v{VERSION_ACTUAL}")
    cuerpo = "\n".join(lineas)

    try:
        msg = MIMEMultipart()
        msg['From']    = cfg["remitente"]
        msg['To']      = ", ".join(cfg["destinatarios"])
        msg['Subject'] = (
            f"ALERTA ALQUILER: {len(vencidos)} vencido(s), "
            f"{len(por_vencer)} por vencer - {now.strftime('%d/%m/%Y')}"
        )
        msg.attach(MIMEText(cuerpo, 'plain', 'utf-8'))
        s = smtplib.SMTP('smtp.gmail.com', 587)
        s.starttls()
        s.login(cfg["remitente"], cfg["password"])
        for dest in cfg["destinatarios"]:
            s.sendmail(cfg["remitente"], dest, msg.as_string())
        s.quit()
        log_auditoria("ALERTA_VENCIMIENTOS_EMAIL",
                      f"{len(vencidos)} vencidos, {len(por_vencer)} por vencer")
        return True, "Alerta enviada correctamente."
    except Exception as e:
        return False, f"Error al enviar: {e}"


def enviar_reporte_por_correo() -> tuple:
    EXCEL_TEMP = ruta_exportacion(f"Reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    try:
        generar_excel(EXCEL_TEMP)
    except Exception as e:
        logger.error("Error generando Excel para correo: %s", e, exc_info=True)
        return False, f"Error al generar Excel: {e}"
    try:
        msg = MIMEMultipart()
        cfg = EMAIL_CONFIG
        if not cfg.get("remitente") or not cfg.get("password"):
            return False, "Faltan credenciales de correo en .env (APP_EMAIL, APP_EMAIL_PASS)"
        msg['From']    = cfg["remitente"]
        msg['To']      = ", ".join(cfg["destinatarios"])
        msg['Subject'] = f"INVENTARIO PRO - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        stats = obtener_reporte_estadisticas()
        body  = (f"Reporte automatico generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}.\n\n"
                 f"Resumen:\n"
                 f"  * Total movimientos: {stats['total_movimientos']}\n"
                 f"  * Entradas: {stats['entradas']} | Salidas: {stats['salidas']}\n"
                 f"  * Materiales distintos: {stats['materiales']}\n"
                 f"  * Material mas activo: {stats['material_top']}\n\n"
                 f"Ver detalle en el archivo adjunto.\n\n")
        from config import VERSION_ACTUAL
        body += f"- INVENTARIO PRO v{VERSION_ACTUAL}"
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        with open(EXCEL_TEMP, "rb") as f:
            parte = MIMEBase('application', 'octet-stream')
            parte.set_payload(f.read())
            encoders.encode_base64(parte)
            parte.add_header('Content-Disposition', f'attachment; filename={os.path.basename(EXCEL_TEMP)}')
            msg.attach(parte)
        s = smtplib.SMTP('smtp.gmail.com', 587)
        s.starttls()
        s.login(cfg["remitente"], cfg["password"])
        for dest in cfg["destinatarios"]:
            s.sendmail(cfg["remitente"], dest, msg.as_string())
        s.quit()
        log_auditoria("EMAIL_ENVIADO", f"Destinatarios: {cfg['destinatarios']}")
        return True, "Reporte enviado correctamente."
    except smtplib.SMTPAuthenticationError:
        return False, ("Error de autenticacion SMTP. Si usas Gmail, necesitas una "
                       "Contrasena de Aplicacion (no la contrasena normal).")
    except smtplib.SMTPException as e:
        logger.error("Error SMTP: %s", e, exc_info=True)
        return False, f"Error SMTP: {e}"
    except Exception as e:
        logger.error("Error al enviar correo: %s", e, exc_info=True)
        return False, f"Error al enviar: {e}"


def enviar_sync_email() -> tuple:
    try:
        cfg = EMAIL_CONFIG
        if not cfg.get("remitente") or not cfg.get("password"):
            return False, "Sin credenciales"
        cambios = exportar_cambios()
        if not cambios.get("movimientos"):
            return True, "Sin cambios para sincronizar"
        payload = json.dumps(cambios, ensure_ascii=False, default=str)
        msg = MIMEMultipart()
        msg['From'] = cfg["remitente"]
        msg['To'] = cfg["remitente"]
        device = obtener_dispositivo_id()
        msg['Subject'] = f"[SYNC] {device} - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        parte = MIMEText(payload, 'plain', 'utf-8')
        parte.add_header('Content-Disposition', 'attachment; filename="sync.json"')
        msg.attach(parte)
        s = smtplib.SMTP('smtp.gmail.com', 587)
        s.starttls()
        s.login(cfg["remitente"], cfg["password"])
        s.sendmail(cfg["remitente"], cfg["remitente"], msg.as_string())
        s.quit()
        log_auditoria("SYNC_ENVIADO", f"{len(cambios['movimientos'])} registros")
        return True, f"Sync enviado: {len(cambios['movimientos'])} registros"
    except Exception as e:
        logger.error("Error enviando sync: %s", e)
        return False, str(e)


def _procesar_sync_email(msg) -> dict:
    for parte in msg.walk():
        filename = parte.get_filename()
        if filename and "sync.json" in filename.lower():
            payload = parte.get_payload(decode=True)
            if payload:
                try:
                    datos = json.loads(payload.decode('utf-8'))
                    return datos
                except Exception:
                    pass
    return {}


class ActualizadorCorreo:
    EXTENSIONES = ('.xlsx', '.xls', '.xlsm')

    def __init__(self, callback_progreso, callback_fin):
        self.cb_prog = callback_progreso
        self.cb_fin  = callback_fin

    def ejecutar(self) -> None:
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self) -> None:
        self._prog("Conectando al servidor de correo...")
        try:
            mail = imaplib.IMAP4_SSL(EMAIL_CONFIG["imap_servidor"], EMAIL_CONFIG["imap_puerto"])
        except Exception as e:
            self._fin(False, {"error": f"No se pudo conectar al servidor IMAP: {e}"})
            return

        self._prog("Iniciando sesion...")
        if not str(EMAIL_CONFIG.get("password") or "").strip():
            try: mail.logout()
            except Exception: pass
            self._fin(False, {"error": "Falta la contrasena de correo."})
            return

        try:
            mail.login(EMAIL_CONFIG["remitente"], EMAIL_CONFIG["password"])
        except imaplib.IMAP4.error as e:
            try: mail.logout()
            except Exception: pass
            self._fin(False, {"error": f"Error de autenticacion: {e}"})
            return

        self._prog("Buscando correos con archivos Excel...")
        try:
            adjunto, meta = self._buscar_adjunto_excel(mail)
        except Exception as e:
            try: mail.logout()
            except Exception: pass
            self._fin(False, {"error": f"Error al buscar correos: {e}"})
            return

        if adjunto is None:
            try: mail.logout()
            except Exception: pass
            self._fin(False, {"error": "No se encontro ningun correo con un adjunto Excel."})
            return

        nombre_archivo = meta["nombre_archivo"]
        self._prog(f"Descargando: {nombre_archivo}...")
        ruta_local = ruta_exportacion(f"correo_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{nombre_archivo}")
        try:
            with open(ruta_local, "wb") as f:
                f.write(adjunto)
        except Exception as e:
            try: mail.logout()
            except Exception: pass
            self._fin(False, {"error": f"No se pudo guardar el archivo: {e}"})
            return

        self._prog("Importando datos al inventario...")
        try:
            resultado = self._importar_excel(ruta_local, meta)
        except Exception as e:
            try: mail.logout()
            except Exception: pass
            self._fin(False, {"error": f"Error al importar el Excel: {e}"})
            return

        try:
            mail.select("INBOX")
            mail.store(meta["uid"], '+FLAGS', '\\Seen')
        except Exception: pass
        try: mail.logout()
        except Exception: pass

        try:
            db_execute(
                "INSERT INTO actualizaciones_correo "
                "(fecha, remitente, asunto, archivo, insertados, omitidos, errores) "
                "VALUES (?,?,?,?,?,?,?)",
                (datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                 meta.get("de", ""), meta.get("asunto", ""),
                 nombre_archivo, resultado["insertados"],
                 resultado["omitidos"], resultado["errores"])
            )
            log_auditoria("ACTUALIZAR_CORREO",
                          f"{nombre_archivo} | +{resultado['insertados']} registros")
        except Exception: pass

        self._fin(True, resultado)

    def _buscar_adjunto_excel(self, mail):
        mail.select("INBOX")

        def _armar_criterio(*grupos):
            partes = []
            for g in grupos:
                partes.extend(g)
            return "(" + " ".join(partes) + ")" if partes else "ALL"

        filtros = []
        if EMAIL_CONFIG.get("asunto_filtro"):
            filtros.append(f'SUBJECT "{EMAIL_CONFIG["asunto_filtro"]}"')
        if EMAIL_CONFIG.get("remitente_filtro"):
            filtros.append(f'FROM "{EMAIL_CONFIG["remitente_filtro"]}"')

        fecha_desde = (datetime.now() - timedelta(days=90)).strftime("%d-%b-%Y")

        # 1) buscar con filtros + ventana de tiempo
        criterio_str = _armar_criterio(filtros, [f'SINCE "{fecha_desde}"'])
        _, data = mail.search(None, criterio_str)
        uids = data[0].split()

        # 2) si no hay, reintentar sin ventana de tiempo
        if not uids and filtros:
            criterio_str = _armar_criterio(filtros)
            _, data = mail.search(None, criterio_str)
            uids = data[0].split()

        # 3) si aun no hay y no hay filtros, buscar todo
        if not uids and not filtros:
            criterio_str = "ALL"
            _, data = mail.search(None, criterio_str)
            uids = data[0].split()

        if not uids:
            return None, {}

        for uid in reversed(uids):
            _, msg_data = mail.fetch(uid, "(RFC822)")
            if not msg_data or not msg_data[0]:
                continue
            raw = msg_data[0][1]
            msg = email.message_from_bytes(raw)
            de      = self._decodificar_header(msg.get("From", ""))
            asunto  = self._decodificar_header(msg.get("Subject", ""))
            for parte in msg.walk():
                filename = parte.get_filename()
                if not filename:
                    continue
                nombre_dec = self._decodificar_header(filename)
                if not nombre_dec.lower().endswith(self.EXTENSIONES):
                    continue
                payload = parte.get_payload(decode=True)
                if not payload:
                    continue
                meta = {
                    "uid": uid, "de": de, "asunto": asunto,
                    "nombre_archivo": nombre_dec,
                }
                return payload, meta
        return None, {}

    @staticmethod
    def _importar_excel(ruta_local, meta):
        from openpyxl import load_workbook
        from services.inventario import obtener_stock_material

        wb   = load_workbook(ruta_local, read_only=True, data_only=True)
        hoja = wb.sheetnames[0]
        ws   = wb[hoja]
        filas = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(filas) < 2:
            return {"insertados": 0, "omitidos": 0, "errores": 0,
                    "nota": "El archivo no tiene filas de datos"}

        enc_upper = [str(h).strip().upper() if h is not None else "" for h in filas[0]]
        mapeo_campos = {
            'nombre':   ['NOMBRE','RESPONSABLE','NAME','USUARIO','RESP'],
            'material': ['MATERIAL','PRODUCTO','ITEM','DESCRIPCION','DESCRIPTION','MAT'],
            'sku':      ['SKU','ID','CODIGO','CODE','REF'],
            'cantidad': ['CANTIDAD','QTY','QUANTITY','CANT'],
            'tipo':     ['TIPO','TYPE','MOVIMIENTO','MOVEMENT'],
            'fecha':    ['FECHA','DATE','DATETIME','TIMESTAMP'],
        }
        indices = {}
        for campo, alternativas in mapeo_campos.items():
            for alt in alternativas:
                if alt in enc_upper:
                    indices[campo] = enc_upper.index(alt)
                    break
        if 'material' not in indices or 'cantidad' not in indices:
            raise ValueError("No se encontraron las columnas 'Material' y 'Cantidad'.")

        conn = DatabaseConnection.get_instance(DB_PATH).get_connection()
        cursor = conn.cursor()
        insertados = errores = omitidos = 0
        ahora = datetime.now().strftime("%d/%m/%Y %H:%M")

        for fila in filas[1:]:
            try:
                i = indices
                nombre   = str(fila[i.get('nombre')] if i.get('nombre') is not None and i['nombre'] < len(fila) else "INVENTARIO").strip().upper()
                material = str(fila[i['material']] if i['material'] < len(fila) else "").strip().upper()
                sku      = str(fila[i.get('sku')] if i.get('sku') is not None and i['sku'] < len(fila) else "").strip()
                raw_cant = fila[i['cantidad']] if i['cantidad'] < len(fila) else None
                raw_tipo = str(fila[i.get('tipo')] if i.get('tipo') is not None and i['tipo'] < len(fila) else "ENTRADA").strip().upper()
                raw_fec  = fila[i.get('fecha')] if i.get('fecha') is not None and i['fecha'] < len(fila) else None

                if not material:
                    omitidos += 1; continue
                try:
                    cantidad = float(str(raw_cant).replace(',', '.'))
                except (TypeError, ValueError):
                    omitidos += 1; continue
                if cantidad <= 0:
                    omitidos += 1; continue

                if any(k in raw_tipo for k in ['ENT','IN','ING','COMP']):
                    tipo = "ENTRADA"
                elif any(k in raw_tipo for k in ['SAL','OUT','EGR','RET']):
                    tipo = "SALIDA"
                else:
                    tipo = "ENTRADA"

                if raw_fec is None:
                    fecha = ahora
                elif isinstance(raw_fec, datetime):
                    fecha = raw_fec.strftime("%d/%m/%Y %H:%M")
                else:
                    fecha = str(raw_fec).strip() or ahora

                cursor.execute(
                    "SELECT 1 FROM movimientos WHERE nombre=? AND material=? AND sku=? AND cantidad=? AND tipo=? AND fecha=? LIMIT 1",
                    (nombre, material, sku, cantidad, tipo, fecha)
                )
                if cursor.fetchone():
                    omitidos += 1; continue

                _, stock_actual = obtener_stock_material(material)
                nuevo_stock = round(stock_actual + (cantidad if tipo == "ENTRADA" else -cantidad), 4)
                cursor.execute(
                    "INSERT INTO movimientos (nombre,material,sku,cantidad,tipo,fecha,stock_registro,dias,retorno,notas,ubicacion) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    (nombre, material, sku, cantidad, tipo, fecha, nuevo_stock, 0, "N/A",
                     "", "")
                )
                insertados += 1
            except Exception:
                errores += 1

        conn.commit()
        return {"insertados": insertados, "omitidos": omitidos, "errores": errores,
                "archivo": meta.get("nombre_archivo", ""), "de": meta.get("de", ""),
                "asunto": meta.get("asunto", "")}

    @staticmethod
    def _decodificar_header(valor):
        if not valor:
            return ""
        partes = decode_header(valor)
        resultado = []
        for parte, enc in partes:
            if isinstance(parte, bytes):
                resultado.append(parte.decode(enc or "utf-8", errors="replace"))
            else:
                resultado.append(str(parte))
        return "".join(resultado)

    def _prog(self, texto):
        from kivy.clock import Clock
        Clock.schedule_once(lambda dt: self.cb_prog(texto))

    def _fin(self, exito, datos):
        from kivy.clock import Clock
        Clock.schedule_once(lambda dt: self.cb_fin(exito, datos))


class AutoActualizador:
    EXTENSIONES = ('.xlsx', '.xls', '.xlsm')

    def __init__(self, callback_fin):
        self.cb_fin = callback_fin

    def ejecutar(self):
        import threading
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self):
        try:
            resultado = self._revisar_e_importar()
            sync_result = self._revisar_sync()
            total = resultado.get("importados", 0) + sync_result.get("importados", 0)
            resultado["importados"] = total
            resultado["sync"] = sync_result.get("importados", 0)
            from kivy.clock import Clock
            Clock.schedule_once(lambda dt: self.cb_fin(True, resultado))
        except Exception as ex:
            err = str(ex)
            from kivy.clock import Clock
            Clock.schedule_once(lambda dt, e=err: self.cb_fin(False, {"error": e}))

    def _revisar_sync(self):
        sync_count = 0
        try:
            mail = imaplib.IMAP4_SSL(EMAIL_CONFIG["imap_servidor"], EMAIL_CONFIG["imap_puerto"])
            mail.login(EMAIL_CONFIG["remitente"], EMAIL_CONFIG["password"])
            mail.select("INBOX")
            _, data = mail.search(None, 'UNSEEN', 'SUBJECT "[SYNC]"')
            uids = data[0].split() if data[0] else []
            for uid in reversed(uids):
                _, msg_data = mail.fetch(uid, "(RFC822)")
                if not msg_data or not msg_data[0]:
                    continue
                raw = msg_data[0][1]
                msg = email.message_from_bytes(raw)
                datos = _procesar_sync_email(msg)
                if datos:
                    res = importar_cambios(datos)
                    sync_count += res.get("importados", 0)
                mail.select("INBOX")
                mail.store(uid, '+FLAGS', '\\Seen')
            mail.logout()
        except Exception as e:
            logger.error("Error en sync check: %s", e)
        return {"importados": sync_count}

    def _revisar_e_importar(self):
        mail = imaplib.IMAP4_SSL(EMAIL_CONFIG["imap_servidor"], EMAIL_CONFIG["imap_puerto"])
        mail.login(EMAIL_CONFIG["remitente"], EMAIL_CONFIG["password"])
        mail.select("INBOX")

        criterios = ["UNSEEN"]
        if EMAIL_CONFIG.get("asunto_filtro"):
            criterios.append(f'SUBJECT "{EMAIL_CONFIG["asunto_filtro"]}"')
        if EMAIL_CONFIG.get("remitente_filtro"):
            criterios.append(f'FROM "{EMAIL_CONFIG["remitente_filtro"]}"')

        _, data = mail.search(None, *criterios)
        uids = data[0].split() if data[0] else []

        if not uids:
            mail.logout()
            return {"importados": 0, "mensaje": "Sin correos nuevos"}

        total_importados = 0
        for uid in reversed(uids):
            _, msg_data = mail.fetch(uid, "(RFC822)")
            if not msg_data or not msg_data[0]:
                continue
            raw = msg_data[0][1]
            msg = email.message_from_bytes(raw)
            de = ActualizadorCorreo._decodificar_header(None, msg.get("From", ""))
            asunto = ActualizadorCorreo._decodificar_header(None, msg.get("Subject", ""))

            for parte in msg.walk():
                filename = parte.get_filename()
                if not filename:
                    continue
                nombre_dec = ActualizadorCorreo._decodificar_header(filename)
                if not nombre_dec.lower().endswith(self.EXTENSIONES):
                    continue
                payload = parte.get_payload(decode=True)
                if not payload:
                    continue

                meta = {"uid": uid, "de": de, "asunto": asunto, "nombre_archivo": nombre_dec}
                ruta_local = ruta_exportacion(f"auto_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{nombre_dec}")
                with open(ruta_local, "wb") as f:
                    f.write(payload)

                resultado = ActualizadorCorreo._importar_excel(ruta_local, meta)
                total_importados += resultado.get("insertados", 0)

                mail.select("INBOX")
                mail.store(uid, '+FLAGS', '\\Seen')

                db_execute(
                    "INSERT INTO actualizaciones_correo "
                    "(fecha, remitente, asunto, archivo, insertados, omitidos, errores) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                     de, asunto, nombre_dec,
                     resultado.get("insertados", 0),
                     resultado.get("omitidos", 0),
                     resultado.get("errores", 0))
                )
                log_auditoria("AUTO_ACTUALIZAR_CORREO",
                              f"{nombre_dec} | +{resultado.get('insertados', 0)} registros")

        mail.logout()
        return {"importados": total_importados}
