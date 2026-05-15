import os
import csv
from datetime import datetime
from database.connection import db_query
from database.connection import DatabaseConnection
from config import DB_PATH, EXPORT_DIR
from services.auditoria import log_auditoria
from services.inventario import obtener_reporte_estadisticas
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def ruta_exportacion(nombre_archivo: str) -> str:
    try:
        os.makedirs(EXPORT_DIR, exist_ok=True)
    except Exception:
        pass
    return os.path.join(EXPORT_DIR, nombre_archivo)


def generar_excel(ruta: str = None) -> str:
    if ruta is None:
        ruta = ruta_exportacion("Reporte_Inventario.xlsx")

    wb = Workbook()
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1E3A5F")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    entrada_fill = PatternFill("solid", fgColor="D4EDDA")
    salida_fill  = PatternFill("solid", fgColor="F8D7DA")
    alerta_fill  = PatternFill("solid", fgColor="FFF3CD")

    ws1 = wb.active
    ws1.title = "Movimientos"
    ws1.freeze_panes = "A2"
    headers = ["#","Responsable","Material","SKU","Cantidad","Tipo","Fecha",
               "Stock Registrado","Dias","Retorno","Notas","Ubicacion"]
    ws1.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center_align; cell.border = thin_border

    movimientos = db_query(
        "SELECT id,nombre,material,sku,cantidad,tipo,fecha,stock_registro,dias,retorno,notas,ubicacion "
        "FROM movimientos ORDER BY id DESC"
    )
    for row_data in movimientos:
        ws1.append(tuple(row_data))
        row_idx = ws1.max_row
        fill = entrada_fill if row_data[5] == "ENTRADA" else salida_fill
        for col_idx in range(1, len(headers)+1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.fill = fill; cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    widths = [5,20,20,12,10,10,18,15,8,10,25,15]
    for col_idx, width in enumerate(widths, 1):
        ws1.column_dimensions[chr(64+col_idx)].width = width

    ws2 = wb.create_sheet("Stock Actual")
    ws2.freeze_panes = "A2"
    ws2.append(["Material","Stock Total Sistema","Stock Disponible Estante","Estado"])
    for col_idx in range(1, 5):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center_align; cell.border = thin_border
    materiales   = db_query("SELECT DISTINCT UPPER(material) FROM movimientos ORDER BY material")
    alertas_conf = {r[0].upper(): r[1] for r in db_query("SELECT material, stock_minimo FROM alertas_stock")}
    for (mat,) in materiales:
        from services.inventario import obtener_stock_material
        fijo, disponible = obtener_stock_material(mat)
        minimo = alertas_conf.get(mat, 5)
        if disponible <= 0:        estado = "SIN STOCK"
        elif disponible <= minimo: estado = "STOCK BAJO"
        else:                      estado = "OK"
        ws2.append([mat, fijo, disponible, estado])
        row_idx  = ws2.max_row
        fill_row = alerta_fill if "BAJO" in estado or "SIN" in estado else PatternFill("solid", fgColor="D4EDDA")
        for col_idx in range(1, 5):
            ws2.cell(row=row_idx, column=col_idx).fill = fill_row
            ws2.cell(row=row_idx, column=col_idx).border = thin_border
    for col_idx, width in enumerate([25,20,22,18], 1):
        ws2.column_dimensions[chr(64+col_idx)].width = width

    ws3 = wb.create_sheet("Limpieza")
    ws3.append(["Material","Cantidad","Estado","Fecha","Notas"])
    for col_idx in range(1, 6):
        cell = ws3.cell(row=1, column=col_idx)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center_align; cell.border = thin_border
    for row_data in db_query("SELECT material,cantidad,estado,fecha,notas FROM limpieza ORDER BY id DESC"):
        ws3.append(tuple(row_data))
        row_idx = ws3.max_row
        f = PatternFill("solid", fgColor="D4EDDA") if row_data[2] == "LIMPIO" else PatternFill("solid", fgColor="FFF3CD")
        for col_idx in range(1, 6):
            ws3.cell(row=row_idx, column=col_idx).fill = f
            ws3.cell(row=row_idx, column=col_idx).border = thin_border

    ws4 = wb.create_sheet("Estadisticas")
    stats = obtener_reporte_estadisticas()
    ws4["A1"] = "RESUMEN EJECUTIVO"
    ws4["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws4["A3"] = "Indicador"; ws4["B3"] = "Valor"
    for cell in [ws4["A3"], ws4["B3"]]:
        cell.font = header_font; cell.fill = header_fill
    filas = [
        ("Total de Movimientos",       stats["total_movimientos"]),
        ("Entradas Registradas",        stats["entradas"]),
        ("Salidas Registradas",         stats["salidas"]),
        ("Materiales Distintos",        stats["materiales"]),
        ("Responsables Activos",        stats["responsables"]),
        ("Material Mas Movido",         stats["material_top"]),
        ("Movimientos (ultimos 7 dias)",stats["movs_7dias"]),
        ("Reporte Generado",            datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    for i, (indicador, valor) in enumerate(filas, start=4):
        ws4.cell(row=i, column=1, value=indicador).border = thin_border
        ws4.cell(row=i, column=2, value=valor).border     = thin_border
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 20
    wb.save(ruta)
    return ruta


