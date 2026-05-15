import os
import threading
from datetime import datetime
from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
from kivy.uix.textinput import TextInput
from kivy.clock import Clock
from kivy.metrics import dp
from kivy.utils import get_color_from_hex
from kivy.core.window import Window
from database.connection import db_query, db_execute
from services.inventario import obtener_stock_material
from services.auditoria import log_auditoria
from widgets.tarjeta import crear_tarjeta
from config import EXPORT_DIR, DB_PATH
from logging_config import logger
from kivy.app import App

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class ExploradorArchivos(Popup):
    EXTENSIONES_EXCEL = ('.xlsx', '.xls', '.xlsm', '.xlsb')

    def __init__(self, callback, **kwargs):
        super().__init__(**kwargs)
        self.callback    = callback
        self.title       = "Seleccionar archivo Excel"
        self.size_hint   = (0.97, 0.93)
        self.ruta_actual = self._ruta_inicial()

        root_layout = BoxLayout(orientation='vertical', spacing=dp(6), padding=dp(8))

        accesos = [("Mis Reportes (app)", EXPORT_DIR)]
        try:
            from kivy.utils import platform
            if platform == 'android':
                accesos += [
                    ("Descargas",    "/storage/emulated/0/Download"),
                    ("Documentos",   "/storage/emulated/0/Documents"),
                    ("WhatsApp",     "/storage/emulated/0/Android/media/com.whatsapp/WhatsApp/Media/WhatsApp Documents"),
                    ("Telegram",     "/storage/emulated/0/Android/media/org.telegram.messenger/Telegram/Telegram Documents"),
                    ("Raiz",         "/storage/emulated/0"),
                ]
            else:
                import pathlib
                accesos += [
                    ("Descargas",    os.path.join(str(pathlib.Path.home()), "Downloads")),
                    ("Inicio",       str(pathlib.Path.home())),
                ]
        except Exception:
            import pathlib
            accesos += [
                ("Descargas", os.path.join(str(pathlib.Path.home()), "Downloads")),
                ("Inicio",    str(pathlib.Path.home())),
            ]

        atajos_scroll = ScrollView(size_hint_y=None, height=dp(48), do_scroll_y=False)
        atajos_box    = BoxLayout(size_hint_x=None, spacing=dp(6))
        atajos_box.bind(minimum_width=atajos_box.setter('width'))
        for nombre, ruta in accesos:
            if os.path.exists(ruta):
                btn = Button(
                    text=nombre, size_hint=(None, 1), width=dp(150),
                    font_size='12sp', background_color=get_color_from_hex('#1E40AF')
                )
                btn.bind(on_release=lambda x, r=ruta: self.listar(r))
                atajos_box.add_widget(btn)
        atajos_scroll.add_widget(atajos_box)
        root_layout.add_widget(atajos_scroll)

        self.lbl_ruta = Label(
            text=self.ruta_actual, size_hint_y=None, height=dp(30),
            font_size='11sp', color=get_color_from_hex('#94A3B8'),
            halign='left', shorten=True, shorten_from='left',
            text_size=(Window.width * 0.92, None)
        )
        root_layout.add_widget(self.lbl_ruta)

        self.filtro = TextInput(
            hint_text='Filtrar archivos...', multiline=False,
            size_hint_y=None, height=dp(42)
        )
        self.filtro.bind(text=lambda inst, val: self.listar(self.ruta_actual))
        root_layout.add_widget(self.filtro)

        self.lbl_debug = Label(
            text='', size_hint_y=None, height=dp(24), font_size='11sp',
            color=get_color_from_hex('#F59E0B'), halign='left',
            text_size=(Window.width * 0.92, None)
        )
        root_layout.add_widget(self.lbl_debug)

        self.scroll = ScrollView()
        self.lista  = GridLayout(cols=1, size_hint_y=None, spacing=dp(5))
        self.lista.bind(minimum_height=self.lista.setter('height'))
        self.scroll.add_widget(self.lista)
        root_layout.add_widget(self.scroll)

        btns = BoxLayout(size_hint_y=None, height=dp(46), spacing=dp(8))
        btn_back = Button(text='SUBIR NIVEL', background_color=get_color_from_hex('#334155'), font_size='13sp')
        btn_back.bind(on_release=lambda x: self.subir_nivel())
        btns.add_widget(btn_back)
        btn_cerrar = Button(text='CANCELAR', background_color=get_color_from_hex('#7F1D1D'), font_size='13sp')
        btn_cerrar.bind(on_release=self.dismiss)
        btns.add_widget(btn_cerrar)
        root_layout.add_widget(btns)
        self.content = root_layout
        self.listar(self.ruta_actual)

    def _ruta_inicial(self):
        if os.path.exists(EXPORT_DIR):
            return EXPORT_DIR
        try:
            from kivy.utils import platform
            if platform == 'android':
                for r in ['/storage/emulated/0/Download', '/storage/emulated/0']:
                    if os.path.exists(r):
                        return r
        except Exception:
            pass
        import pathlib
        descargas = os.path.join(str(pathlib.Path.home()), 'Downloads')
        return descargas if os.path.exists(descargas) else str(pathlib.Path.home())

    def listar(self, ruta):
        self.lista.clear_widgets()
        self.lbl_ruta.text = ruta
        self.ruta_actual   = ruta
        filtro_txt = self.filtro.text.lower().strip() if hasattr(self, 'filtro') else ''
        try:
            entradas = sorted(os.listdir(ruta))
        except PermissionError:
            self.lista.add_widget(Label(
                text='Sin permiso para acceder a esta carpeta',
                color=get_color_from_hex('#F87171'), size_hint_y=None, height=dp(60), halign='center'
            ))
            return
        except Exception as e:
            self.lista.add_widget(Label(text=f'Error: {e}', color=get_color_from_hex('#F87171'), size_hint_y=None, height=dp(50)))
            return

        carpetas = []; archivos = []; total_enc = 0
        for nombre in entradas:
            ns = nombre.strip()
            if ns.startswith('.'):
                continue
            rc = os.path.join(ruta, ns)
            if filtro_txt and filtro_txt not in ns.lower():
                continue
            try:
                es_dir = os.path.isdir(rc)
            except Exception:
                continue
            if es_dir:
                carpetas.append((ns, rc))
            elif ns.lower().endswith(self.EXTENSIONES_EXCEL):
                archivos.append((ns, rc))
                total_enc += 1

        total_en_carpeta = len([n for n in entradas if not n.startswith('.')])
        if hasattr(self, 'lbl_debug'):
            self.lbl_debug.text = f"{total_en_carpeta} entradas  |  {len(carpetas)} carpetas  |  {total_enc} Excel encontrado(s)"

        if not carpetas and not archivos:
            self.lista.add_widget(Label(
                text='No hay archivos Excel en esta carpeta',
                color=get_color_from_hex('#94A3B8'), size_hint_y=None, height=dp(50), halign='center'
            ))
            return

        for ns, rc in carpetas:
            btn = Button(text=f'DIR  {ns}', size_hint_y=None, height=dp(52), halign='left',
                         font_size='13sp', background_color=get_color_from_hex('#1E3A5F'))
            btn.bind(on_release=lambda x, r=rc: self.listar(r))
            self.lista.add_widget(btn)

        for ns, rc in archivos:
            try:
                size_txt = f'{os.path.getsize(rc)/1024:.1f} KB'
            except Exception:
                size_txt = ''
            btn = Button(
                text=f'XLS  {ns}\n[size=11sp][color=#94A3B8]{size_txt}[/color][/size]',
                markup=True, size_hint_y=None, height=dp(60), halign='left',
                font_size='13sp', background_color=get_color_from_hex('#065F46')
            )
            btn.bind(on_release=lambda x, r=rc: self._seleccionar(r))
            self.lista.add_widget(btn)

    def subir_nivel(self):
        padre = os.path.dirname(self.ruta_actual)
        if padre and padre != self.ruta_actual:
            self.listar(padre)

    def _seleccionar(self, ruta):
        self.dismiss()
        Clock.schedule_once(lambda dt: self.callback(ruta), 0.15)


class ImportarExcelScreen(Screen):
    _columnas_excel = []

    def on_enter(self):
        self._columnas_excel = []
        self.ids.preview_container.clear_widgets()
        self.ids.btn_importar.disabled = True
        self.ids.estado_importar.text  = "Pulsa EXPLORAR para buscar tu archivo Excel"
        try:
            self.ids.lbl_hint_carpeta.text = f"Reportes de la app en: {EXPORT_DIR}"
        except Exception:
            pass

    def abrir_explorador(self) -> None:
        try:
            explorador = ExploradorArchivos(callback=self._al_seleccionar_archivo)
            explorador.open()
        except Exception as e:
            self.ids.estado_importar.text = f"[color=#F87171]Error al abrir explorador: {e}[/color]"

    def _al_seleccionar_archivo(self, ruta):
        self.ids.ruta_excel.text      = ruta
        self.ids.estado_importar.text = f"[color=#10B981]Archivo seleccionado[/color]"
        Clock.schedule_once(lambda dt: self.previsualizar(), 0.1)

    def pegar_ruta(self) -> None:
        try:
            from kivy.core.clipboard import Clipboard
            texto = Clipboard.paste()
            if texto:
                texto = texto.strip().strip('\n').strip('\r')
                texto = texto.strip('"').strip("'").strip('\u201c').strip('\u201d')
                self.ids.ruta_excel.text = texto
        except Exception:
            self.ids.estado_importar.text = "[color=#F87171]No se pudo acceder al portapapeles[/color]"

    def previsualizar(self) -> None:
        ruta = self._normalizar_ruta(self.ids.ruta_excel.text)
        if not ruta:
            self.ids.estado_importar.text = "[color=#F87171]Primero selecciona un archivo[/color]"
            return
        if not os.path.exists(ruta):
            self.ids.estado_importar.text = (
                f"[color=#F87171]Archivo no encontrado: {ruta}[/color]"
            )
            return
        self.ids.ruta_excel.text       = ruta
        self.ids.estado_importar.text  = "Leyendo archivo..."
        self.ids.btn_importar.disabled = True
        self.ids.preview_container.clear_widgets()
        threading.Thread(target=self._leer_excel_seguro, args=(ruta,), daemon=True).start()

    def _normalizar_ruta(self, ruta_raw) -> str:
        if not ruta_raw:
            return ""
        ruta = ruta_raw.strip()
        for ch in ('"', "'", '\u201c', '\u201d', '\u2018', '\u2019'):
            ruta = ruta.strip(ch)
        ruta = ruta.strip('\n').strip('\r').strip()
        ruta = os.path.expanduser(ruta)
        ruta = os.path.normpath(ruta)
        return ruta

    def _leer_excel_seguro(self, ruta):
        try:
            if load_workbook is None:
                Clock.schedule_once(lambda dt: self._set_estado("[color=#F87171]openpyxl no instalado[/color]"))
                return
            wb    = load_workbook(ruta, read_only=True, data_only=True)
            hojas = wb.sheetnames
            Clock.schedule_once(lambda dt: self._set_hojas(hojas))
            hoja_sel    = self.ids.hoja_spinner.text
            hoja_nombre = hoja_sel if hoja_sel in hojas else hojas[0]
            ws          = wb[hoja_nombre]
            filas       = list(ws.iter_rows(values_only=True))
            wb.close()
            if len(filas) < 2:
                Clock.schedule_once(lambda dt: self._set_estado("[color=#F87171]El archivo no tiene datos[/color]"))
                return
            encabezados = [str(h).strip() if h is not None else "" for h in filas[0]]
            datos       = filas[1:]
            Clock.schedule_once(lambda dt: self._mostrar_preview(encabezados, datos, hoja_nombre))
        except MemoryError:
            Clock.schedule_once(lambda dt: self._set_estado("[color=#F87171]Archivo demasiado grande[/color]"))
        except Exception as e:
            logger.error("Error leyendo Excel: %s", e)
            Clock.schedule_once(lambda dt: self._set_estado(f"[color=#F87171]Error al leer: {e}[/color]"))

    def _set_hojas(self, hojas):
        self.ids.hoja_spinner.values = hojas
        if self.ids.hoja_spinner.text == "-- detectar --" and hojas:
            self.ids.hoja_spinner.text = hojas[0]

    def _set_estado(self, texto):
        self.ids.estado_importar.text = texto

    def _mostrar_preview(self, encabezados, datos, hoja_nombre):
        self._columnas_excel = encabezados
        self.ids.preview_container.clear_widgets()
        mapeo_auto = {
            'col_nombre':   ['nombre','responsable','name','usuario','resp'],
            'col_material': ['material','producto','item','descripcion','description','mat'],
            'col_sku':      ['sku','id','codigo','code','ref'],
            'col_cantidad': ['cantidad','qty','quantity','cant'],
            'col_tipo':     ['tipo','type','movimiento','movement'],
            'col_fecha':    ['fecha','date','datetime','timestamp'],
        }
        enc_lower = [e.lower() for e in encabezados]
        for campo, alternativas in mapeo_auto.items():
            for alt in alternativas:
                if alt in enc_lower:
                    self.ids[campo].text = encabezados[enc_lower.index(alt)]
                    break

        card_enc = crear_tarjeta(dp(75), '#0F3460')
        card_enc.add_widget(Label(
            text=f"[b]Hoja:[/b] {hoja_nombre}  |  [b]Columnas:[/b] {len(encabezados)}",
            markup=True, color=get_color_from_hex('#38BDF8'), font_size='13sp',
            size_hint_y=None, height=dp(24),
            halign='left', valign='middle',
            text_size=(Window.width * 0.85, None)
        ))
        card_enc.add_widget(Label(
            text="  ".join(encabezados[:5]),
            font_size='11sp', color=get_color_from_hex('#94A3B8'),
            size_hint_y=None, height=dp(22),
            halign='left', valign='middle',
            text_size=(Window.width * 0.85, None),
            shorten=True, shorten_from='right'
        ))
        self.ids.preview_container.add_widget(card_enc)

        total = len(datos)
        for fila in datos[:10]:
            vals = [str(v) if v is not None else "-" for v in fila]
            card = crear_tarjeta(dp(50))
            card.add_widget(Label(
                text="  |  ".join(vals[:5]),
                font_size='11sp', color=get_color_from_hex('#CBD5E1'),
                size_hint_y=None, height=dp(28),
                halign='left', valign='middle',
                text_size=(Window.width * 0.85, None),
                shorten=True, shorten_from='right'
            ))
            self.ids.preview_container.add_widget(card)

        resumen = f"[color=#10B981]{total} fila(s) listas - revisa el mapeo y pulsa IMPORTAR[/color]"
        if total > 10:
            resumen += f"  [color=#94A3B8](mostrando 10 de {total})[/color]"
        self._set_estado(resumen)
        self.ids.btn_importar.disabled = False

    def importar(self) -> None:
        ruta = self._normalizar_ruta(self.ids.ruta_excel.text)
        if not ruta or not os.path.exists(ruta):
            self.ids.estado_importar.text = f"[color=#F87171]Archivo no encontrado.[/color]"
            return
        self.ids.btn_importar.disabled = True
        self.ids.estado_importar.text  = "Importando datos..."
        threading.Thread(target=self._importar_worker, args=(ruta,), daemon=True).start()

    def _importar_worker(self, ruta):
        try:
            if load_workbook is None:
                Clock.schedule_once(lambda dt: self._set_estado("[color=#F87171]openpyxl no instalado[/color]"))
                Clock.schedule_once(lambda dt: setattr(self.ids.btn_importar, 'disabled', False))
                return
            wb          = load_workbook(ruta, read_only=True, data_only=True)
            hoja_nombre = self.ids.hoja_spinner.text
            if hoja_nombre not in wb.sheetnames:
                hoja_nombre = wb.sheetnames[0]
            ws    = wb[hoja_nombre]
            filas = list(ws.iter_rows(values_only=True))
            wb.close()
            if len(filas) < 2:
                Clock.schedule_once(lambda dt: self._set_estado("[color=#F87171]No hay datos[/color]"))
                Clock.schedule_once(lambda dt: setattr(self.ids.btn_importar, 'disabled', False))
                return

            enc_upper = [str(h).strip().upper() if h is not None else "" for h in filas[0]]

            def idx(campo_id):
                nombre_col = self.ids[campo_id].text.strip().upper()
                try:
                    return enc_upper.index(nombre_col)
                except ValueError:
                    return None

            i_nombre   = idx('col_nombre')
            i_material = idx('col_material')
            i_sku      = idx('col_sku')
            i_cantidad = idx('col_cantidad')
            i_tipo     = idx('col_tipo')
            i_fecha    = idx('col_fecha')

            if i_material is None or i_cantidad is None:
                Clock.schedule_once(lambda dt: self._set_estado("[color=#F87171]Columnas 'Material' y 'Cantidad' requeridas[/color]"))
                Clock.schedule_once(lambda dt: setattr(self.ids.btn_importar, 'disabled', False))
                return

            insertados = errores = omitidos = 0
            ahora = datetime.now().strftime("%d/%m/%Y %H:%M")

            for fila in filas[1:]:
                try:
                    def celda(i):
                        if i is None or i >= len(fila):
                            return None
                        return fila[i]

                    nombre   = str(celda(i_nombre)   or "INVENTARIO").strip().upper()
                    material = str(celda(i_material) or "").strip().upper()
                    sku      = str(celda(i_sku)      or "").strip()
                    raw_cant = celda(i_cantidad)
                    raw_tipo = str(celda(i_tipo)     or "ENTRADA").strip().upper()
                    raw_fec  = celda(i_fecha)

                    if not material:
                        omitidos += 1; continue
                    try:
                        cantidad = float(str(raw_cant).replace(',', '.'))
                    except (TypeError, ValueError):
                        omitidos += 1; continue
                    if cantidad <= 0:
                        omitidos += 1; continue

                    if any(k in raw_tipo for k in ['ENT','IN','ING','COMP']):
                        tipo = "ENTRADA"
                    elif any(k in raw_tipo for k in ['SAL','OUT','EGR','RET']):
                        tipo = "SALIDA"
                    else:
                        tipo = "ENTRADA"

                    if raw_fec is None:
                        fecha = ahora
                    elif isinstance(raw_fec, datetime):
                        fecha = raw_fec.strftime("%d/%m/%Y %H:%M")
                    else:
                        fecha = str(raw_fec).strip() or ahora

                    _, stock_actual = obtener_stock_material(material)
                    nuevo_stock = round(stock_actual + (cantidad if tipo == "ENTRADA" else -cantidad), 4)
                    db_execute(
                        "INSERT INTO movimientos (nombre,material,sku,cantidad,tipo,fecha,stock_registro,dias,retorno,notas,ubicacion) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                        (nombre, material, sku, cantidad, tipo, fecha, nuevo_stock, 0, "N/A", "", "")
                    )
                    insertados += 1
                except Exception:
                    errores += 1

            log_auditoria("IMPORTAR_EXCEL", f"{insertados} importadas, {errores} errores, {omitidos} omitidas")
            resumen = (
                f"[color=#10B981]Importacion completada[/color]\n"
                f"Insertados: [b]{insertados}[/b]  |  "
                f"Omitidos: {omitidos}  |  Errores: {errores}"
            )
            Clock.schedule_once(lambda dt: self._finalizar(resumen))
        except MemoryError:
            Clock.schedule_once(lambda dt: self._set_estado("[color=#F87171]Memoria insuficiente[/color]"))
            Clock.schedule_once(lambda dt: setattr(self.ids.btn_importar, 'disabled', False))
        except Exception as e:
            logger.error("Error importando: %s", e)
            Clock.schedule_once(lambda dt: self._set_estado(f"[color=#F87171]Error: {e}[/color]"))
            Clock.schedule_once(lambda dt: setattr(self.ids.btn_importar, 'disabled', False))

    def _finalizar(self, resumen):
        self._set_estado(resumen)
        self.ids.btn_importar.disabled = False
        self.ids.preview_container.clear_widgets()
        try:
            inv = App.get_running_app().root.get_screen('inventario')
            inv.actualizar_spinners()
            inv._actualizar_dashboard()
            inv._cargar_actividad_reciente()
            inv._verificar_alertas_silencioso()
        except Exception:
            pass
